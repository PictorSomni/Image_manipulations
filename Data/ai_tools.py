# -*- coding: utf-8 -*-
"""
Utilitaires IA partagés entre Dashboard.pyw et SidePanel.pyw.

Fonctions et constantes exposées :
  _fetch_url_content(url, max_chars)         — récupère le texte d'une URL HTTP(S)
  _web_search(query, max_results)            — recherche DuckDuckGo, retourne les résultats formatés
  _ollama_chat_once(url, model, messages)    — appel non-streaming à /api/chat, retourne le dict message
  _ollama_chat_stream(url, model, messages)  — appel streaming à /api/chat, génère les tokens

Outils dossier (partagés) :
  _folder_tool_definitions(folder_path)      — retourne les 7 définitions d'outils dossier pour Ollama/Gemma
  _gemini_tool_definitions(folder_path)       — version allégée pour Gemini (sans analyze_images)
  _folder_list_contents(folder_path)         — liste les fichiers avec taille et date
  _folder_read_file(folder_path, filename)   — lit le contenu texte d'un fichier
  _folder_create_file(folder_path, filename, content)
                                             — crée un fichier texte dans le dossier ouvert
  _encode_image_for_analysis(image_path)     — encode une image en base64 JPEG pour un modèle vision
  _analyze_images_batched(...)               — analyse des images par lots et retourne les résultats
  _gemini_generate_image(prompt, ...)        — génère/modifie une image via Nano Banana 2
  _gemini_interactions_create(input, ...)   — Interactions API : appel simple avec previous_interaction_id
  _gemini_interactions_get(interaction_id)  — Interactions API : polling pour background=True (agents)
  _gemini_generate_music(prompt, model, images)
                                             — génère de la musique via Lyria 3 (retourne bytes MP3/WAV)

Helpers système (partagés) :
  _WEB_TOOLS                                 — liste des 2 définitions d'outils web (web_search + fetch_url)
  _TERMINAL_TOOLS                            — définition de l'outil run_terminal_command
  _MEMORY_TOOLS                              — définition de l'outil update_memory_file
  _run_terminal_command(command, cwd, timeout, admin)
                                             — exécute une commande shell et retourne la sortie
                                               (admin=True → élévation via _run_elevated)
  _run_elevated(command, cwd, timeout)       — exécute une commande avec élévation de
                                               privilèges (UAC / osascript / pkexec selon l'OS)
  _update_memory_file(target, action, content, old_text)
                                             — met à jour memory.md / user.md / skills.md (retourne JSON)
  _read_memory_file(target)                  — lit un fichier mémoire brut (str)
  _build_system_content(base_prompt, folder_path, today_date_str)
                                             — assemble le message système complet (system.md + mémoire + date + dossier)
  _get_gemini_cached_content(client, model, system_instr, raw_tools)
                                             — réutilise/crée un cache de contexte Gemini (cachedContent)
  _compact_history_summary(ai_conversation, history_limit, state)
                                             — résume (Gemini, repli Gemma) les tours qui sortent de la fenêtre d'historique, pour tous les modèles

Voix — TTS :
  _gemini_tts(text, ...)                     — génère l'audio TTS en une seule requête (bytes PCM)
  _gemini_tts_stream(text, ...)              — TTS multi-requêtes pipeline (compatible tous modèles)
  _gemini_live_tts_stream(text, ...)         — TTS via Gemini Live WebSocket (voix cohérente, modèle Live)
  _voice_play_audio(pcm_bytes, ...)          — joue des bytes PCM via sounddevice

Dictée — STT (push-to-talk) :
  _MicRecorder(sample_rate)                  — enregistre le micro (start/stop → WAV bytes)
  _gemini_transcribe_audio(wav_bytes, ...)   — transcrit un WAV via Gemini, retourne le texte
"""

import ast as _ast
import gzip
import hashlib as _hashlib
import json
import logging
import os as _os
import re
import sys as _sys
import time as _time
import urllib.request
import urllib.parse
import html.parser

import CONSTANTS

# Erreurs "best-effort" (fallback silencieux volontaire) — voir chaque site
# d'appel pour le contexte. Un sous-ensemble seulement est loggé ici : ceux
# qui peuvent masquer un vrai bug (ex. arguments d'un tool call perdus),
# pas les dizaines de fallbacks cosmétiques (DPI, audio, dates...).
_logger = logging.getLogger("ai_tools")
if not _logger.handlers:
    _log_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), ".ai_tools_errors.log")
    _handler = logging.FileHandler(_log_path, encoding="utf-8")
    _handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    _logger.addHandler(_handler)
    _logger.setLevel(logging.WARNING)

# Rend le processus DPI-aware sous Windows : sans ça, si l'affichage a une
# mise à l'échelle > 100%, Windows "virtualise" les coordonnées de la souris
# (SetCursorPos/click) sur une résolution logique réduite alors que les
# captures d'écran (pyautogui.screenshot) restent en pixels physiques —
# décalage systématique entre ce que le modèle voit et où le clic atterrit.
if _sys.platform == "win32":
    try:
        import ctypes as _ctypes_dpi
        _ctypes_dpi.windll.shcore.SetProcessDpiAwareness(2)  # PROCESS_PER_MONITOR_DPI_AWARE
    except Exception:
        try:
            _ctypes_dpi.windll.user32.SetProcessDPIAware()  # fallback Windows 7/8
        except Exception:
            pass


_GEMINI_API_KEY_CACHE = None


class _HTMLTextExtractor(html.parser.HTMLParser):
    """Extrait le texte brut d'un document HTML de façon propre et structurée."""
    _SKIP_TAGS = {
        "script", "style", "noscript", "head",
        "nav", "header", "footer", "aside",
        "form", "button", "select", "option", "input", "textarea",
        "iframe", "figure", "figcaption", "picture",
        "dialog", "menu", "menuitem",
    }
    _CONTENT_TAGS = {"main", "article", "section"}
    
    # Balises induisant un saut de ligne sémantique
    _BLOCK_TAGS = {
        "p", "div", "h1", "h2", "h3", "h4", "h5", "h6", 
        "li", "tr", "td", "blockquote", "pre", "code", "br"
    }

    def __init__(self):
        super().__init__()
        self._skip_depth   = 0
        self._content_depth = 0
        self._parts        = []
        self._content_parts = []

    def handle_starttag(self, tag, attrs):
        if tag in self._SKIP_TAGS:
            self._skip_depth += 1
        if tag in self._CONTENT_TAGS:
            self._content_depth += 1
        if tag in self._BLOCK_TAGS:
            self._parts.append("\n")
            if self._content_depth > 0:
                self._content_parts.append("\n")

    def handle_endtag(self, tag):
        if tag in self._SKIP_TAGS and self._skip_depth > 0:
            self._skip_depth -= 1
        if tag in self._CONTENT_TAGS and self._content_depth > 0:
            self._content_depth -= 1
        if tag in self._BLOCK_TAGS:
            self._parts.append("\n")
            if self._content_depth > 0:
                self._content_parts.append("\n")

    def handle_data(self, data):
        if self._skip_depth > 0:
            return
        if not data.strip():
            return
        self._parts.append(data)
        if self._content_depth > 0:
            self._content_parts.append(data)

    def get_text(self):
        target = self._content_parts if len(self._content_parts) > 20 else self._parts
        raw_text = "".join(target)
        # Normalisation des espaces et tabulations
        text = re.sub(r"[ \t]+", " ", raw_text)
        # Normalisation des sauts de lignes successifs
        text = re.sub(r"\n\s*\n\s*\n+", "\n\n", text)
        lines = [line.strip() for line in text.split("\n")]
        return "\n".join(line for line in lines if line).strip()


def _get_gemini_api_key():
    """Tente de récupérer la clé d'API Gemini de façon extrêmement robuste sur Windows, macOS et Linux."""
    global _GEMINI_API_KEY_CACHE
    if _GEMINI_API_KEY_CACHE is not None:
        return _GEMINI_API_KEY_CACHE

    import os
    import re
    import subprocess
    
    key = ""
    # 1. Directement dans l'environnement
    key = os.environ.get("GEMINI_API_KEY", "").strip()
    if key:
        _GEMINI_API_KEY_CACHE = key
        return key
        
    # 2. Dans un fichier .env (relatif au dossier de l'app ou racine du dépôt)
    for env_dir in [os.getcwd(), os.path.dirname(os.path.abspath(__file__)), os.path.dirname(os.path.dirname(os.path.abspath(__file__)))]:
        env_path = os.path.join(env_dir, ".env")
        if os.path.exists(env_path):
            try:
                with open(env_path, "r", encoding="utf-8") as f:
                    for line in f:
                        if line.strip() and not line.startswith("#") and "=" in line:
                            k, v = line.split("=", 1)
                            if k.strip() == "GEMINI_API_KEY":
                                key = v.strip().strip('"').strip("'")
                                if key:
                                    os.environ["GEMINI_API_KEY"] = key
                                    _GEMINI_API_KEY_CACHE = key
                                    return key
            except Exception:
                pass

    # 3. macOS / Linux uniquement : charger depuis le login shell (pour .zshrc, .bashrc, .profile)
    if os.name != "nt":
        for shell in ["/bin/zsh", "/bin/bash"]:
            if os.path.exists(shell):
                try:
                    # "-li" (login + interactif), pas seulement "-l" : la
                    # plupart des .bashrc (Debian/Raspberry Pi OS...)
                    # commencent par "si non interactif, ne rien faire" et
                    # s'arrêtent avant d'atteindre l'export de la clé sinon.
                    result = subprocess.run(
                        [shell, "-li", "-c", "echo $GEMINI_API_KEY"],
                        capture_output=True, text=True, timeout=2
                    )
                    key = result.stdout.strip()
                    if key:
                        os.environ["GEMINI_API_KEY"] = key
                        _GEMINI_API_KEY_CACHE = key
                        return key
                except Exception:
                    pass

        # Fallback : lecture directe des fichiers RC de l'utilisateur
        home = os.path.expanduser("~")
        for rc_file in [".zshrc", ".bashrc", ".bash_profile", ".profile"]:
            rc_path = os.path.join(home, rc_file)
            if os.path.isfile(rc_path):
                try:
                    with open(rc_path, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read()
                    # match "export GEMINI_API_KEY=xxx" or "GEMINI_API_KEY=xxx"
                    # re.MULTILINE : sans lui, "$" ne matche que la fin du
                    # fichier entier (pas la fin de chaque ligne), donc la
                    # recherche échouait dès que la ligne n'était pas la
                    # toute dernière du fichier RC.
                    match = re.search(
                        r'(?:export\s+)?GEMINI_API_KEY\s*=\s*["\']?(.*?)["\']?\s*(?:#|$)',
                        content, re.MULTILINE)
                    if match:
                        key = match.group(1).strip()
                        if key:
                            os.environ["GEMINI_API_KEY"] = key
                            _GEMINI_API_KEY_CACHE = key
                            return key
                except Exception:
                    pass
                    
    return ""


def _fetch_url_content(url, max_chars=12_000):
    """
    Récupère le contenu textuel d'une URL HTTP(S).

    Retourne une chaîne tronquée à ``max_chars`` caractères,
    ou un message d'erreur si la récupération échoue.
    """
    try:
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (compatible; ImageManipBot/1.0)",
                "Accept-Encoding": "gzip, deflate",
            },
        )
        with urllib.request.urlopen(req, timeout=15) as response:
            content_type = response.headers.get("Content-Type", "")
            content_encoding = response.headers.get("Content-Encoding", "").lower()
            
            charset = "utf-8"
            if "charset=" in content_type.lower():
                try:
                    charset = content_type.lower().split("charset=")[-1].split(";")[0].strip()
                except Exception:
                    pass
                    
            raw_bytes = response.read()
            if "gzip" in content_encoding:
                try:
                    raw_bytes = gzip.decompress(raw_bytes)
                except Exception as gz_err:
                    return f"[Erreur de décompression Gzip : {gz_err}]"
                    
        try:
            raw_text = raw_bytes.decode(charset, errors="replace")
        except LookupError:
            raw_text = raw_bytes.decode("utf-8", errors="replace")
            
        if "</" in raw_text or "<br" in raw_text.lower():
            extractor = _HTMLTextExtractor()
            extractor.feed(raw_text)
            plain_text = extractor.get_text()
        else:
            plain_text = raw_text
            
        plain_text = re.sub(r"\n{3,}", "\n\n", plain_text).strip()
        if len(plain_text) > max_chars:
            plain_text = plain_text[:max_chars] + f"\n\n[… contenu tronqué à {max_chars} caractères]"
        return plain_text
    except Exception as fetch_error:
        return f"[Impossible de récupérer l'URL : {fetch_error}]"


def _clean_ddg_url(url):
    """Extrait la vraie URL de destination à partir d'un lien de redirection DuckDuckGo."""
    if not url:
        return ""
    if url.startswith("//"):
        url = "https:" + url
    if "uddg=" in url or "/l/?" in url:
        try:
            parsed = urllib.parse.urlparse(url)
            query_params = urllib.parse.parse_qs(parsed.query)
            if "uddg" in query_params:
                return urllib.parse.unquote(query_params["uddg"][0])
        except Exception:
            pass
    return url


class _DDGResultsParser(html.parser.HTMLParser):
    """Parse les résultats de recherche de la page HTML DuckDuckGo."""

    def __init__(self):
        super().__init__()
        self._results   = []   # [{"title": ..., "url": ..., "snippet": ...}]
        self._capturing = None  # "title" | "snippet"
        self._current   = None
        self._buf       = []

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        cls = a.get("class") or ""
        if tag == "a" and "result__a" in cls:
            raw_url = a.get("href", "")
            clean_url = _clean_ddg_url(raw_url)
            self._current   = {"title": "", "url": clean_url, "snippet": ""}
            self._capturing = "title"
            self._buf       = []
        elif tag == "a" and "result__snippet" in cls:
            self._capturing = "snippet"
            self._buf       = []

    def handle_endtag(self, tag):
        if tag != "a" or self._capturing is None:
            return
        text = " ".join(self._buf).strip()
        if self._capturing == "title" and self._current:
            self._current["title"] = text
            self._results.append(self._current)
            self._current = None
        elif self._capturing == "snippet" and self._results:
            self._results[-1]["snippet"] = text
        self._capturing = None

    def handle_data(self, data):
        if self._capturing:
            s = data.strip()
            if s:
                self._buf.append(s)

    def get_results(self):
        return self._results


def _ollama_chat_once(ollama_url, model, messages, tools=None, temperature=0.7, keep_alive=-1, timeout=120):
    """
    Envoie un appel non-streaming à /api/chat d'Ollama.
    Retourne le dict ``message`` de la réponse (clés "content" et éventuellement "tool_calls").
    Lève une exception en cas d'erreur réseau ou HTTP.
    """
    body = {
        "model": model,
        "messages": messages,
        "stream": False,
        "think": True,
        "keep_alive": keep_alive,
        "options": {"temperature": temperature, "num_ctx": 8192, "num_predict": -1},
    }
    if tools:
        body["tools"] = tools
    payload = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        f"{ollama_url}/api/chat",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        result = json.loads(resp.read().decode("utf-8"))
    return result.get("message", {})


def _ollama_chat_stream(ollama_url, model, messages, temperature=0.7, keep_alive=-1, timeout=300):
    """
    Envoie un appel streaming à /api/chat d'Ollama.
    Génère les tokens (str) au fur et à mesure de leur réception.
    Lève une exception en cas d'erreur réseau ou HTTP.
    """
    body = {
        "model": model,
        "messages": messages,
        "stream": True,
        "keep_alive": keep_alive,
        "options": {"temperature": temperature},
    }
    payload = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        f"{ollama_url}/api/chat",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        for raw_line in resp:
            try:
                chunk = json.loads(raw_line.decode("utf-8"))
            except json.JSONDecodeError:
                continue
            token = chunk.get("message", {}).get("content", "")
            if token:
                yield token
            if chunk.get("done"):
                break


def _ollama_chat_stream_with_tools(ollama_url, model, messages, tools=None, temperature=0.7, keep_alive=-1, timeout=300, think=False):
    """
    Streaming /api/chat avec thinking natif Ollama et capture des tool_calls.

    Génère des tuples :
      ("token",     str)  — token texte au fur et à mesure
      ("thinking",  str)  — token de réflexion (message.thinking, natif Ollama)
      ("tool_calls", list) — appels d'outils cumulés depuis tous les chunks
    """
    body = {
        "model": model,
        "messages": messages,
        "stream": True,
        "think": think,
        "keep_alive": keep_alive,
        "options": {
            "temperature": temperature,
            "num_ctx": 8192,    # prompt système + 19 outils ≈ 4300 tokens ; 2048 (défaut) est trop court
            "num_predict": -1,  # pas de limite sur la longueur de la réponse générée
        },
    }
    if tools:
        body["tools"] = tools
    payload = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        f"{ollama_url}/api/chat",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        for raw_line in resp:
            try:
                chunk = json.loads(raw_line.decode("utf-8"))
            except json.JSONDecodeError:
                continue
            msg = chunk.get("message", {})
            thinking = msg.get("thinking", "")
            if thinking:
                yield ("thinking", thinking)
            token = msg.get("content", "")
            if token:
                yield ("token", token)
            tool_calls = msg.get("tool_calls") or []
            if tool_calls:
                yield ("tool_calls", tool_calls)
            if chunk.get("done"):
                break


def _md_dark(text: str) -> str:
    """Remplace les blockquotes Markdown par un équivalent lisible sur thème sombre."""
    lines = text.split("\n")
    result = []
    for line in lines:
        if line.startswith("> "):
            result.append("**›** " + line[2:])
        elif line == ">":
            result.append("")
        else:
            result.append(line)
    return "\n".join(result)


def _ai_save_history(conversation, file_path, history_compaction=None):
    """
    Sauvegarde la conversation (role + content) dans un fichier JSON, avec
    l'état de compactage de l'historique (résumé cumulatif — voir
    _compact_history_summary) pour qu'il survive à un redémarrage de l'app
    ou à un changement de machine.
    """
    try:
        serializable = [
            {"role": msg["role"], "content": msg["content"]}
            for msg in conversation
            if msg.get("role") in ("user", "assistant")
        ]
        payload = {
            "messages": serializable,
            "history_compaction": history_compaction or {"summary": "", "summarized_count": 0},
        }
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _ensure_ollama_ready(model_name, add_bubble_fn, page, ollama_process):
    """
    Vérifie qu'Ollama est lancé et que le modèle est disponible.
    Lance le serveur et/ou télécharge le modèle si nécessaire.
    Retourne True si tout est prêt, False en cas d'erreur bloquante.
    Doit être appelé depuis un thread secondaire (bloquant).
    """
    import time as _time
    if (model_name or "").startswith(("gemini", "claude")):
        return True

    def _is_ollama_up():
        try:
            with urllib.request.urlopen(
                f"{CONSTANTS.AI_OLLAMA_URL}/api/tags", timeout=3
            ) as resp:
                return resp.status == 200
        except Exception:
            return False

    if not _is_ollama_up():
        add_bubble_fn("assistant", "⚙️ Démarrage d'Ollama en arrière-plan…")
        try:
            ollama_process["proc"] = _subprocess.Popen(
                ["ollama", "serve"],
                stdout=_subprocess.DEVNULL,
                stderr=_subprocess.DEVNULL,
            )
        except FileNotFoundError:
            add_bubble_fn(
                "assistant",
                "[ERREUR] Ollama n'est pas installé sur cette machine.\n"
                "Téléchargez-le sur https://ollama.com",
            )
            return False
        for _ in range(40):
            _time.sleep(0.5)
            if _is_ollama_up():
                break
        else:
            add_bubble_fn(
                "assistant",
                "[ERREUR] Ollama n'a pas démarré dans les délais impartis.",
            )
            return False

    try:
        with urllib.request.urlopen(
            f"{CONSTANTS.AI_OLLAMA_URL}/api/tags", timeout=5
        ) as resp:
            available_names = [
                m.get("name", "")
                for m in json.loads(resp.read().decode("utf-8")).get("models", [])
            ]
        model_present = any(
            name == model_name or name.startswith(model_name + ":")
            for name in available_names
        )
    except Exception:
        model_present = False

    if not model_present:
        pull_status_ctrl = add_bubble_fn(
            "assistant",
            f"⬇️ Téléchargement de {model_name}…\n"
            "(première utilisation — peut prendre quelques minutes)",
        )
        try:
            pull_payload = json.dumps(
                {"name": model_name, "stream": True}
            ).encode("utf-8")
            pull_req = urllib.request.Request(
                f"{CONSTANTS.AI_OLLAMA_URL}/api/pull",
                data=pull_payload,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(pull_req, timeout=3600) as pull_resp:
                for raw_line in pull_resp:
                    try:
                        chunk = json.loads(raw_line.decode("utf-8"))
                    except json.JSONDecodeError:
                        continue
                    status = chunk.get("status", "")
                    completed = chunk.get("completed", 0)
                    total = chunk.get("total", 0)
                    if total:
                        pct = int(completed / total * 100)
                        pull_status_ctrl.value = (
                            f"⬇️ {model_name} — {status} {pct}%"
                        )
                    elif status:
                        pull_status_ctrl.value = f"⬇️ {model_name} — {status}"
                    try:
                        page.update()
                    except Exception:
                        pass
            pull_status_ctrl.value = f"✅ {model_name} téléchargé et prêt !"
            try:
                page.update()
            except Exception:
                pass
        except Exception as exc:
            add_bubble_fn(
                "assistant", f"[ERREUR] Téléchargement du modèle : {exc}"
            )
            return False

    return True


def _eval_str_node(node):
    """Évalue un nœud AST en chaîne (littéral ou concaténation de littéraux)."""
    if isinstance(node, _ast.Constant) and isinstance(node.value, str):
        return node.value
    if isinstance(node, _ast.BinOp) and isinstance(node.op, _ast.Add):
        left = _eval_str_node(node.left)
        right = _eval_str_node(node.right)
        if left is not None and right is not None:
            return left + right
    return None


def _parse_text_tool_calls(text):
    """
    Détecte les appels d'outils écrits en texte brut par le modèle (format de repli).

    Supporte trois formats émis par Gemma :
      <execute_tool> fn(key='val') </execute_tool>
      <|tool_call>call:fn{key:<|"|>val<|"|>}<tool_call|>
      <tool_code>print(obj.fn(key="val"))</tool_code>  — style Google/Gemma

    Retourne une liste de dicts {"function": {"name": ..., "arguments": {...}}}
    compatibles avec le format tool_calls d'Ollama.
    """
    calls = []

    # Format 1 : <execute_tool> fn(key='value') </execute_tool>
    pat1 = re.compile(r'<execute_tool>\s*(\w+)\(([^)]*)\)\s*</execute_tool>',
                      re.IGNORECASE | re.DOTALL)
    for m in pat1.finditer(text):
        fn_name  = m.group(1)
        args_str = m.group(2)
        args = {}
        for kv in re.finditer(r'(\w+)\s*=\s*["\']([^"\']*)["\']', args_str):
            args[kv.group(1)] = kv.group(2)
        calls.append({"function": {"name": fn_name, "arguments": args}})

    # Format 2 (tokens natifs Gemma) : <|tool_call>call:fn{key:<|"|>val<|"|>}<tool_call|>
    pat2 = re.compile(r'<\|tool_call>call:(\w+)\{(.*?)\}<tool_call\|>', re.DOTALL)
    for m in pat2.finditer(text):
        fn_name  = m.group(1)
        args_str = m.group(2)
        args = {}
        for kv in re.finditer(r'(\w+):<\|"\|>(.*?)<\|"\|>', args_str, re.DOTALL):
            args[kv.group(1)] = kv.group(2)
        for kv in re.finditer(r'(\w+):([^,}]+)', args_str):
            key = kv.group(1)
            if key not in args:
                args[key] = kv.group(2).strip()
        calls.append({"function": {"name": fn_name, "arguments": args}})

    # Format 3 : <tool_code>print(obj.fn(key="val"))</tool_code> — style Google/Gemma
    _GEMMA_FN_ALIASES = {
        "write_file":       "create_file",
        "update_file":      "create_file",
        "save_file":        "create_file",
        "create_text_file": "create_file",
    }
    _GEMMA_PARAM_ALIASES = {
        "create_file": {
            "file_name":    "filename",
            "file_content": "content",
            "file_path":    "filename",
            "path":         "filename",
        },
        "read_file_content": {
            "file_name": "filename",
            "file_path": "filename",
            "path":      "filename",
        },
    }
    pat3 = re.compile(r'<tool_code>(.*?)</tool_code>', re.DOTALL | re.IGNORECASE)
    for m in pat3.finditer(text):
        code_block = m.group(1).strip()
        try:
            tree = _ast.parse(code_block, mode='eval')
            call_node = tree.body
            # Déballer print(...) si présent
            if (
                isinstance(call_node, _ast.Call)
                and isinstance(call_node.func, _ast.Name)
                and call_node.func.id == 'print'
                and call_node.args
            ):
                call_node = call_node.args[0]
            # obj.fn_name(...)
            if isinstance(call_node, _ast.Call) and isinstance(call_node.func, _ast.Attribute):
                fn_name = call_node.func.attr
                # Normaliser le nom de fonction (ex. write_file → create_file)
                fn_name = _GEMMA_FN_ALIASES.get(fn_name, fn_name)
                args = {}
                for kw in call_node.keywords:
                    if not kw.arg:
                        continue
                    # Essayer concaténation de chaînes d'abord, puis literal_eval
                    arg_val = _eval_str_node(kw.value)
                    if arg_val is None:
                        try:
                            arg_val = _ast.literal_eval(kw.value)
                        except Exception:
                            arg_val = None
                    if arg_val is not None:
                        args[kw.arg] = arg_val
                # Normaliser les noms de paramètres Gemma → noms de nos outils
                aliases = _GEMMA_PARAM_ALIASES.get(fn_name, {})
                for gemma_key, our_key in aliases.items():
                    if gemma_key in args:
                        args[our_key] = args.pop(gemma_key)
                if fn_name:  # args peut être vide (ex. list_folder_contents())
                    calls.append({"function": {"name": fn_name, "arguments": args}})
        except (SyntaxError, ValueError, AttributeError):
            # Fallback regex pour les blocs avec chaînes multi-lignes
            # (ast.parse échoue si le contenu contient de vraies newlines)
            _fb_match = re.search(
                r'(?:print\s*\(\s*)?(?:\w+\.)?(\w+)\s*\(', code_block
            )
            if _fb_match:
                _fb_fn = _fb_match.group(1)
                if _fb_fn == "print":
                    _inner_match = re.search(
                        r'print\s*\(\s*(?:\w+\.)?(\w+)\s*\(', code_block
                    )
                    _fb_fn = _inner_match.group(1) if _inner_match else ""
                # Normaliser le nom de fonction (ex. write_file → create_file)
                _fb_fn = _GEMMA_FN_ALIASES.get(_fb_fn, _fb_fn)
                _fb_args = {}
                for _kv in re.finditer(
                    r'(\w+)\s*=\s*"((?:[^"\\]|\\.)*)"', code_block, re.DOTALL
                ):
                    _fb_args[_kv.group(1)] = (
                        _kv.group(2)
                        .replace('\\n', '\n').replace('\\t', '\t')
                        .replace('\\\\', '\\').replace('\\"', '"')
                    )
                for _kv in re.finditer(
                    r"(\w+)\s*=\s*'((?:[^'\\]|\\.)*)'", code_block, re.DOTALL
                ):
                    if _kv.group(1) not in _fb_args:
                        _fb_args[_kv.group(1)] = (
                            _kv.group(2)
                            .replace('\\n', '\n').replace('\\t', '\t')
                            .replace('\\\\', '\\').replace("\\'", "'")
                        )
                _fb_aliases = _GEMMA_PARAM_ALIASES.get(_fb_fn, {})
                for _gk, _ok in _fb_aliases.items():
                    if _gk in _fb_args:
                        _fb_args[_ok] = _fb_args.pop(_gk)
                if _fb_fn:
                    calls.append({"function": {"name": _fb_fn, "arguments": _fb_args}})

    return calls


def _strip_text_tool_calls(text):
    """Supprime les blocs d'appels d'outils textuels du contenu."""
    text = re.sub(r'<execute_tool>.*?</execute_tool>', '', text,
                  flags=re.IGNORECASE | re.DOTALL)
    text = re.sub(r'<\|tool_call>.*?<tool_call\|>', '', text, flags=re.DOTALL)
    text = re.sub(r'<tool_code>.*?</tool_code>', '', text,
                  flags=re.IGNORECASE | re.DOTALL)
    return text.strip()


def _format_ai_conversation(conversation, user_name="Toi", separator_width=80):
    """
    Formate une liste de messages IA en texte Markdown exportable.

    Chaque tour est un bloc H1 avec le nom du rôle. Les blocs de réflexion
    (thinking) sont rendus en blockquote avant le contenu. Les blocs sont
    séparés par une ligne de tirets.

    Args:
        conversation : liste de dicts avec les clés 'role', 'content', 'thinking'.
        user_name    : nom affiché pour le rôle 'user'.

    Returns:
        Chaîne Markdown prête à être copiée ou insérée dans le bloc-notes.
    """
    blocks = []
    separator = "\n\n" + "#" * separator_width + "\n\n"
    for message in conversation:
        role = message.get("role", "")
        content = message.get("content", "")
        thinking = message.get("thinking", "")
        events = message.get("events", [])
        if role == "user":
            prefix = user_name
        elif role == "assistant":
            prefix = "IA"
        else:
            continue
        if thinking:
            thinking_lines = "\n".join(
                f"> {line}" if line.strip() else ">" for line in thinking.split("\n")
            )
            block = f"> 💭 **Réflexion**\n>\n{thinking_lines}\n\n# {prefix}\n\n"
        else:
            block = f"# {prefix}\n\n"
        if events:
            events_text = "\n".join(f"  • {e}" for e in events)
            block += f"**Actions :**\n{events_text}\n\n"
        block += content.strip()
        blocks.append(block.strip())
    return separator.join(blocks).strip()


def _web_search(query, max_results=5):
    """
    Effectue une recherche sur DuckDuckGo et retourne les résultats formatés.

    Retourne une chaîne listant titres, URLs et extraits,
    ou un message d'erreur si la recherche échoue.
    """
    try:
        encoded = urllib.parse.quote_plus(query)
        url = f"https://html.duckduckgo.com/html/?q={encoded}&kl=fr-fr"
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
            },
        )
        with urllib.request.urlopen(req, timeout=12) as resp:
            raw_html = resp.read().decode("utf-8", errors="replace")
        parser = _DDGResultsParser()
        parser.feed(raw_html)
        results = parser.get_results()[:max_results]
        if not results:
            return "[Aucun résultat trouvé]"
        lines = []
        for i, r in enumerate(results, 1):
            title   = r["title"]   or "(sans titre)"
            snippet = r["snippet"] or ""
            link    = r["url"]
            lines.append(f"{i}. **{title}**\n   {link}\n   {snippet}")
        return "\n\n".join(lines)
    except Exception as exc:
        return f"[Erreur de recherche web : {exc}]"


# ─── Outils dossier partagés ──────────────────────────────────────────────────
import os as _os
import datetime as _datetime_module

_FOLDER_DOCUMENT_EXTS_DEFAULT = frozenset({
    ".txt", ".md", ".py", ".js", ".ts", ".json", ".csv", ".xml",
    ".html", ".htm", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".log",
    ".rst", ".rtf",
})
_FOLDER_IMAGE_EXTS_DEFAULT = frozenset({
    ".jpg", ".jpeg", ".png", ".gif", ".bmp",
    ".webp", ".ico", ".tiff", ".tif",
})


def _folder_tool_definitions(folder_path):
    """
    Retourne les définitions d'outils pour l'API Ollama/Claude.
    Les outils fichiers (lecture, écriture, suppression, déplacement…) sont toujours disponibles
    car ils acceptent des chemins absolus. Les outils dossier (organize_files, analyze_images,
    edit_image) nécessitent un dossier ouvert valide.
    """
    folder_valid = bool(folder_path and _os.path.isdir(folder_path))

    tools = [
        # ── Outils fichiers (toujours disponibles, acceptent chemins absolus) ──
        {
            "type": "function",
            "function": {
                "name": "list_folder_contents",
                "description": (
                    "Liste les fichiers et sous-dossiers d'un dossier avec taille et date. "
                    "Si 'path' est omis, liste le dossier actuellement ouvert dans l'interface. "
                    "Accepte un chemin absolu pour lister n'importe quel dossier du système."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": (
                                "Chemin absolu du dossier à lister. "
                                "Laisser vide pour lister le dossier actuellement ouvert."
                            ),
                        }
                    },
                    "required": [],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_file_content",
                "description": (
                    "Lit le contenu d'un fichier texte. "
                    "Accepte un chemin absolu (ex. '/Users/charles/projet/script.py') "
                    "ou un nom de fichier relatif au dossier ouvert. "
                    "Extensions supportées : txt, md, py, js, json, csv, yaml, pyw, toml…"
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "filename": {
                            "type": "string",
                            "description": "Chemin absolu ou nom de fichier relatif au dossier ouvert",
                        }
                    },
                    "required": ["filename"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "create_file",
                "description": (
                    "Crée ou remplace un fichier texte avec le contenu fourni. "
                    "Accepte un chemin absolu (ex. '/Users/charles/projet/script.py') "
                    "ou un nom/chemin relatif au dossier ouvert (ex. 'sous-dossier/notes.md'). "
                    "Crée automatiquement les répertoires parents si nécessaire. "
                    "Fonctionne aussi pour modifier un fichier existant : lire avec read_file_content, "
                    "modifier en mémoire, puis réécrire avec create_file."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "filename": {
                            "type": "string",
                            "description": "Chemin absolu ou nom/chemin relatif au dossier ouvert",
                        },
                        "content": {
                            "type": "string",
                            "description": "Contenu textuel complet du fichier",
                        },
                    },
                    "required": ["filename", "content"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "delete_files",
                "description": (
                    "Supprime une liste de fichiers ou dossiers. "
                    "Accepte des chemins absolus ou relatifs au dossier ouvert. "
                    "Une confirmation est demandée à l'utilisateur avant la suppression "
                    "(désactivable via CONSTANTS.AI_DELETE_CONFIRM). "
                    "Les dossiers sont supprimés récursivement avec leur contenu."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "paths": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Liste des chemins à supprimer (absolus ou relatifs au dossier ouvert)",
                        },
                        "summary": {
                            "type": "string",
                            "description": "Résumé de ce qui sera supprimé et pourquoi",
                        },
                    },
                    "required": ["paths"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "move_file",
                "description": (
                    "Déplace ou renomme un fichier ou dossier. "
                    "Source et destination acceptent des chemins absolus ou relatifs au dossier ouvert. "
                    "Si la destination est un dossier existant, le fichier y est déplacé. "
                    "Sinon, le fichier est renommé/déplacé au chemin exact spécifié."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source": {
                            "type": "string",
                            "description": "Chemin source (absolu ou relatif au dossier ouvert)",
                        },
                        "destination": {
                            "type": "string",
                            "description": "Chemin destination (absolu ou relatif au dossier ouvert)",
                        },
                    },
                    "required": ["source", "destination"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "copy_file",
                "description": (
                    "Copie un fichier ou un dossier (copie récursive pour les dossiers). "
                    "Source et destination acceptent des chemins absolus ou relatifs au dossier ouvert."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source": {
                            "type": "string",
                            "description": "Chemin source (absolu ou relatif au dossier ouvert)",
                        },
                        "destination": {
                            "type": "string",
                            "description": "Chemin destination (absolu ou relatif au dossier ouvert)",
                        },
                    },
                    "required": ["source", "destination"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "create_folder",
                "description": (
                    "Crée un dossier et ses parents si nécessaire (équivalent de mkdir -p). "
                    "Accepte un chemin absolu ou relatif au dossier ouvert."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Chemin absolu ou relatif au dossier ouvert",
                        }
                    },
                    "required": ["path"],
                },
            },
        },
        # ── EXIF / ZIP (toujours disponibles, acceptent chemins absolus) ──────
        {
            "type": "function",
            "function": {
                "name": "read_exif",
                "description": (
                    "Lit les métadonnées EXIF d'une ou plusieurs images "
                    "(date de prise, appareil, objectif, réglages exposition, coordonnées GPS…). "
                    "Accepte des chemins absolus ou relatifs au dossier ouvert. "
                    "Utile pour trier par date réelle de prise de vue, identifier l'appareil, "
                    "ou récupérer une localisation GPS."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "filenames": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Liste de fichiers images (chemins absolus ou relatifs au dossier ouvert)",
                        }
                    },
                    "required": ["filenames"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "zip_files",
                "description": (
                    "Crée une archive ZIP à partir d'une liste de fichiers et/ou dossiers. "
                    "Accepte des chemins absolus ou relatifs au dossier ouvert. "
                    "L'archive est créée dans le dossier ouvert par défaut "
                    "(ou dans 'destination' si fourni)."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "paths": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Liste des fichiers/dossiers à archiver",
                        },
                        "zip_name": {
                            "type": "string",
                            "description": "Nom de l'archive (ex. 'photos_2024.zip'). Défaut : 'archive.zip'.",
                        },
                        "destination": {
                            "type": "string",
                            "description": "Dossier de destination (absolu ou relatif). Défaut : dossier ouvert.",
                        },
                    },
                    "required": ["paths"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "unzip_file",
                "description": (
                    "Extrait une archive ZIP. "
                    "Accepte un chemin absolu ou relatif au dossier ouvert. "
                    "Détecte automatiquement si l'archive contient un dossier racine unique "
                    "pour éviter les dossiers imbriqués inutiles."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source": {
                            "type": "string",
                            "description": "Chemin du fichier .zip (absolu ou relatif au dossier ouvert)",
                        },
                        "destination": {
                            "type": "string",
                            "description": "Dossier de destination (optionnel). Par défaut : même dossier que le .zip.",
                        },
                    },
                    "required": ["source"],
                },
            },
        },
        # ── Génération d'image (toujours disponible) ───────────────────────────
        {
            "type": "function",
            "function": {
                "name": "generate_image",
                "description": (
                    "Génère une image à partir d'un prompt texte avec Nano Banana 2 "
                    "(gemini-3.1-flash-image-preview). "
                    "L'image est sauvegardée dans le dossier ouvert et affichée dans le chat. "
                    "Utilise cet outil quand l'utilisateur demande de créer, dessiner ou générer une image."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt": {
                            "type": "string",
                            "description": (
                                "Description détaillée de l'image à générer. "
                                "Plus la description est précise, meilleur est le résultat."
                            ),
                        },
                        "filename": {
                            "type": "string",
                            "description": (
                                "Nom du fichier de sortie (ex. 'portrait.png'). "
                                "Laisser vide pour nommer automatiquement."
                            ),
                        },
                        "aspect_ratio": {
                            "type": "string",
                            "enum": ["1:1", "16:9", "9:16", "4:3", "3:4", "3:2", "2:3"],
                            "description": "Format de l'image. Défaut : 1:1.",
                        },
                    },
                    "required": ["prompt"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "generate_music",
                "description": (
                    "Génère de la musique via Lyria 3 (Google). "
                    "L'audio MP3 est sauvegardé dans le dossier ouvert ou Generated/. "
                    "Utilise cet outil quand l'utilisateur demande de créer ou générer "
                    "de la musique, une chanson, un morceau, une ambiance sonore, etc."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt": {
                            "type": "string",
                            "description": (
                                "Description détaillée de la musique à générer. "
                                "Inclure : genre, instruments, tempo (BPM), ambiance, "
                                "structure ([Intro] [Verse] [Chorus]…). "
                                "Plus le prompt est précis, meilleur est le résultat."
                            ),
                        },
                        "filename": {
                            "type": "string",
                            "description": (
                                "Nom du fichier de sortie (ex. 'theme_principal.mp3'). "
                                "Laisser vide pour nommer automatiquement."
                            ),
                        },
                        "model": {
                            "type": "string",
                            "enum": [
                                "lyria-3-clip-preview",
                                "lyria-3-pro-preview",
                            ],
                            "description": (
                                "Modèle Lyria. "
                                "lyria-3-clip-preview : 30 s fixe (défaut). "
                                "lyria-3-pro-preview : ~2 min avec structure complète."
                            ),
                        },
                    },
                    "required": ["prompt"],
                },
            },
        },
    ]

    if folder_valid:
        tools += [
            # ── Outils dossier (nécessitent un dossier ouvert) ────────────────
            {
                "type": "function",
                "function": {
                    "name": "organize_files",
                    "description": (
                        "Déplace des fichiers du dossier ouvert vers des sous-dossiers. "
                        "Une boîte de confirmation est présentée à l'utilisateur avant toute exécution. "
                        "IMPORTANT : utilise le dossier parent pour la catégorie générale "
                        "(ex. 'Audio' pour un fichier .mp3), et un sous-dossier imbriqué uniquement "
                        "pour les sous-catégories bien distinctes "
                        "(ex. 'Audio/Midi' pour les fichiers .mid/.midi). "
                        "Ne range pas un fichier générique dans un sous-dossier spécialisé."
                    ),
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "actions": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "filename": {
                                            "type": "string",
                                            "description": "Nom du fichier à déplacer",
                                        },
                                        "destination_subfolder": {
                                            "type": "string",
                                            "description": (
                                                "Sous-dossier de destination (créé si absent). "
                                                "Peut être imbriqué avec '/' (ex. 'Audio/Midi'). "
                                                "Préférer le dossier parent pour les types généraux, "
                                                "les sous-dossiers pour les sous-catégories précises."
                                            ),
                                        },
                                    },
                                    "required": ["filename", "destination_subfolder"],
                                },
                                "description": "Liste des déplacements à effectuer",
                            },
                            "summary": {
                                "type": "string",
                                "description": "Résumé de l'organisation proposée",
                            },
                        },
                        "required": ["actions"],
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "analyze_images",
                    "description": (
                        "Analyse visuellement les images du dossier ouvert, une par une, "
                        "pour répondre à une question. Exemples : trouver les personnes "
                        "portant du rouge, identifier les photos floues, décrire chaque image, "
                        "sélectionner les meilleures photos selon des critères. "
                        "IMPORTANT : mets dans 'question' TOUS les critères de l'utilisateur, "
                        "de façon précise et exhaustive — ce qu'il faut GARDER comme ce qu'il "
                        "faut ÉCARTER (ex. : garder les photos de groupe nettes et les "
                        "personnes mises en valeur ; écarter les photos de dos, floues ou ratées). "
                        "La qualité du tri dépend directement de la précision de cette question. "
                        "Fonde ensuite ta sélection sur le verdict renvoyé pour chaque fichier, "
                        "sans deviner. "
                        "N'utilise PAS cet outil si l'image est déjà jointe directement "
                        "dans le message de l'utilisateur — réponds directement à partir de cette image."
                    ),
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "filenames": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": (
                                    "Liste des noms de fichiers images à analyser. "
                                    "Laisser vide pour analyser toutes les images du dossier."
                                ),
                            },
                            "question": {
                                "type": "string",
                                "description": "Question visuelle à poser pour chaque image",
                            },
                        },
                        "required": ["question"],
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "score_photos",
                    "description": (
                        "Note chaque image du dossier ouvert sur des critères fixes "
                        "(netteté, cadrage, expression, exposition — voir "
                        "CONSTANTS.AI_PHOTO_SCORE_CRITERIA) plus d'éventuels critères "
                        "additionnels propres à ce tri précis. Contrairement à "
                        "analyze_images (verdict libre), écrit un score structuré "
                        "(note 0-10 + raison courte par critère, score global) dans "
                        "un fichier .ai_photo_scores.json du dossier — exploitable "
                        "ensuite par Dashboard pour copier automatiquement les images "
                        "au-dessus du seuil, et par Charles pour affiner les notes à "
                        "la main. Utilise cet outil (pas analyze_images) dès que "
                        "Charles demande de 'noter', 'scorer' ou 'trier par qualité' "
                        "ses photos. Si le contexte du tri (type de shooting, ce qui "
                        "compte le plus) n'est pas clair, pose la question via "
                        "ask_clarifying_question avant d'appeler cet outil plutôt que "
                        "de deviner."
                    ),
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "filenames": {
                                "type": "array",
                                "items": {"type": "string"},
                                "description": (
                                    "Liste des noms de fichiers images à noter. "
                                    "Laisser vide pour noter toutes les images du dossier."
                                ),
                            },
                            "contexte": {
                                "type": "string",
                                "description": (
                                    "Contexte du tri pour affiner le jugement, ex. "
                                    "'mariage, préférer les sourires naturels' ou "
                                    "'photos d'identité, cadrage strict'. Laisser vide "
                                    "si Charles n'a rien précisé."
                                ),
                            },
                            "criteres_additionnels": {
                                "type": "array",
                                "items": {
                                    "type": "object",
                                    "properties": {
                                        "nom": {"type": "string"},
                                        "description": {"type": "string"},
                                    },
                                    "required": ["nom", "description"],
                                },
                                "description": (
                                    "Critères propres à ce tri, en plus des critères "
                                    "fixes (netteté, cadrage, expression, exposition)."
                                ),
                            },
                        },
                        "required": [],
                    },
                },
            },
            {
                "type": "function",
                "function": {
                    "name": "edit_image",
                    "description": (
                        "Modifie une image existante du dossier ouvert avec un prompt texte "
                        "via Nano Banana 2 (gemini-3.1-flash-image-preview). "
                        "Idéal pour : changer le style, ajouter/supprimer des éléments, "
                        "changer les couleurs, transformer une photo en illustration. "
                        "L'image modifiée est sauvegardée dans le dossier ouvert."
                    ),
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "source_filename": {
                                "type": "string",
                                "description": "Nom du fichier image source à modifier (dans le dossier ouvert).",
                            },
                            "prompt": {
                                "type": "string",
                                "description": (
                                    "Instructions de modification, ex. : "
                                    "'Transforme en style aquarelle', "
                                    "'Remplace le fond par un coucher de soleil'."
                                ),
                            },
                            "output_filename": {
                                "type": "string",
                                "description": (
                                    "Nom du fichier de sortie (ex. 'photo_aquarelle.png'). "
                                    "Laisser vide pour nommer automatiquement."
                                ),
                            },
                            "aspect_ratio": {
                                "type": "string",
                                "enum": ["1:1", "16:9", "9:16", "4:3", "3:4", "3:2", "2:3"],
                                "description": "Format de sortie. Défaut : même que l'original.",
                            },
                        },
                        "required": ["source_filename", "prompt"],
                    },
                },
            },
        ]

    return tools


def _gemini_tool_definitions(folder_path):
    """
    Retourne tous les outils dossier pour Gemini, y compris analyze_images.
    Les outils web (web_search, fetch_url) sont filtrés dans
    _gemini_chat_stream_with_tools et remplacés par google_search natif.
    """
    return _folder_tool_definitions(folder_path)


def build_tool_list(folder_path, mcp_tools=None, extra_tools=None):
    """
    Construit la liste complète d'outils envoyée au modèle — MUTUALISÉE entre
    Dashboard et SidePanel (source unique, plus de double maintenance).

    Paramètres :
      folder_path : dossier ouvert (outils dossier + generate/edit/iterate image)
      mcp_tools   : sortie de mcp_client.mcp_get_all_tools(), passée par
                    l'appelant (ai_tools reste découplé de mcp_client)
      extra_tools : outils propres à une seule app (ex. _IMAGE_ITERATE_TOOLS
                    pour Dashboard) ; None ailleurs

    L'ordre reproduit exactement l'assemblage historique (important pour la
    stabilité du cache de préfixe outils). Le filtrage web_search/fetch_url
    pour Gemini est fait en aval dans _gemini_chat_stream_with_tools : la liste
    renvoyée est la même quel que soit le modèle.
    """
    new_tools = (
        _EDIT_TOOLS + _READ_LINES_TOOLS + _SEARCH_TOOLS + _GIT_TOOLS
        + _TASK_TOOLS + _PDF_TOOLS + _SUBAGENT_TOOLS + _SCHEDULE_TOOLS
        + _HTTP_TOOLS + _SPREADSHEET_TOOLS + _PYAUTOGUI_TOOLS + _SSH_TOOLS
        + list(extra_tools or [])
    )
    return (
        _WEB_TOOLS + _TERMINAL_TOOLS + _MEMORY_TOOLS + _SCREENSHOT_TOOLS
        + _NOTEPAD_TOOLS + _UI_TOOLS + new_tools + list(mcp_tools or [])
        + _folder_tool_definitions(folder_path)
    )


# Sentinelle : renvoyée par dispatch_folder_tool quand fn_name n'appartient pas
# au groupe d'outils « purs » géré ici (l'appelant traite alors ses propres
# branches UI-lourdes : generate_image, organize_files, delete_files, etc.).
DISPATCH_UNHANDLED = object()


def dispatch_folder_tool(fn_name, fn_args, folder_path, ui):
    """
    Exécute une branche d'outil « pure » (résultat = chaîne) — MUTUALISÉE entre
    Dashboard et SidePanel. Renvoie la chaîne résultat, ou DISPATCH_UNHANDLED si
    fn_name n'est pas géré par ce groupe.

    Traduction fidèle des branches read_exif → read_spreadsheet, jadis dupliquées
    à l'identique dans les deux apps. L'appelant fait :
        _r = dispatch_folder_tool(fn_name, fn_args, path, ui)
        if _r is not DISPATCH_UNHANDLED:
            _folder_tool_results.append((fn_name, _r))

    `ui` est un objet fournissant les rappels d'interface (chaque app le construit
    à partir de ses propres widgets) :
      ui.set_status(text)      — texte de statut (sans repaint)
      ui.bubble(text)          — bulle assistant dans le chat (sans repaint)
      ui.event(text)           — événement du tour (export)
      ui.refresh()             — rafraîchit la vue dossier
      ui.paint()               — repaint (try/except géré côté app)
      ui.credential(host, user)— demande/récupère un mot de passe (SSH), ou None
    """
    if fn_name == "move_file":
        _mv_src = fn_args.get("source", "").strip()
        _mv_dst = fn_args.get("destination", "").strip()
        if not _mv_src or not _mv_dst:
            return "Paramètres source ou destination manquants."
        _mv_res = _folder_move_file(folder_path, _mv_src, _mv_dst)
        ui.refresh()
        ui.bubble(f"📦 Déplacement : {_os.path.basename(_mv_src)} → {_mv_dst}")
        ui.event(f"📦 Déplacement : {_os.path.basename(_mv_src)} → {_mv_dst}")
        ui.paint()
        return _mv_res
    elif fn_name == "copy_file":
        _cp_src = fn_args.get("source", "").strip()
        _cp_dst = fn_args.get("destination", "").strip()
        if not _cp_src or not _cp_dst:
            return "Paramètres source ou destination manquants."
        _cp_res = _folder_copy_file(folder_path, _cp_src, _cp_dst)
        ui.refresh()
        ui.bubble(f"📋 Copie : {_os.path.basename(_cp_src)} → {_cp_dst}")
        ui.event(f"📋 Copie : {_os.path.basename(_cp_src)} → {_cp_dst}")
        ui.paint()
        return _cp_res
    elif fn_name == "create_folder":
        _mkdir_path = fn_args.get("path", "").strip()
        if not _mkdir_path:
            return "Chemin manquant."
        _mkdir_res = _folder_create_folder(folder_path, _mkdir_path)
        ui.refresh()
        ui.bubble(f"📁 Dossier créé : {_mkdir_path}")
        ui.event(f"📁 Dossier créé : {_mkdir_path}")
        ui.paint()
        return _mkdir_res
    elif fn_name == "mouse_click":
        _mc_x = int(fn_args.get("x", 0))
        _mc_y = int(fn_args.get("y", 0))
        _mc_button = fn_args.get("button", "left")
        _mc_clicks = fn_args.get("clicks", 1)
        ui.set_status(f"🖱️ Clic ({_mc_x}, {_mc_y})…")
        ui.bubble(f"🖱️ Clic {_mc_button} à ({_mc_x}, {_mc_y})")
        ui.event(f"🖱️ Clic à ({_mc_x}, {_mc_y})")
        ui.paint()
        return _mouse_click(_mc_x, _mc_y, _mc_button, _mc_clicks)
    elif fn_name == "keyboard_type":
        _kt_text = fn_args.get("text", "")
        _kt_short = (_kt_text[:30] + "…") if len(_kt_text) > 30 else _kt_text
        ui.set_status(f"⌨️ Saisie : {_kt_short}…")
        ui.bubble(f"⌨️ Saisie : « {_kt_short} »")
        ui.event(f"⌨️ Saisie : « {_kt_short} »")
        ui.paint()
        return _keyboard_type(_kt_text)
    elif fn_name == "keyboard_hotkey":
        _kh_keys = fn_args.get("keys", [])
        _kh_str = "+".join(_kh_keys)
        ui.set_status(f"⌨️ Raccourci : {_kh_str}…")
        ui.bubble(f"⌨️ Raccourci : {_kh_str}")
        ui.event(f"⌨️ Raccourci : {_kh_str}")
        ui.paint()
        return _keyboard_hotkey(*_kh_keys)
    elif fn_name == "read_exif":
        _exif_files = fn_args.get("filenames", [])
        if not _exif_files:
            return "Aucun fichier fourni."
        return _folder_read_exif(folder_path, _exif_files)
    elif fn_name == "zip_files":
        _zip_paths = fn_args.get("paths", [])
        _zip_name = fn_args.get("zip_name", "archive") or "archive"
        _zip_dest = fn_args.get("destination", "") or None
        if not _zip_paths:
            return "Aucun fichier fourni."
        _zip_res = _folder_zip_files(folder_path, _zip_paths, _zip_name, _zip_dest)
        ui.refresh()
        ui.bubble(f"🗜️ Archive créée : {_zip_name}")
        ui.event(f"🗜️ Archive créée : {_zip_name}")
        ui.paint()
        return _zip_res
    elif fn_name == "unzip_file":
        _unzip_src = fn_args.get("source", "").strip()
        _unzip_dest = fn_args.get("destination", "") or None
        if not _unzip_src:
            return "Source manquante."
        _unzip_res = _folder_unzip_file(folder_path, _unzip_src, _unzip_dest)
        ui.refresh()
        ui.bubble(f"📦 Extrait : {_os.path.basename(_unzip_src)}")
        ui.event(f"📦 Extrait : {_os.path.basename(_unzip_src)}")
        ui.paint()
        return _unzip_res
    elif fn_name == "edit_file":
        _ef_path = fn_args.get("filepath", "").strip()
        _ef_old = fn_args.get("old_string", "")
        _ef_new = fn_args.get("new_string", "")
        if not _ef_path or _ef_old == "":
            return "Paramètres filepath / old_string manquants."
        ui.set_status(f"✏️ Édition : {_os.path.basename(_ef_path)}…")
        ui.bubble(f"✏️ Édition : {_ef_path}")
        ui.paint()
        _ef_res = _edit_file(folder_path, _ef_path, _ef_old, _ef_new)
        ui.refresh()
        return _ef_res
    elif fn_name == "read_file_lines":
        _rl_path = fn_args.get("filepath", "").strip()
        _rl_start = fn_args.get("start_line", 1)
        _rl_end = fn_args.get("end_line", None)
        if not _rl_path:
            return "Paramètre filepath manquant."
        _rl_end_str = str(_rl_end) if _rl_end else "fin"
        ui.set_status(
            f"📄 Lignes {_rl_start}–{_rl_end_str} : {_os.path.basename(_rl_path)}…"
        )
        ui.paint()
        return _read_file_lines(folder_path, _rl_path, _rl_start, _rl_end)
    elif fn_name == "search_in_files":
        _si_pattern = fn_args.get("pattern", "")
        _si_path = (fn_args.get("path", "") or "").strip() or None
        _si_glob = fn_args.get("file_glob", "*") or "*"
        _si_max = int(fn_args.get("max_results", 50) or 50)
        _si_case = bool(fn_args.get("case_sensitive", False))
        if not _si_pattern:
            return "Paramètre 'pattern' manquant."
        ui.set_status(f"🔎 Grep : {_si_pattern}…")
        ui.paint()
        return _search_in_files(
            folder_path, _si_pattern,
            path=_si_path, file_glob=_si_glob,
            max_results=_si_max, case_sensitive=_si_case,
        )
    elif fn_name == "find_files":
        _ff_pattern = fn_args.get("pattern", "")
        _ff_basepath = (fn_args.get("base_path", "") or "").strip() or None
        _ff_max = int(fn_args.get("max_results", 200) or 200)
        if not _ff_pattern:
            return "Paramètre 'pattern' manquant."
        ui.set_status(f"🔎 Glob : {_ff_pattern}…")
        ui.paint()
        return _find_files(
            folder_path, _ff_pattern,
            base_path=_ff_basepath, max_results=_ff_max,
        )
    elif fn_name == "git_command":
        _git_args = fn_args.get("args", [])
        _git_cwd = (fn_args.get("cwd", "") or "").strip() or folder_path or None
        if not _git_args:
            return "Paramètre 'args' manquant."
        _git_label = " ".join(str(a) for a in _git_args[:3])
        ui.set_status(f"🔀 git {_git_label}…")
        ui.bubble(f"🔀 git {' '.join(str(a) for a in _git_args)}")
        ui.paint()
        return _git_command(_git_args, cwd=_git_cwd)
    elif fn_name == "manage_tasks":
        return _manage_tasks(
            fn_args.get("action", "list"),
            task_id=fn_args.get("task_id") or None,
            title=fn_args.get("title") or None,
            status=fn_args.get("status") or None,
            notes=fn_args.get("notes") or None,
        )
    elif fn_name == "read_pdf":
        _pdf_path = fn_args.get("filepath", "").strip()
        _pdf_pages = fn_args.get("pages") or None
        if not _pdf_path:
            return "Paramètre 'filepath' manquant."
        ui.set_status(f"📄 PDF : {_os.path.basename(_pdf_path)}…")
        ui.paint()
        return _read_pdf(folder_path, _pdf_path, pages=_pdf_pages)
    elif fn_name == "ask_subagent":
        _sa_task = fn_args.get("task", "")
        _sa_context = fn_args.get("context") or None
        _sa_model = fn_args.get("model") or None
        if not _sa_task:
            return "Paramètre 'task' manquant."
        _sa_short = (_sa_task[:50] + "…") if len(_sa_task) > 50 else _sa_task
        ui.set_status(f"🤖 Sous-agent : {_sa_short}…")
        ui.bubble(f"🤖 Sous-agent : {_sa_short}")
        ui.paint()
        return _ask_subagent(_sa_task, context=_sa_context, model=_sa_model)
    elif fn_name == "schedule_task":
        ui.set_status(f"⏰ Planificateur : {fn_args.get('action', 'list')}…")
        ui.paint()
        return _schedule_task(
            fn_args.get("action", "list"),
            name=fn_args.get("name") or None,
            command=fn_args.get("command") or None,
            when=fn_args.get("when") or None,
        )
    elif fn_name == "http_request":
        _hr_method = (fn_args.get("method", "GET") or "GET").upper()
        _hr_url = fn_args.get("url", "").strip()
        if not _hr_url:
            return "Paramètre 'url' manquant."
        ui.set_status(f"🌐 {_hr_method} {_hr_url[:60]}…")
        ui.paint()
        return _http_request(
            _hr_method, _hr_url,
            headers=fn_args.get("headers") or None,
            body=fn_args.get("body") or None,
            timeout=fn_args.get("timeout") or 30,
        )
    elif fn_name == "ssh_command":
        _ssh_host = fn_args.get("host", "").strip()
        _ssh_user = fn_args.get("username", "").strip()
        _ssh_port = fn_args.get("port") or 22
        _ssh_name = (fn_args.get("name") or "").strip()
        if _ssh_name:
            _srv = next((s for s in CONSTANTS.SSH_SERVERS
                        if s["name"].lower() == _ssh_name.lower()), None)
            if _srv is None:
                _known = ", ".join(s["name"] for s in CONSTANTS.SSH_SERVERS) or "aucun"
                return f"Serveur '{_ssh_name}' inconnu. Serveurs connus : {_known}."
            _ssh_host = _srv["host"]
            _ssh_user = _srv["username"]
            _ssh_port = fn_args.get("port") or _srv.get("port", 22)
        _ssh_cmd = fn_args.get("command", "")
        if not _ssh_host or not _ssh_user or not _ssh_cmd:
            return "Paramètres 'host'/'username' (ou 'name') et 'command' requis."
        ui.set_status(f"🔐 SSH {_ssh_user}@{_ssh_host}…")
        ui.paint()
        _ssh_pwd = ui.credential(_ssh_host, _ssh_user)
        if _ssh_pwd is None:
            return "Connexion annulée par l'utilisateur (mot de passe non fourni)."
        return _ssh_command(
            _ssh_host, _ssh_user, _ssh_pwd, _ssh_cmd,
            port=_ssh_port,
            timeout=fn_args.get("timeout") or 30,
        )
    elif fn_name == "read_spreadsheet":
        _ss_path = fn_args.get("filepath", "").strip()
        if not _ss_path:
            return "Paramètre 'filepath' manquant."
        ui.set_status(f"📊 Tableur : {_os.path.basename(_ss_path)}…")
        ui.paint()
        return _read_spreadsheet(
            folder_path, _ss_path,
            sheet=fn_args.get("sheet") or None,
            max_rows=fn_args.get("max_rows") or 100,
        )
    return DISPATCH_UNHANDLED


def _folder_list_contents(folder_path):
    """
    Retourne une chaîne listant fichiers et sous-dossiers avec taille et date.
    """
    try:
        scan_entries = sorted(
            _os.scandir(folder_path),
            key=lambda entry: (entry.is_file(), entry.name.lower()),
        )
        lines = [f"Dossier : {folder_path}", ""]
        for scan_entry in scan_entries:
            if scan_entry.is_dir():
                lines.append(f"  {scan_entry.name}/  [dossier]")
            else:
                scan_stat = scan_entry.stat()
                size_kb   = scan_stat.st_size / 1024
                size_str  = (
                    f"{size_kb:.0f} KB" if size_kb < 1024 else f"{size_kb / 1024:.1f} MB"
                )
                mtime_str = _datetime_module.datetime.fromtimestamp(
                    scan_stat.st_mtime
                ).strftime("%Y-%m-%d %H:%M")
                lines.append(f"  {scan_entry.name}  ({size_str}, {mtime_str})")
        return "\n".join(lines)
    except Exception as exc:
        return f"Erreur lors de la lecture du dossier : {exc}"


def _backup_file(path):
    """
    Copie un fichier/dossier existant dans le dossier de sauvegarde AVANT qu'il
    ne soit écrasé ou supprimé — filet anti-perte de données, général (appelé
    par toute opération destructrice locale). Un index.jsonl trace
    origine → sauvegarde pour permettre une restauration manuelle.

    Ne lève jamais : retourne le chemin de sauvegarde, ou None si rien à sauver.
    """
    try:
        if not getattr(CONSTANTS, "AI_BACKUP_ENABLED", True):
            return None
        if not path or not _os.path.exists(path):
            return None
        import shutil as _sh_bak
        _dirname = getattr(CONSTANTS, "AI_BACKUP_DIRNAME", ".ai_backups")
        _base_dir = _os.path.join(
            _os.path.dirname(_os.path.abspath(__file__)), _dirname, "files")
        _os.makedirs(_base_dir, exist_ok=True)
        _ts = _time.strftime("%Y%m%d_%H%M%S")
        _name = _os.path.basename(path.rstrip("/\\")) or "root"
        _dest = _os.path.join(_base_dir, f"{_ts}_{_name}")
        _n = 1
        while _os.path.exists(_dest):
            _dest = _os.path.join(_base_dir, f"{_ts}_{_name}_{_n}")
            _n += 1
        if _os.path.isdir(path):
            _sh_bak.copytree(path, _dest)
        else:
            _sh_bak.copy2(path, _dest)
        try:
            with open(_os.path.join(_base_dir, "index.jsonl"),
                      "a", encoding="utf-8") as _idx:
                _idx.write(json.dumps(
                    {"timestamp": _ts,
                     "original": _os.path.abspath(path),
                     "backup": _dest},
                    ensure_ascii=False) + "\n")
        except Exception:
            pass
        return _dest
    except Exception as exc:
        _logger.warning("backup fichier échoué pour %r : %r", path, exc)
        return None


def _folder_create_file(folder_path, filename, content):
    """
    Crée ou remplace un fichier texte.
    Accepte un chemin absolu ou un nom/chemin relatif au dossier ouvert.
    Crée les répertoires parents si nécessaire.
    """
    try:
        if not filename:
            return "Nom de fichier invalide."
        file_path = _resolve_path(folder_path, filename)
        if _os.path.exists(file_path):
            _backup_file(file_path)  # écrasement -> on sauve l'ancienne version
        _os.makedirs(_os.path.dirname(_os.path.abspath(file_path)), exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as file_handle:
            file_handle.write(content)
        return f"Fichier créé : {file_path} ({len(content)} caractère(s))"
    except Exception as exc:
        return f"Erreur lors de la création du fichier : {exc}"


def _resolve_path(folder_path, path_arg):
    """Résout path_arg en chemin absolu. Absolu → inchangé. Relatif → joint à folder_path."""
    if _os.path.isabs(path_arg):
        return path_arg
    return _os.path.join(folder_path, path_arg) if folder_path else path_arg


def _folder_read_file(folder_path, filename, document_exts=None, max_chars=CONSTANTS.AI_FILE_MAX_CHARS):
    """
    Lit le contenu texte d'un fichier.
    Accepte un chemin absolu ou un nom/chemin relatif au dossier ouvert.
    """
    if document_exts is None:
        document_exts = _FOLDER_DOCUMENT_EXTS_DEFAULT
    try:
        if _os.path.isabs(filename):
            file_path = filename
        else:
            file_path = _os.path.join(folder_path, filename) if folder_path else filename
            # Vérification d'extension seulement pour les chemins relatifs
            if _os.path.splitext(filename)[1].lower() not in document_exts:
                return f"Type de fichier non lisible en texte : {filename}"
        if folder_path:
            # ponytail: realpath résout symlinks + ../ avant la comparaison
            real_file = _os.path.realpath(file_path)
            real_folder = _os.path.realpath(folder_path)
            if not real_file.startswith(real_folder + _os.sep) and real_file != real_folder:
                return "Accès refusé : fichier hors du dossier autorisé."
        if not _os.path.isfile(file_path):
            return f"Fichier introuvable : {file_path}"
        with open(file_path, "r", encoding="utf-8", errors="replace") as file_handle:
            content = file_handle.read(max_chars)
        if len(content) == max_chars:
            content += f"\n… (tronqué à {max_chars:,} caractères)"
        return content
    except Exception as exc:
        return f"Erreur : {exc}"


def _folder_delete_files(folder_path, paths):
    """
    Supprime une liste de fichiers ou dossiers vides.
    Chaque entrée peut être un chemin absolu ou relatif au dossier ouvert.
    """
    import shutil as _shutil_del
    results = []
    for p in paths:
        target = _resolve_path(folder_path, p)
        try:
            if not _os.path.exists(target):
                results.append(f"✗ Introuvable : {target}")
            elif _os.path.isdir(target):
                _backup_file(target)  # sauvegarde avant suppression
                _shutil_del.rmtree(target)
                results.append(f"✓ Dossier supprimé : {target}")
            else:
                _backup_file(target)  # sauvegarde avant suppression
                _os.remove(target)
                results.append(f"✓ Supprimé : {_os.path.basename(target)}")
        except Exception as exc:
            results.append(f"✗ Erreur ({_os.path.basename(target)}) : {exc}")
    return "\n".join(results) if results else "Aucun fichier à supprimer."


def _folder_move_file(folder_path, source, destination):
    """
    Déplace ou renomme un fichier ou dossier.
    Source et destination peuvent être des chemins absolus ou relatifs au dossier ouvert.
    """
    import shutil as _shutil_mv
    try:
        src = _resolve_path(folder_path, source)
        dst = _resolve_path(folder_path, destination)
        if not _os.path.exists(src):
            return f"Source introuvable : {src}"
        if _os.path.isdir(dst):
            dst = _os.path.join(dst, _os.path.basename(src))
        if _os.path.exists(dst):
            _backup_file(dst)  # écrasement de la destination -> sauvegarde
        _os.makedirs(_os.path.dirname(_os.path.abspath(dst)), exist_ok=True)
        _shutil_mv.move(src, dst)
        return f"Déplacé : {src} → {dst}"
    except Exception as exc:
        return f"Erreur lors du déplacement : {exc}"


def _folder_copy_file(folder_path, source, destination):
    """
    Copie un fichier ou un dossier (récursif).
    Source et destination peuvent être des chemins absolus ou relatifs au dossier ouvert.
    """
    import shutil as _shutil_cp
    try:
        src = _resolve_path(folder_path, source)
        dst = _resolve_path(folder_path, destination)
        if not _os.path.exists(src):
            return f"Source introuvable : {src}"
        if _os.path.isdir(src):
            if _os.path.isdir(dst):
                dst = _os.path.join(dst, _os.path.basename(src))
            _shutil_cp.copytree(src, dst)
        else:
            if _os.path.isdir(dst):
                dst = _os.path.join(dst, _os.path.basename(src))
            _os.makedirs(_os.path.dirname(_os.path.abspath(dst)), exist_ok=True)
            _shutil_cp.copy2(src, dst)
        return f"Copié : {src} → {dst}"
    except Exception as exc:
        return f"Erreur lors de la copie : {exc}"


def _folder_create_folder(folder_path, path):
    """
    Crée un dossier (et ses parents si nécessaire).
    Accepte un chemin absolu ou relatif au dossier ouvert.
    """
    try:
        target = _resolve_path(folder_path, path)
        _os.makedirs(target, exist_ok=True)
        return f"Dossier créé : {target}"
    except Exception as exc:
        return f"Erreur lors de la création du dossier : {exc}"


def _folder_read_exif(folder_path, paths):
    """
    Lit les métadonnées EXIF d'une liste d'images.
    Retourne les infos clés (date, appareil, objectif, réglages, GPS).
    """
    try:
        from PIL import Image as _PilImg
        from PIL.ExifTags import TAGS as _ETAGS
    except ImportError:
        return "Erreur : Pillow n'est pas installé."

    _WANTED = {
        "DateTimeOriginal", "DateTime", "Make", "Model", "LensModel",
        "ExposureTime", "FNumber", "ISOSpeedRatings", "FocalLength",
        "Flash", "WhiteBalance", "Software",
    }

    def _to_float(v):
        if hasattr(v, "numerator"):
            return float(v)
        if isinstance(v, tuple) and len(v) == 2:
            return v[0] / v[1] if v[1] else 0.0
        return float(v)

    def _fmt(tag, val):
        try:
            if tag == "ExposureTime":
                f = _to_float(val)
                return f"1/{round(1/f)}s" if f < 1 else f"{f:.1f}s"
            if tag == "FNumber":
                return f"f/{_to_float(val):.1f}"
            if tag == "FocalLength":
                return f"{_to_float(val):.0f} mm"
        except Exception:
            pass
        return str(val).strip()

    results = []
    for p in paths:
        file_path = _resolve_path(folder_path, p)
        label = _os.path.basename(file_path)
        try:
            with _PilImg.open(file_path) as img:
                exif = img.getexif()
                if not exif:
                    results.append(f"{label} : Pas de données EXIF.")
                    continue
                lines = [f"{label} :"]
                for tag_id, val in exif.items():
                    name = _ETAGS.get(tag_id, "")
                    if name in _WANTED:
                        lines.append(f"  {name} : {_fmt(name, val)}")
                gps_ifd = exif.get_ifd(34853)
                if gps_ifd:
                    lat_dms = gps_ifd.get(2)
                    lat_ref = gps_ifd.get(1, "N")
                    lon_dms = gps_ifd.get(4)
                    lon_ref = gps_ifd.get(3, "E")
                    if lat_dms and lon_dms:
                        def _dms(dms, ref):
                            d, m, s = [_to_float(x) for x in dms]
                            dd = d + m / 60 + s / 3600
                            return -dd if ref in ("S", "W") else dd
                        lines.append(f"  GPS : {_dms(lat_dms, lat_ref):.6f}, {_dms(lon_dms, lon_ref):.6f}")
                if len(lines) == 1:
                    lines.append("  (Aucune donnée EXIF exploitable)")
                results.append("\n".join(lines))
        except Exception as exc:
            results.append(f"{label} : Erreur — {exc}")
    return "\n\n".join(results) if results else "Aucun fichier fourni."


def _folder_zip_files(folder_path, paths, zip_name="archive", destination=None):
    """
    Crée une archive ZIP à partir d'une liste de fichiers/dossiers.
    Accepte des chemins absolus ou relatifs au dossier ouvert.
    """
    import zipfile as _zf
    if not paths:
        return "Aucun fichier à zipper."
    if destination:
        dest_dir = _resolve_path(folder_path, destination)
    elif folder_path and _os.path.isdir(folder_path):
        dest_dir = folder_path
    else:
        first = _resolve_path(folder_path, paths[0])
        dest_dir = _os.path.dirname(first)
    base = zip_name if zip_name else "archive"
    if not base.lower().endswith(".zip"):
        base += ".zip"
    candidate = _os.path.join(dest_dir, base)
    stem, ext = _os.path.splitext(candidate)
    counter = 1
    while _os.path.exists(candidate):
        candidate = f"{stem}_{counter}{ext}"
        counter += 1
    log = []
    try:
        with _zf.ZipFile(candidate, "w", _zf.ZIP_DEFLATED) as archive:
            for p in paths:
                item = _resolve_path(folder_path, p)
                if not _os.path.exists(item):
                    log.append(f"✗ Introuvable : {item}")
                    continue
                if _os.path.isdir(item):
                    dir_name = _os.path.basename(item)
                    for root, _dirs, files in _os.walk(item):
                        for fname in files:
                            full = _os.path.join(root, fname)
                            archive.write(full, arcname=_os.path.join(dir_name, _os.path.relpath(full, item)))
                    log.append(f"✓ Dossier : {_os.path.basename(item)}")
                else:
                    archive.write(item, arcname=_os.path.basename(item))
                    log.append(f"✓ Fichier : {_os.path.basename(item)}")
        return f"Archive créée : {candidate}\n" + "\n".join(log)
    except Exception as exc:
        return f"Erreur lors de la création de l'archive : {exc}"


def _folder_unzip_file(folder_path, source, destination=None):
    """
    Extrait une archive ZIP dans le dossier de destination.
    Détecte automatiquement si l'archive a un dossier racine unique.
    """
    import zipfile as _zf
    file_path = _resolve_path(folder_path, source)
    if not _os.path.isfile(file_path):
        return f"Fichier introuvable : {file_path}"
    if not file_path.lower().endswith(".zip"):
        return f"Ce n'est pas un fichier .zip : {_os.path.basename(file_path)}"
    if destination:
        extract_to = _resolve_path(folder_path, destination)
        _os.makedirs(extract_to, exist_ok=True)
    else:
        dest_dir = _os.path.dirname(file_path)
        zip_name = _os.path.splitext(_os.path.basename(file_path))[0]
        try:
            with _zf.ZipFile(file_path, "r") as zf:
                names = zf.namelist()
                top_levels = {n.split("/")[0] for n in names if n}
                if len(top_levels) == 1 and any("/" in n for n in names):
                    extract_to = dest_dir
                else:
                    extract_to = _os.path.join(dest_dir, zip_name)
                    _os.makedirs(extract_to, exist_ok=True)
        except Exception as exc:
            return f"Erreur à l'ouverture de l'archive : {exc}"
    try:
        with _zf.ZipFile(file_path, "r") as zf:
            zf.extractall(extract_to)
        return f"Extrait dans : {extract_to}"
    except Exception as exc:
        return f"Erreur lors de l'extraction : {exc}"


def _encode_image_for_analysis(image_path, max_size=1024, quality=70):
    """
    Encode une image en JPEG base64 redimensionnée pour l'envoi à un modèle vision.
    Retourne la chaîne base64 ou None en cas d'échec.
    """
    try:
        import base64 as _base64
        import io as _io
        from PIL import Image as _PilImage, ImageOps as _PilImageOps
        with _PilImage.open(image_path) as pil_img:
            pil_img = _PilImageOps.exif_transpose(pil_img).convert("RGB")
            pil_img.thumbnail((max_size, max_size), _PilImage.LANCZOS)
            buffer = _io.BytesIO()
            pil_img.save(buffer, format="JPEG", quality=quality, optimize=True)
            return _base64.b64encode(buffer.getvalue()).decode("utf-8")
    except Exception:
        return None


def _take_screenshot(max_size=1920, quality=75, region=None):
    """
    Capture l'écran et retourne {"text": str, "b64": str} ou None en cas d'échec.
    region : [x, y, width, height] pour capturer une zone précise (optionnel).
    """
    try:
        import base64 as _base64
        import io as _io
        import sys as _sys
        from PIL import Image as _PilImage

        img = None
        _region_pag = tuple(region) if region else None

        # pyautogui en premier — cross-platform, support région natif
        try:
            import pyautogui as _pag
            img = _pag.screenshot(region=_region_pag)
        except Exception:
            pass

        # ponytail: fallback PIL si pyautogui absent
        if img is None:
            try:
                from PIL import ImageGrab as _ImageGrab
                img = _ImageGrab.grab()
            except Exception:
                pass

        # ponytail: fallback Linux via outils système — scrot est le plus courant
        if img is None and _sys.platform.startswith("linux"):
            import subprocess as _sp
            import tempfile as _tf
            import os as _os2
            with _tf.NamedTemporaryFile(suffix=".png", delete=False) as _tmp:
                _tmp_path = _tmp.name
            try:
                for _cmd in (
                    ["scrot", _tmp_path],
                    ["gnome-screenshot", "-f", _tmp_path],
                    ["import", "-window", "root", _tmp_path],
                ):
                    if _sp.call(_cmd, stdout=_sp.DEVNULL, stderr=_sp.DEVNULL) == 0:
                        img = _PilImage.open(_tmp_path)
                        img.load()
                        break
            finally:
                try:
                    _os2.unlink(_tmp_path)
                except Exception:
                    pass

        if img is None:
            return None

        w, h = img.size
        img = img.convert("RGB")

        # Dessiner la position du curseur (absent des captures natives) pour
        # que le modèle voie où la souris se trouve réellement.
        try:
            import pyautogui as _pag
            from PIL import ImageDraw as _ImageDraw
            _cx, _cy = _pag.position()
            if region:
                _cx -= region[0]
                _cy -= region[1]
            if 0 <= _cx < w and 0 <= _cy < h:
                _draw = _ImageDraw.Draw(img)
                _r = max(8, w // 150)
                _draw.ellipse(
                    [_cx - _r, _cy - _r, _cx + _r, _cy + _r],
                    outline="red", width=max(2, _r // 4),
                )
                _draw.line([_cx - _r * 2, _cy, _cx + _r * 2, _cy], fill="red", width=2)
                _draw.line([_cx, _cy - _r * 2, _cx, _cy + _r * 2], fill="red", width=2)
        except Exception:
            pass

        img.thumbnail((max_size, max_size), _PilImage.LANCZOS)
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        b64 = _base64.b64encode(buf.getvalue()).decode("utf-8")
        region_str = f" région ({region[0]},{region[1]} {region[2]}×{region[3]})" if region else ""
        return {
            "text": (
                f"Screenshot capturé ({w}×{h} px{region_str}). "
                "Le cercle rouge avec croix marque la position actuelle du curseur "
                "(absent nativement des captures) — pas un élément de l'interface."
            ),
            "b64": b64,
        }
    except Exception:
        return None


def _mouse_warp_macos(x, y):
    """
    macOS uniquement : repositionne le curseur via CGWarpMouseCursorPosition.

    Après une longue session d'appels répétés, le déplacement standard de
    pyautogui (CGEventPost d'un évènement mouseMoved) reste parfois bloqué :
    la file d'évènements HID sature et le curseur ne bouge plus, sans lever
    d'erreur. CGWarpMouseCursorPosition force la position au niveau système,
    contournement plus fiable que le simple post d'évènement.
    """
    try:
        import Quartz as _Quartz
        _Quartz.CGWarpMouseCursorPosition((x, y))
        _Quartz.CGAssociateMouseAndMouseCursorPosition(True)
        return True
    except Exception:
        return False


def _mouse_click(x, y, button="left", clicks=1):
    """Clique à la position (x, y) sur l'écran, et vérifie que le curseur y est bien arrivé."""
    try:
        import pyautogui as _pag
        _is_macos = _sys.platform == "darwin"

        for _attempt in range(2):
            if _is_macos:
                _mouse_warp_macos(x, y)
            _pag.click(x, y, button=button, clicks=int(clicks))
            _actual_x, _actual_y = _pag.position()
            if abs(_actual_x - x) <= 3 and abs(_actual_y - y) <= 3:
                break

        _suffix = f' ×{clicks}' if int(clicks) > 1 else ''
        if abs(_actual_x - x) > 3 or abs(_actual_y - y) > 3:
            return (
                f"[ATTENTION] Clic {button} demandé à ({x}, {y}){_suffix} mais le curseur "
                f"est en réalité en ({_actual_x}, {_actual_y}) après le clic — le curseur "
                f"n'a probablement pas atteint la bonne position, le clic a pu manquer sa cible."
            )
        return f"Clic {button} à ({x}, {y}){_suffix}."
    except ImportError:
        return "[Erreur] pyautogui n'est pas installé. Installe : pip install pyautogui"
    except Exception as exc:
        return f"[Erreur] {exc}"


def _keyboard_type(text):
    """Saisit du texte via le presse-papiers (unicode complet, cross-platform)."""
    import platform as _plt
    import subprocess as _spc
    try:
        import pyautogui as _pag
    except ImportError:
        return "[Erreur] pyautogui n'est pas installé. Installe : pip install pyautogui"
    try:
        plat = _plt.system()
        if plat == "Darwin":
            _spc.run(["pbcopy"], input=text.encode("utf-8"), check=True)
            _pag.hotkey("command", "v")
        elif plat == "Windows":
            _spc.run(["clip"], input=text.encode("utf-16-le"), check=True)
            _pag.hotkey("ctrl", "v")
        else:
            _spc.run(["xclip", "-selection", "clipboard"],
                     input=text.encode("utf-8"), check=True)
            _pag.hotkey("ctrl", "v")
        short = (text[:40] + "…") if len(text) > 40 else text
        return f"Texte saisi : « {short} »"
    except Exception as exc:
        return f"[Erreur] {exc}"


def _keyboard_hotkey(*keys):
    """Appuie sur un raccourci clavier."""
    try:
        import pyautogui as _pag
        _pag.hotkey(*keys)
        return f"Raccourci : {'+'.join(keys)}"
    except ImportError:
        return "[Erreur] pyautogui n'est pas installé. Installe : pip install pyautogui"
    except Exception as exc:
        return f"[Erreur] {exc}"


def _analyze_images_batched(
    ollama_url,
    model,
    folder_path,
    filenames,
    question,
    batch_size=5,
    image_exts=None,
    max_size=1024,
    quality=70,
    on_progress=None,
    is_running=None,
):
    """
    Analyse visuellement une liste d'images par lots en posant une question à chaque lot.

    Paramètres :
      ollama_url  — URL de l'API Ollama (ex. "http://localhost:11434")
      model       — nom du modèle vision
      folder_path — chemin du dossier contenant les images
      filenames   — liste de noms de fichiers à analyser (liste vide = toutes les images)
      question    — question à poser pour chaque image
      batch_size  — nombre d'images par appel IA
      image_exts  — frozenset d'extensions images (défaut : _FOLDER_IMAGE_EXTS_DEFAULT)
      max_size    — résolution max des images encodées (px)
      quality     — qualité JPEG des images encodées
      on_progress — callback(batch_num, total_batches) appelé avant chaque lot
      is_running  — callable() → bool, interrompt la boucle si False (optionnel)

    Retourne une liste de chaînes de résultats (une par lot).
    """
    if image_exts is None:
        image_exts = _FOLDER_IMAGE_EXTS_DEFAULT
    if not filenames:
        try:
            filenames = sorted([
                entry.name for entry in _os.scandir(folder_path)
                if entry.is_file()
                and _os.path.splitext(entry.name)[1].lower() in image_exts
            ])
        except Exception:
            filenames = []

    total         = len(filenames)
    total_batches = (total + batch_size - 1) // batch_size if total else 0
    results       = []

    for batch_idx, batch_start in enumerate(range(0, total, batch_size)):
        if is_running is not None and not is_running():
            break
        batch_num   = batch_idx + 1
        batch_names = filenames[batch_start : batch_start + batch_size]
        if on_progress:
            on_progress(batch_num, total_batches)

        b64_list      = []
        encoded_names = []
        for fname in batch_names:
            b64 = _encode_image_for_analysis(
                _os.path.join(folder_path, fname),
                max_size=max_size,
                quality=quality,
            )
            if b64:
                b64_list.append(b64)
                encoded_names.append(fname)
        if not b64_list:
            continue

        prompt = (
            f"Tâche : {question}\n\n"
            f"Voici {len(encoded_names)} image(s), dans cet ordre : "
            f"{', '.join(encoded_names)}.\n"
            "Examine CHAQUE image individuellement et avec rigueur (netteté, "
            "cadrage, qui est visible et comment, dos vs visages, etc.). "
            "Réponds sur UNE ligne par image, au format strict :\n"
            "NomFichier : verdict précis pour CETTE image, répondant "
            "directement à la tâche. Si l'image ne correspond pas aux "
            "critères, dis-le explicitement et pourquoi.\n"
            "Ne regroupe pas les images, ne sois pas complaisant, n'invente "
            "rien : juge ce que tu vois réellement sur chaque photo."
        )
        try:
            if model.startswith("gemini"):
                import concurrent.futures as _cf_ab
                from google import genai as _genai_ab
                from google.genai import types as _gtypes_ab
                import base64 as _b64_ab
                _client_ab = _genai_ab.Client()
                _parts_ab = [
                    _gtypes_ab.Part.from_bytes(
                        data=_b64_ab.b64decode(b64), mime_type="image/jpeg"
                    )
                    for b64 in b64_list
                ]
                _parts_ab.append(_gtypes_ab.Part(text=prompt))
                _contents_ab = [_gtypes_ab.Content(role="user", parts=_parts_ab)]
                _config_ab = _gtypes_ab.GenerateContentConfig(temperature=0.2)
                with _cf_ab.ThreadPoolExecutor(max_workers=1) as _ex_ab:
                    _fut_ab = _ex_ab.submit(
                        _client_ab.models.generate_content,
                        model=model,
                        contents=_contents_ab,
                        config=_config_ab,
                    )
                    try:
                        _resp_ab = _fut_ab.result(timeout=120)
                        results.append(_resp_ab.text or "")
                    except _cf_ab.TimeoutError:
                        results.append(
                            f"(lot {batch_num} — timeout Gemini après 120 s)"
                        )
            else:
                response = _ollama_chat_once(
                    ollama_url,
                    model,
                    [{"role": "user", "content": prompt, "images": b64_list}],
                    temperature=0.2,
                )
                results.append(response.get("content", ""))
        except Exception as exc:
            results.append(f"(lot {batch_num} — erreur : {exc})")

    return results


def _score_images_batched(
    ollama_url,
    model,
    folder_path,
    filenames,
    contexte="",
    criteres_additionnels=None,
    batch_size=5,
    image_exts=None,
    max_size=1024,
    quality=70,
    on_progress=None,
    is_running=None,
):
    """
    Note une liste d'images par lots sur les critères fixes de
    CONSTANTS.AI_PHOTO_SCORE_CRITERIA (+ criteres_additionnels/contexte),
    et écrit le résultat fusionné dans folder_path/CONSTANTS.AI_PHOTO_SCORE_FILE
    (les entrées non retraitées à ce tour sont conservées).

    Retourne une chaîne de résumé (nb noté, nb au-dessus du seuil).
    """
    if image_exts is None:
        image_exts = _FOLDER_IMAGE_EXTS_DEFAULT
    if not filenames:
        try:
            filenames = sorted([
                entry.name for entry in _os.scandir(folder_path)
                if entry.is_file()
                and _os.path.splitext(entry.name)[1].lower() in image_exts
            ])
        except Exception:
            filenames = []

    criteres_additionnels = criteres_additionnels or []
    criteres_lignes = [
        f"- {cle} ({label})"
        for cle, label in CONSTANTS.AI_PHOTO_SCORE_CRITERIA.items()
    ]
    criteres_cles = list(CONSTANTS.AI_PHOTO_SCORE_CRITERIA.keys())
    for crit in criteres_additionnels:
        nom  = (crit.get("nom") or "").strip()
        desc = (crit.get("description") or "").strip()
        if nom:
            criteres_lignes.append(f"- {nom} ({desc})" if desc else f"- {nom}")
            criteres_cles.append(nom)
    criteres_bloc = "\n".join(criteres_lignes)

    total         = len(filenames)
    total_batches = (total + batch_size - 1) // batch_size if total else 0
    scored        = {}   # nom de fichier -> dict de score

    for batch_idx, batch_start in enumerate(range(0, total, batch_size)):
        if is_running is not None and not is_running():
            break
        batch_num   = batch_idx + 1
        batch_names = filenames[batch_start : batch_start + batch_size]
        if on_progress:
            on_progress(batch_num, total_batches)

        b64_list      = []
        encoded_names = []
        for fname in batch_names:
            b64 = _encode_image_for_analysis(
                _os.path.join(folder_path, fname),
                max_size=max_size,
                quality=quality,
            )
            if b64:
                b64_list.append(b64)
                encoded_names.append(fname)
        if not b64_list:
            continue

        prompt = (
            "Tâche : note chaque image ci-dessous sur les critères suivants, "
            "de 0 (très mauvais) à 10 (irréprochable) :\n"
            f"{criteres_bloc}\n\n"
            + (f"Contexte de ce tri : {contexte}\n\n" if contexte else "")
            + f"Voici {len(encoded_names)} image(s), dans cet ordre : "
              f"{', '.join(encoded_names)}.\n"
            "Sois SÉVÈRE : ce tri sert à ne garder que les meilleurs "
            "clichés qu'un photographe professionnel retouchera ensuite à "
            "la main. Ne sur-note pas par complaisance — réserve les notes "
            "hautes (8-10) aux images réellement irréprochables sur le "
            "critère jugé. Pour CHAQUE image et CHAQUE critère, donne une "
            "note et une raison courte (quelques mots) qui justifie la "
            "note. Calcule aussi un score_global (moyenne des critères) et "
            "un commentaire global d'une phrase."
        )

        _batch_scores = []
        if model.startswith("gemini"):
            try:
                from google import genai as _genai_sc
                from google.genai import types as _gtypes_sc
                import base64 as _b64_sc
                import concurrent.futures as _cf_sc

                _crit_schema_sc = _gtypes_sc.Schema(
                    type=_gtypes_sc.Type.OBJECT,
                    properties={
                        "note": _gtypes_sc.Schema(type=_gtypes_sc.Type.NUMBER),
                        "raison": _gtypes_sc.Schema(type=_gtypes_sc.Type.STRING),
                    },
                    required=["note", "raison"],
                )
                _schema_sc = _gtypes_sc.Schema(
                    type=_gtypes_sc.Type.ARRAY,
                    items=_gtypes_sc.Schema(
                        type=_gtypes_sc.Type.OBJECT,
                        properties={
                            "fichier": _gtypes_sc.Schema(
                                type=_gtypes_sc.Type.STRING),
                            "scores": _gtypes_sc.Schema(
                                type=_gtypes_sc.Type.OBJECT,
                                properties={
                                    cle: _crit_schema_sc
                                    for cle in criteres_cles
                                },
                                required=criteres_cles,
                            ),
                            "score_global": _gtypes_sc.Schema(
                                type=_gtypes_sc.Type.NUMBER),
                            "commentaire": _gtypes_sc.Schema(
                                type=_gtypes_sc.Type.STRING),
                        },
                        required=[
                            "fichier", "scores", "score_global", "commentaire"
                        ],
                    ),
                )
                _client_sc = _genai_sc.Client()
                _parts_sc = [
                    _gtypes_sc.Part.from_bytes(
                        data=_b64_sc.b64decode(b64), mime_type="image/jpeg"
                    )
                    for b64 in b64_list
                ]
                _parts_sc.append(_gtypes_sc.Part(text=prompt))
                _contents_sc = [_gtypes_sc.Content(role="user", parts=_parts_sc)]
                _config_sc = _gtypes_sc.GenerateContentConfig(
                    temperature=0.2,
                    response_mime_type="application/json",
                    response_schema=_schema_sc,
                )
                with _cf_sc.ThreadPoolExecutor(max_workers=1) as _ex_sc:
                    _fut_sc = _ex_sc.submit(
                        _client_sc.models.generate_content,
                        model=model,
                        contents=_contents_sc,
                        config=_config_sc,
                    )
                    _resp_sc = _fut_sc.result(timeout=120)
                    _batch_scores = json.loads(_resp_sc.text or "[]")
            except Exception:
                _batch_scores = []
        else:
            # Ollama : pas de sortie structurée garantie, best-effort JSON.
            try:
                _prompt_json = prompt + (
                    "\n\nRéponds UNIQUEMENT avec un tableau JSON valide, "
                    "sans texte autour, au format : "
                    '[{"fichier": ..., "scores": {"<critere>": '
                    '{"note": ..., "raison": ...}, ...}, '
                    '"score_global": ..., "commentaire": ...}, ...]'
                )
                response = _ollama_chat_once(
                    ollama_url,
                    model,
                    [{"role": "user", "content": _prompt_json,
                      "images": b64_list}],
                    temperature=0.2,
                )
                _batch_scores = json.loads(response.get("content", "") or "[]")
            except Exception:
                _batch_scores = []

        for entry in _batch_scores:
            if not isinstance(entry, dict):
                continue
            fname = (entry.get("fichier") or "").strip()
            if fname in encoded_names:
                scored[fname] = entry

    # Fusion avec le fichier existant : seules les entrées retraitées à ce
    # tour sont remplacées, le reste (dont d'éventuelles corrections
    # manuelles) est conservé.
    scores_path = _os.path.join(folder_path, CONSTANTS.AI_PHOTO_SCORE_FILE)
    existing = []
    if _os.path.isfile(scores_path):
        try:
            with open(scores_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            existing = []
    merged = {
        e.get("fichier", ""): e
        for e in existing if isinstance(e, dict) and e.get("fichier")
    }
    merged.update(scored)
    merged_list = list(merged.values())
    try:
        with open(scores_path, "w", encoding="utf-8") as f:
            json.dump(merged_list, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    above_threshold = sum(
        1 for e in merged_list
        if isinstance(e.get("score_global"), (int, float))
        and e["score_global"] >= CONSTANTS.AI_PHOTO_SCORE_THRESHOLD
    )
    return (
        f"{len(scored)} image(s) notée(s) sur {total} — "
        f"{above_threshold}/{len(merged_list)} au total ≥ seuil "
        f"{CONSTANTS.AI_PHOTO_SCORE_THRESHOLD} dans "
        f"{CONSTANTS.AI_PHOTO_SCORE_FILE}."
    )


def _copy_scored_photos(folder_path):
    """
    Copie dans folder_path/CONSTANTS.AI_PHOTO_SCORE_SELECTION_FOLDER les
    images du fichier de scores (CONSTANTS.AI_PHOTO_SCORE_FILE) dont le
    score_global atteint CONSTANTS.AI_PHOTO_SCORE_THRESHOLD.

    Retourne une chaîne de résumé (copiés / déjà présents / introuvables).
    Action déterministe déclenchée par Charles (bouton Dashboard), pas un
    outil IA — la copie n'est jamais décidée par le modèle seul.
    """
    import shutil as _shutil_cp2

    scores_path = _os.path.join(folder_path, CONSTANTS.AI_PHOTO_SCORE_FILE)
    if not _os.path.isfile(scores_path):
        return f"Aucun fichier de scores trouvé ({CONSTANTS.AI_PHOTO_SCORE_FILE})."
    try:
        with open(scores_path, "r", encoding="utf-8") as f:
            entries = json.load(f)
    except Exception as exc:
        return f"Fichier de scores illisible : {exc}"

    dest_dir = _os.path.join(folder_path, CONSTANTS.AI_PHOTO_SCORE_SELECTION_FOLDER)
    copied, already, missing = [], [], []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        score = entry.get("score_global")
        if not isinstance(score, (int, float)):
            continue
        if score < CONSTANTS.AI_PHOTO_SCORE_THRESHOLD:
            continue
        fname = (entry.get("fichier") or "").strip()
        if not fname:
            continue
        source = _os.path.join(folder_path, fname)
        dest   = _os.path.join(dest_dir, fname)
        if not _os.path.isfile(source):
            missing.append(fname)
            continue
        if _os.path.isfile(dest):
            already.append(fname)
            continue
        try:
            _os.makedirs(dest_dir, exist_ok=True)
            _shutil_cp2.copy2(source, dest)
            copied.append(fname)
        except Exception as exc:
            missing.append(f"{fname} ({exc})")

    lines = [f"{len(copied)} image(s) copiée(s) dans "
             f"{CONSTANTS.AI_PHOTO_SCORE_SELECTION_FOLDER}/."]
    if already:
        lines.append(f"{len(already)} déjà présente(s).")
    if missing:
        lines.append(f"{len(missing)} introuvable(s)/erreur : "
                      + ", ".join(missing[:10]))
    return " ".join(lines)


# ─── Outils terminal partagés ───────────────────────────────────────────────
import subprocess as _subprocess


def _run_elevated(command, cwd=None, timeout=120):
    """
    Exécute une commande shell avec élévation de privilèges (admin/root),
    via l'invite native de l'OS (jamais de mot de passe géré par l'appli) :
      - macOS   : osascript ... with administrator privileges
      - Linux   : pkexec (nécessite un agent PolicyKit graphique, ex. KDE/GNOME)
      - Windows : Start-Process -Verb RunAs (UAC), sortie récupérée via un
                  script .bat temporaire (une élévation UAC ne partage pas
                  ses flux stdout/stderr avec le process parent)
    """
    import platform as _platform_el
    import shlex as _shlex_el

    system = _platform_el.system()
    command = command.strip()
    if command.startswith("sudo "):
        command = command[5:]

    try:
        if system == "Darwin":
            if cwd:
                command = f"cd {_shlex_el.quote(cwd)} && {command}"
            _escaped = command.replace("\\", "\\\\").replace('"', '\\"')
            result = _subprocess.run(
                ["osascript", "-e",
                 f'do shell script "{_escaped}" with administrator privileges'],
                capture_output=True, text=True, timeout=timeout,
            )
            output = (result.stdout + result.stderr).strip()

        elif system == "Linux":
            if cwd:
                command = f"cd {_shlex_el.quote(cwd)} && {command}"
            result = _subprocess.run(
                ["pkexec", "bash", "-c", command],
                capture_output=True, text=True, timeout=timeout,
            )
            output = (result.stdout + result.stderr).strip()

        elif system == "Windows":
            import tempfile
            import os as _os_el
            _script_fd, _script_path = tempfile.mkstemp(suffix=".bat")
            _out_path = _script_path + ".out"
            with _os_el.fdopen(_script_fd, "w") as f:
                if cwd:
                    f.write(f'cd /d "{cwd}"\n')
                f.write(f'{command} > "{_out_path}" 2>&1\n')
            _ps = (
                f'Start-Process -FilePath "{_script_path}" '
                f'-Verb RunAs -Wait -WindowStyle Hidden'
            )
            _subprocess.run(
                ["powershell", "-NoProfile", "-Command", _ps],
                timeout=timeout,
            )
            output = ""
            if _os_el.path.exists(_out_path):
                with open(_out_path, encoding="utf-8", errors="replace") as f:
                    output = f.read().strip()
                _os_el.remove(_out_path)
            _os_el.remove(_script_path)

        else:
            return f"[Erreur] OS non reconnu pour l'élévation : {system}"

        return output or "(Commande exécutée en administrateur, pas de sortie)"

    except _subprocess.TimeoutExpired:
        return f"[Timeout : la commande a dépassé {timeout} secondes]"
    except FileNotFoundError as exc:
        return f"[Erreur : outil d'élévation introuvable ({exc})]"
    except Exception as exc:
        return f"[Erreur d'exécution élevée : {exc}]"


def _run_terminal_command(command, cwd=None, timeout=120, admin=False):
    """
    Exécute une commande shell et retourne la sortie combinée stdout + stderr.
    cwd : répertoire de travail (dossier ouvert si fourni).
    admin : si True, exécute via _run_elevated (invite native OS).
    """
    if admin:
        return _run_elevated(command, cwd=cwd, timeout=timeout)
    import os as _os_tc
    _env = _os_tc.environ.copy()
    # Ajouter les chemins qui manquent dans les apps lancées hors terminal
    if _os_tc.name == "nt":  # Windows
        _sep = ";"
        _home = _os_tc.environ.get("USERPROFILE", "C:\\Users\\User")
        _extra = _sep.join([
            _os_tc.path.join(_home, "AppData", "Local", "Programs", "Ollama"),
            "C:\\Program Files\\Ollama",
            "C:\\Program Files\\Git\\bin",
            _os_tc.path.join(_home, "AppData\\Local\\Microsoft\\WindowsApps"),
        ])
    else:  # macOS / Linux
        _sep = ":"
        _extra = _sep.join([
            "/usr/local/bin", "/opt/homebrew/bin", "/opt/homebrew/sbin",
            "/usr/local/sbin", _os_tc.path.expanduser("~/.local/bin"), "/snap/bin",
        ])
    _env["PATH"] = _extra + _sep + _env.get("PATH", "")
    try:
        result = _subprocess.run(
            command,
            shell=True,
            capture_output=True,
            text=True,
            cwd=cwd,
            timeout=timeout,
            env=_env,
        )
        output = result.stdout
        if result.stderr:
            output += ("\n" if output else "") + result.stderr
        output = output.strip()
        if not output:
            output = f"(Commande exécutée, code de retour : {result.returncode})"
        elif result.returncode != 0:
            output += f"\n(Code de retour : {result.returncode})"
        return output
    except _subprocess.TimeoutExpired:
        return f"[Timeout : la commande a dépassé {timeout} secondes]"
    except Exception as exc:
        return f"[Erreur d'exécution : {exc}]"


_TERMINAL_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "run_terminal_command",
            "description": (
                "Exécute une commande shell sur la machine locale. "
                "Utile pour installer des paquets, lancer des scripts, "
                "convertir des formats, lancer des processus, etc. "
                "NE PAS utiliser pour lister le contenu d'un dossier "
                "(utiliser list_folder_contents à la place) ni pour créer un fichier texte "
                "(utiliser create_file à la place). "
                "Une confirmation est toujours demandée à l'utilisateur avant exécution. "
                "Si la commande nécessite des droits administrateur/root, "
                "mettre admin=true au lieu d'écrire 'sudo' dans command : "
                "cela déclenche l'invite native du système (UAC sur Windows, "
                "mot de passe/Touch ID sur macOS, PolicyKit sur Linux)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "command": {
                        "type": "string",
                        "description": "Commande shell à exécuter",
                    },
                    "description": {
                        "type": "string",
                        "description": "Explication courte de ce que fait la commande",
                    },
                    "admin": {
                        "type": "boolean",
                        "description": (
                            "true si la commande nécessite des droits "
                            "administrateur/root (déclenche l'invite native "
                            "de l'OS). Ne jamais préfixer command par 'sudo'."
                        ),
                    },
                },
                "required": ["command"],
            },
        },
    },
]


# ─── Outils web partagés ──────────────────────────────────────────────────────

_WEB_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "web_search",
            "description": (
                "Recherche des informations récentes sur internet via DuckDuckGo. "
                "À utiliser pour les actualités, événements récents, prix, météo, "
                "ou toute information susceptible d'avoir changé depuis la date d'entraînement."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "Requête de recherche",
                    }
                },
                "required": ["query"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "fetch_url",
            "description": (
                "Lit le contenu textuel d'une page web à partir de son URL. "
                "À utiliser pour approfondir un résultat de recherche ou consulter "
                "une page spécifique."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "url": {
                        "type": "string",
                        "description": "URL complète (https://…) de la page à lire",
                    }
                },
                "required": ["url"],
            },
        },
    },
]


# ─── Système de mémoire persistante ─────────────────────────────────────────

_DATA_DIR = _os.path.dirname(_os.path.abspath(__file__))

_MEMORY_CHAR_LIMITS = {
    "memory": 2200,
    "user":   1375,
    "skills": 3000,
}

_MEMORY_FILE_NAMES = {
    "memory": "memory.md",
    "user":   "user.md",
    "skills": "skills.md",
}


def _memory_file_path(target):
    return _os.path.join(_DATA_DIR, _MEMORY_FILE_NAMES[target])


def _read_memory_file(target):
    """Lit un fichier mémoire, retourne le contenu brut (chaîne vide si absent)."""
    try:
        with open(_memory_file_path(target), encoding="utf-8") as file_handle:
            return file_handle.read()
    except (FileNotFoundError, OSError):
        return ""


def _update_memory_file(target, action, content="", old_text=""):
    """
    Met à jour un fichier mémoire persistant (memory.md, user.md, skills.md).

    Actions :
        add     — ajoute une nouvelle entrée (séparée par §)
        replace — remplace l'entrée contenant old_text par content
        remove  — supprime l'entrée contenant old_text

    Retourne une chaîne JSON avec success, message et usage.
    """
    if target not in _MEMORY_CHAR_LIMITS:
        return json.dumps({
            "success": False,
            "error": f"Cible invalide : {target!r}. Utilise 'memory', 'user' ou 'skills'.",
        })

    file_path = _memory_file_path(target)
    char_limit = _MEMORY_CHAR_LIMITS[target]

    try:
        with open(file_path, encoding="utf-8") as file_handle:
            raw_content = file_handle.read()
    except FileNotFoundError:
        raw_content = ""
    except OSError as error:
        return json.dumps({"success": False, "error": str(error)})

    # Séparer les lignes de commentaires HTML du corps
    all_lines = raw_content.splitlines()
    comment_lines = [line for line in all_lines if line.startswith("<!--")]
    header = ("\n".join(comment_lines) + "\n") if comment_lines else ""
    body_lines = [line for line in all_lines if not line.startswith("<!--")]
    body = "\n".join(body_lines).strip()
    entries = [entry.strip() for entry in body.split("§") if entry.strip()]

    if action == "add":
        if not content.strip():
            return json.dumps({"success": False, "error": "content est vide."})
        new_entry = content.strip()
        if new_entry in entries:
            return json.dumps({"success": True, "note": "Entrée déjà présente, aucun doublon ajouté."})
        candidate_entries = entries + [new_entry]
        candidate_body = " §\n".join(candidate_entries)
        candidate_full = header + candidate_body
        if len(candidate_full) > char_limit:
            _entry_sep = " §\n"
            current_usage = f"{len(header + _entry_sep.join(entries))}/{char_limit}"
            return json.dumps({
                "success": False,
                "error": (
                    f"Mémoire à {current_usage} chars. "
                    f"Ajouter cette entrée ({len(new_entry)} chars) dépasserait la limite. "
                    "Consolide ou supprime des entrées existantes d'abord."
                ),
                "current_entries": entries,
                "usage": current_usage,
            })
        entries = candidate_entries

    elif action == "replace":
        if not old_text.strip():
            return json.dumps({"success": False, "error": "old_text est requis pour 'replace'."})
        matching_indices = [index for index, entry in enumerate(entries) if old_text.strip() in entry]
        if not matching_indices:
            return json.dumps({
                "success": False,
                "error": f"Aucune entrée ne contient : {old_text!r}",
                "current_entries": entries,
            })
        if len(matching_indices) > 1:
            return json.dumps({
                "success": False,
                "error": f"old_text ambigu : {len(matching_indices)} entrées correspondent. Précise davantage.",
                "matches": [entries[index] for index in matching_indices],
            })
        entries[matching_indices[0]] = content.strip()
        candidate_body = " §\n".join(entries)
        candidate_full = header + candidate_body
        if len(candidate_full) > char_limit:
            return json.dumps({
                "success": False,
                "error": f"Contenu trop long après remplacement ({len(candidate_full)}/{char_limit} chars).",
            })

    elif action == "remove":
        if not old_text.strip():
            return json.dumps({"success": False, "error": "old_text est requis pour 'remove'."})
        matching_indices = [index for index, entry in enumerate(entries) if old_text.strip() in entry]
        if not matching_indices:
            return json.dumps({
                "success": False,
                "error": f"Aucune entrée ne contient : {old_text!r}",
                "current_entries": entries,
            })
        if len(matching_indices) > 1:
            return json.dumps({
                "success": False,
                "error": f"old_text ambigu : {len(matching_indices)} entrées correspondent.",
                "matches": [entries[index] for index in matching_indices],
            })
        entries.pop(matching_indices[0])

    else:
        return json.dumps({
            "success": False,
            "error": f"Action invalide : {action!r}. Utilise 'add', 'replace' ou 'remove'.",
        })

    new_body = " §\n".join(entries)
    new_full = header + new_body
    try:
        with open(file_path, "w", encoding="utf-8") as file_handle:
            file_handle.write(new_full)
    except OSError as error:
        return json.dumps({"success": False, "error": str(error)})

    usage = f"{len(new_full)}/{char_limit}"
    return json.dumps({
        "success": True,
        "action": action,
        "target": target,
        "entries_count": len(entries),
        "usage": usage,
    })


# ─── Outil mémoire persistante ────────────────────────────────────────────────

_MEMORY_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "update_memory_file",
            "description": (
                "Met à jour ta mémoire persistante entre les sessions. "
                "Utilise cet outil proactivement quand tu apprends quelque chose d'important sur "
                "l'utilisateur, son environnement, ses préférences ou une procédure utile.\n"
                "Cibles :\n"
                "  - 'memory' : tes notes personnelles (environnement, conventions, leçons apprises) — 2 200 chars max\n"
                "  - 'user'   : profil utilisateur (préférences, habitudes, style de communication) — 1 375 chars max\n"
                "  - 'skills' : procédures et techniques apprises (étapes, commandes, workflows) — 3 000 chars max\n"
                "Actions :\n"
                "  - 'add'     : ajoute une nouvelle entrée\n"
                "  - 'replace' : remplace l'entrée contenant old_text par content\n"
                "  - 'remove'  : supprime l'entrée contenant old_text\n"
                "IMPORTANT — le fichier est une liste d'entrées séparées par '§' (visibles telles quelles "
                "dans les sections MÉMOIRE/PROFIL UTILISATEUR/SKILLS injectées plus haut dans ce prompt) : "
                "'replace'/'remove' ciblent UNE SEULE entrée à la fois. old_text doit être un extrait exact "
                "et unique tiré d'UNE SEULE entrée existante — jamais un texte qui couvre plusieurs entrées "
                "ou le fichier entier. Pour corriger plusieurs entrées, appelle cet outil séparément pour "
                "chacune. Si l'appel retourne success:false (old_text introuvable ou ambigu), NE PRÉTENDS "
                "JAMAIS que la mise à jour a réussi : indique l'échec à Charles et corrige old_text à partir "
                "de current_entries/matches renvoyés, ou réessaie avec un extrait plus court et plus précis."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "target": {
                        "type": "string",
                        "enum": ["memory", "user", "skills"],
                        "description": "Fichier cible : 'memory', 'user' ou 'skills'",
                    },
                    "action": {
                        "type": "string",
                        "enum": ["add", "replace", "remove"],
                        "description": "Action à effectuer",
                    },
                    "content": {
                        "type": "string",
                        "description": "Contenu de la nouvelle entrée (requis pour 'add' et 'replace')",
                    },
                    "old_text": {
                        "type": "string",
                        "description": (
                            "Sous-chaîne unique identifiant l'entrée à remplacer ou supprimer "
                            "(requis pour 'replace' et 'remove')"
                        ),
                    },
                },
                "required": ["target", "action"],
            },
        },
    },
]


# ─── Outils bloc-notes ───────────────────────────────────────────────────────

_SCREENSHOT_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "take_screenshot",
            "description": (
                "Capture une image de l'écran de l'utilisateur. "
                "Utilise cet outil quand l'utilisateur demande de regarder son écran, "
                "de voir un site web, une application ou tout ce qui est affiché. "
                "Retourne une image que tu peux analyser visuellement. "
                "Paramètre optionnel region pour capturer une zone précise et réduire la taille envoyée."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "region": {
                        "type": "array",
                        "items": {"type": "integer"},
                        "description": "Zone à capturer [x, y, largeur, hauteur] en pixels. Omis = plein écran.",
                    },
                },
                "required": [],
            },
        },
    },
]


_PYAUTOGUI_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "mouse_click",
            "description": (
                "Clique à une position précise sur l'écran. "
                "Utilise take_screenshot avant pour identifier les coordonnées. "
                "button : 'left' (défaut), 'right', 'middle'. "
                "clicks : 1 (défaut), 2 pour double-clic."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "x": {"type": "integer", "description": "Coordonnée X en pixels"},
                    "y": {"type": "integer", "description": "Coordonnée Y en pixels"},
                    "button": {
                        "type": "string",
                        "enum": ["left", "right", "middle"],
                        "description": "Bouton de la souris (défaut : left)",
                    },
                    "clicks": {
                        "type": "integer",
                        "description": "Nombre de clics (1 = simple, 2 = double)",
                    },
                },
                "required": ["x", "y"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "keyboard_type",
            "description": (
                "Saisit du texte dans le champ actif (celui qui a le focus clavier). "
                "Supporte tous les caractères unicode via le presse-papiers. "
                "Utilise mouse_click d'abord pour donner le focus au bon champ."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "text": {"type": "string", "description": "Texte à saisir"},
                },
                "required": ["text"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "keyboard_hotkey",
            "description": (
                "Appuie sur un raccourci clavier. "
                "Exemples : ['ctrl', 'c'] pour copier, ['alt', 'F4'] pour fermer, "
                "['ctrl', 'alt', 't'] pour ouvrir un terminal. "
                "Sur macOS, utilise 'command' à la place de 'ctrl'."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "keys": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Touches à appuyer simultanément, ex: ['ctrl', 'c']",
                    },
                },
                "required": ["keys"],
            },
        },
    },
]


_NOTEPAD_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_notepad",
            "description": (
                "Lit le contenu actuel du bloc-notes (éditeur de texte intégré). "
                "Utilise cet outil pour consulter les notes de l'utilisateur avant de les modifier."
            ),
            "parameters": {
                "type": "object",
                "properties": {},
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_notepad",
            "description": (
                "Écrit du contenu dans le bloc-notes intégré. "
                "RÈGLE ABSOLUE : appelle TOUJOURS read_notepad d'abord pour vérifier si le bloc-notes contient déjà du texte. "
                "Ne remplace JAMAIS ('replace') le contenu existant sans avoir d'abord demandé confirmation explicite "
                "à Charles dans la conversation et attendu sa réponse — même s'il semble le demander. "
                "Par défaut et en cas de doute, utilise 'append' (aucune confirmation nécessaire pour ajouter). "
                "Si le bloc-notes n'est pas vide, un appel avec action='replace' sera automatiquement rétrogradé en 'append' par l'application. "
                "'prepend' pour ajouter au début."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "content": {
                        "type": "string",
                        "description": "Texte à écrire dans le bloc-notes",
                    },
                    "action": {
                        "type": "string",
                        "enum": ["replace", "append", "prepend"],
                        "description": "replace : remplace tout le contenu. append : ajoute à la fin. prepend : ajoute au début.",
                    },
                },
                "required": ["content", "action"],
            },
        },
    },
]


# ─── Outils interface utilisateur ────────────────────────────────────────────

_UI_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "navigate_to_folder",
            "description": (
                "Navigue vers un dossier dans l'interface (change le dossier courant affiché dans le navigateur de fichiers). "
                "Utilise cet outil pour ouvrir un dossier sans que l'utilisateur ait à le faire manuellement."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Chemin absolu du dossier à ouvrir",
                    },
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "select_files_in_ui",
            "description": (
                "Sélectionne ou désélectionne des fichiers dans l'interface (dans le dossier courant). "
                "Particulièrement utile après analyze_images pour sélectionner automatiquement "
                "les fichiers correspondant à des critères (ex: images floues, portraits, etc.)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filenames": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Liste des noms de fichiers à traiter (noms seuls, sans chemin complet)",
                    },
                    "mode": {
                        "type": "string",
                        "enum": ["replace", "add", "remove"],
                        "description": "replace : remplace toute la sélection actuelle. add : ajoute à la sélection existante. remove : retire de la sélection.",
                    },
                },
                "required": ["filenames", "mode"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "ask_clarifying_question",
            "description": (
                "Pose UNE question à choix limité à Charles pour lever une "
                "ambiguïté AVANT d'agir, plutôt que de deviner ou de partir "
                "dans une mauvaise direction. À utiliser dès qu'une demande "
                "a plusieurs interprétations raisonnables ou qu'il manque "
                "une information structurante (ex. : contexte d'un tri "
                "photo, seuil à utiliser, dossier de destination). Ne PAS "
                "l'utiliser pour des détails mineurs déductibles ou déjà "
                "couverts par une valeur par défaut dans CONSTANTS.py — "
                "une seule question à la fois, jamais en rafale."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "question": {
                        "type": "string",
                        "description": "Question posée à Charles, claire et autonome.",
                    },
                    "options": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": (
                            "2 à 5 choix courts et concrets parmi lesquels "
                            "Charles peut choisir. Charles pourra toujours "
                            "répondre autre chose si aucun ne convient."
                        ),
                    },
                },
                "required": ["question", "options"],
            },
        },
    },
]


def _build_system_content(folder_path=None, today_date_str=None):
    """
    Assemble le contenu complet du message système envoyé à l'IA.

    Lit system.md (obligatoire), puis injecte memory.md, user.md, skills.md.
    Ajoute la date du jour et le contexte du dossier ouvert.

    Args:
        folder_path     : chemin absolu du dossier ouvert, ou None.
        today_date_str  : date du jour pré-formatée (ex. "25 mai 2026"), ou None.

    Returns:
        Chaîne complète prête à être passée comme message système.

    Raises:
        FileNotFoundError : si system.md est absent.
        OSError           : si system.md ne peut pas être lu.
    """
    # ── Prompt système de base ────────────────────────────────────────────────
    system_md_path = _os.path.join(_DATA_DIR, "system.md")
    if not _os.path.exists(system_md_path):
        raise FileNotFoundError(
            f"Fichier system.md introuvable : {system_md_path}\n"
            "L'IA ne peut pas démarrer sans ce fichier."
        )
    try:
        with open(system_md_path, encoding="utf-8") as file_handle:
            system_content = file_handle.read().strip()
    except OSError as exc:
        raise OSError(
            f"Impossible de lire system.md : {exc}\n"
            "L'IA ne peut pas démarrer sans ce fichier."
        ) from exc

    # ── Fichiers mémoire persistante ──────────────────────────────────────────
    _memory_label_map = {
        "memory": "MÉMOIRE (notes personnelles)",
        "user":   "PROFIL UTILISATEUR",
        "skills": "SKILLS (procédures apprises)",
    }

    for target in ("memory", "user", "skills"):
        raw = _read_memory_file(target)
        all_lines = raw.splitlines()
        body_lines = [line for line in all_lines if not line.startswith("<!--")]
        body = "\n".join(body_lines).strip()
        entries = [entry.strip() for entry in body.split("§") if entry.strip()]
        if not entries:
            continue
        char_limit = _MEMORY_CHAR_LIMITS[target]
        char_count = len(raw)
        pct = int(char_count * 100 / char_limit)
        label = _memory_label_map[target]
        separator = "═" * 46
        content_str = " §\n".join(entries)
        system_content += f"\n\n{separator}\n{label} [{pct}% — {char_count}/{char_limit} chars]\n{separator}\n{content_str}"

    # ── Date du jour ──────────────────────────────────────────────────────────
    if today_date_str:
        system_content += f"\n\nDate du jour : {today_date_str}."

    # ── Contexte du dossier ouvert ────────────────────────────────────────────
    if folder_path and _os.path.isdir(folder_path):
        folder_name = _os.path.basename(folder_path)
        system_content += (
            f"\n\nDOSSIER OUVERT : « {folder_name} » (`{folder_path}`).\n"
            "Outils disponibles pour ce dossier : list_folder_contents, "
            "read_file_content, create_file, edit_file, delete_files, move_file, "
            "copy_file, create_folder, organize_files, search_in_files, find_files, "
            "analyze_images, generate_image, edit_image, read_exif, "
            "zip_files, unzip_file.\n"
            "Utilise-les quand l'utilisateur te demande d'explorer, résumer, "
            "organiser ou analyser visuellement le contenu de ce dossier. "
            "Pour toute question sur ce que contiennent les images "
            "(couleurs, personnes, lieux, objets…), utilise analyze_images. "
            "Pour générer une nouvelle image depuis un prompt texte, utilise generate_image. "
            "Pour modifier une image existante du dossier, utilise edit_image. "
            "RÈGLES ABSOLUES :\n"
            "- Pour lister le contenu du dossier, utilise TOUJOURS "
            "list_folder_contents — JAMAIS ls, find ou toute autre commande shell via run_terminal_command.\n"
            "- Pour modifier un fichier texte existant, utilise TOUJOURS edit_file "
            "(lit → remplace old_string par new_string) — "
            "JAMAIS créer une copie .txt ou un nouveau fichier à la place de l'original. "
            "Si tu dois réécrire entièrement le fichier, utilise create_file sur le même chemin.\n"
            "- Pour créer un nouveau fichier (script, note, liste, config…), utilise TOUJOURS create_file — "
            "JAMAIS run_terminal_command avec une redirection (>, tee, etc.).\n"
            "- Pour chercher du texte dans des fichiers, utilise search_in_files (grep) — "
            "JAMAIS run_terminal_command avec grep ou rg.\n"
            "Le paramètre 'content' de create_file doit contenir UNIQUEMENT le texte final du fichier, "
            "recopié mot pour mot depuis les résultats des outils — "
            "sans aucun raisonnement, auto-correction, note entre parenthèses ou placeholder."
        )
    system_content += (
        "\n\nOutil terminal disponible : run_terminal_command. "
        "Utilise-le pour exécuter des commandes shell si l'utilisateur le demande "
        "(installation de paquets, conversion de fichiers, scripts, etc.). "
        "NE PAS l'utiliser pour lister des fichiers ou créer des fichiers texte : "
        "utilise list_folder_contents et create_file pour ça. "
        "NE PAS l'utiliser pour ssh/scp/sftp : utilise ssh_command, qui gère "
        "le mot de passe via overlay. "
        "Une confirmation sera toujours demandée avant exécution.\n\n"
        "IMPORTANT — ssh_command cible une machine DISTANTE (serveur SSH), "
        "un système de fichiers totalement séparé du DOSSIER OUVERT en local. "
        "Ne jamais utiliser les chemins du dossier ouvert dans une commande "
        "ssh_command, et ne jamais utiliser list_folder_contents/read_file_content/"
        "etc. pour explorer le serveur distant — tout passe par le paramètre "
        "'command' de ssh_command (ex: 'ls', 'cat fichier'). "
        "Serveurs SSH connus (nom → host) : "
        + (", ".join(f"{s['name']} → {s['host']}" for s in CONSTANTS.SSH_SERVERS)
           or "aucun enregistré") + "."
    )
    return system_content


# ─── Intégration Google Gemini ───────────────────────────────────────────────

def _format_gemini_error(exc, *, prefix="Erreur Gemini"):
    """Rend les erreurs Gemini plus lisibles sans masquer le message brut."""
    import re as _re_ge

    raw = str(exc).strip()
    compact = " ".join(raw.split())
    details = []

    if "429" in compact or "RESOURCE_EXHAUSTED" in compact:
        details.append("quota/rate limit Google atteint")
    elif "503" in compact or "UNAVAILABLE" in compact:
        details.append("service Google temporairement indisponible")

    retry_match = _re_ge.search(r'retryDelay[^0-9]*(\d+(?:\.\d+)?)', compact)
    if retry_match:
        details.append(f"retryDelay={retry_match.group(1)}s")

    if "GenerateContentResponse" in compact:
        compact = compact.split("GenerateContentResponse", 1)[0].strip(" :-")

    suffix = f" ({', '.join(details)})" if details else ""
    if compact:
        return f"[{prefix}{suffix} : {compact}]"
    return f"[{prefix}{suffix}]"


def _extract_gemini_feedback_messages(response):
    """Extrait les signaux de blocage/arrêt Gemini (prompt_feedback, safety, finish_reason)."""
    messages = []

    def _normalize_enum_name(value):
        raw = str(value).strip()
        if not raw:
            return ""
        if "." in raw:
            raw = raw.split(".")[-1]
        return raw.upper()

    prompt_feedback = getattr(response, "prompt_feedback", None)
    if prompt_feedback is not None:
        block_reason = getattr(prompt_feedback, "block_reason", None)
        block_msg = getattr(prompt_feedback, "block_reason_message", None) or getattr(prompt_feedback, "message", None)
        if block_reason and str(block_reason) not in {"BLOCK_REASON_UNSPECIFIED", "0"}:
            messages.append(f"Prompt bloqué ({block_reason})")
        if block_msg:
            messages.append(str(block_msg).strip())

    for cand in (getattr(response, "candidates", None) or []):
        finish_reason = getattr(cand, "finish_reason", None)
        finish_name = _normalize_enum_name(finish_reason) if finish_reason else ""
        if finish_name and finish_name not in {"STOP", "FINISH_REASON_UNSPECIFIED", "UNSPECIFIED", "0"}:
            messages.append(f"Arrêt modèle : {finish_reason}")

        for rating in (getattr(cand, "safety_ratings", None) or []):
            blocked = getattr(rating, "blocked", False)
            if blocked:
                category = getattr(rating, "category", "UNKNOWN")
                probability = getattr(rating, "probability", "UNKNOWN")
                messages.append(f"Blocage sécurité : {category} ({probability})")

    # Dédupliquer en conservant l'ordre
    deduped = []
    seen = set()
    for msg in messages:
        key = msg.strip()
        if key and key not in seen:
            deduped.append(key)
            seen.add(key)
    return deduped

def _ollama_tools_to_gemini(tools):
    """
    Convertit une liste de définitions d'outils au format Ollama (JSON Schema)
    en un seul objet types.Tool Gemini contenant les FunctionDeclaration.
    Retourne None si la liste est vide.
    """
    if not tools:
        return None
    try:
        from google.genai import types as _gtypes
    except ImportError:
        return None

    def _convert_schema(schema, _gtypes):
        """Convertit récursivement un JSON Schema dict en types.Schema Gemini."""
        raw_type = schema.get("type", "string").upper()
        prop_type = getattr(_gtypes.Type, raw_type, _gtypes.Type.STRING)
        kwargs = {
            "type": prop_type,
            "description": schema.get("description", ""),
        }
        if "enum" in schema:
            kwargs["enum"] = schema["enum"]
        # Tableaux : items obligatoire pour Gemini
        if prop_type == _gtypes.Type.ARRAY and "items" in schema:
            kwargs["items"] = _convert_schema(schema["items"], _gtypes)
        # Objets imbriqués
        if prop_type == _gtypes.Type.OBJECT and "properties" in schema:
            kwargs["properties"] = {
                k: _convert_schema(v, _gtypes)
                for k, v in schema["properties"].items()
            }
            if "required" in schema:
                kwargs["required"] = schema["required"]
        return _gtypes.Schema(**kwargs)

    declarations = []
    for tool in tools:
        fn = tool.get("function", {})
        params = fn.get("parameters", {})
        properties = {
            prop_name: _convert_schema(schema, _gtypes)
            for prop_name, schema in params.get("properties", {}).items()
        }
        gemini_params = _gtypes.Schema(
            type=_gtypes.Type.OBJECT,
            properties=properties,
            required=params.get("required", []),
        )
        declarations.append(_gtypes.FunctionDeclaration(
            name=fn.get("name", ""),
            description=fn.get("description", ""),
            parameters=gemini_params,
        ))
    return _gtypes.Tool(function_declarations=declarations)


def _ollama_messages_to_gemini(messages):
    """
    Convertit une liste de messages au format Ollama en :
      (system_instruction_str_or_None, [types.Content, ...])
    Gère : texte, images base64 (clé 'images'), PDFs base64 (clé 'pdfs'),
    tool_calls (role assistant) et résultats d'outils (role 'tool').
    """
    try:
        import base64 as _b64
        from google.genai import types as _gtypes
    except ImportError:
        return None, []

    system_instr = None
    gemini_contents = []
    _prev_role_was_tool = False

    for msg in messages:
        role = msg.get("role", "user")
        content = msg.get("content", "")

        if role == "system":
            system_instr = content if isinstance(content, str) else ""
            continue

        gemini_role = "model" if role == "assistant" else "user"
        parts = []

        # Contenu textuel
        if isinstance(content, str) and content:
            parts.append(_gtypes.Part(text=content))
        elif isinstance(content, list):
            for item in content:
                if not isinstance(item, dict):
                    continue
                if item.get("type") == "text" and item.get("text"):
                    parts.append(_gtypes.Part(text=item["text"]))
                elif item.get("type") == "image_url":
                    url = (item.get("image_url") or {}).get("url", "")
                    if url.startswith("data:") and "," in url:
                        header, b64_data = url.split(",", 1)
                        mime_type = header.split(":")[1].split(";")[0]
                        try:
                            parts.append(_gtypes.Part.from_bytes(
                                data=_b64.b64decode(b64_data),
                                mime_type=mime_type,
                            ))
                        except Exception:
                            pass

        # Images base64 (format Ollama : liste de chaînes b64 JPEG)
        for b64_img in msg.get("images", []):
            try:
                parts.append(_gtypes.Part.from_bytes(
                    data=_b64.b64decode(b64_img),
                    mime_type="image/jpeg",
                ))
            except Exception:
                pass

        # PDFs base64 (format étendu : clé 'pdfs', liste de chaînes b64)
        for b64_pdf in msg.get("pdfs", []):
            try:
                parts.append(_gtypes.Part.from_bytes(
                    data=_b64.b64decode(b64_pdf),
                    mime_type="application/pdf",
                ))
            except Exception:
                pass

        # Appels d'outils (rôle assistant)
        if role == "assistant":
            for tc in msg.get("tool_calls", []):
                fn = tc.get("function", {})
                args = fn.get("arguments", {})
                if isinstance(args, str):
                    try:
                        args = json.loads(args)
                    except Exception as exc:
                        _logger.warning(
                            "arguments de tool call (Gemini) illisibles, "
                            "remplacés par {} : %r — brut : %r", exc, args)
                        args = {}
                fc_part = _gtypes.Part.from_function_call(
                    name=fn.get("name", ""),
                    args=args,
                )
                # Rendre au tour suivant le thought_signature capté au moment
                # de l'appel (cf. _gemini_chat_stream_with_tools) — requis par
                # Gemini 3.x, sinon 400 INVALID_ARGUMENT sur les tours d'outils
                # suivants (surtout visible avec les outils MCP).
                sig_b64 = tc.get("thought_signature")
                if sig_b64:
                    try:
                        fc_part.thought_signature = _b64.b64decode(sig_b64)
                    except Exception:
                        pass
                parts.append(fc_part)

        # Résultats d'outils (rôle 'tool')
        if role == "tool":
            tool_name = msg.get("tool_name") or msg.get("name") or "tool_result"
            result_text = content if isinstance(content, str) else json.dumps(content)
            parts.append(_gtypes.Part.from_function_response(
                name=tool_name,
                response={"result": result_text},
            ))
            gemini_role = "user"

        if parts:
            # Regrouper les réponses d'outils consécutives (appels parallèles,
            # fréquent en modification de fichiers via MCP) dans UN SEUL tour
            # "user" : Gemini exige que toutes les function_response d'un
            # même tour de function_call arrivent ensemble, sinon l'alternance
            # user/model est rompue et le tour d'appel suivant échoue avec
            # 400 INVALID_ARGUMENT "function call turn comes immediately
            # after a user turn or after a function response turn".
            if role == "tool" and _prev_role_was_tool and gemini_contents:
                gemini_contents[-1].parts.extend(parts)
            else:
                gemini_contents.append(_gtypes.Content(role=gemini_role, parts=parts))

        _prev_role_was_tool = (role == "tool")

    return system_instr, gemini_contents


# ── Cache de contexte Gemini ─────────────────────────────────────────────────
# system.md + memoire/user/skills + les ~18 definitions d'outils sont
# quasi-identiques a chaque tour (~9000 tokens) et etaient jusqu'ici renvoyes
# et factures en entier a chaque message. Gemini permet de les mettre en
# cache cote serveur (cachedContent) : on ne paie le plein tarif que la
# premiere fois, les tours suivants relisent le cache pour une fraction du
# prix, tant que system_instruction/tools n'ont pas change.
_GEMINI_CACHE_REGISTRY = {}   # cache_key (sha256) -> (cache_name, expire_epoch)
_GEMINI_CACHE_TTL_SECONDS = 3600
_GEMINI_CACHE_MIN_CHARS = 4000  # sous ce seuil l'API refuse la mise en cache de toute facon


def _get_gemini_cached_content(client, model, system_instr, raw_tools):
    """
    Retourne le nom d'un cachedContent Gemini reutilisable pour (model,
    system_instr, raw_tools), ou None si la mise en cache n'est pas possible
    ou pas utile (contenu trop court, modele non supporte, erreur API…).
    Un nouveau cache est cree automatiquement des que le contenu change.
    """
    if not system_instr or len(system_instr) < _GEMINI_CACHE_MIN_CHARS:
        return None
    try:
        from google.genai import types as _gtypes
    except ImportError:
        return None

    cache_key = _hashlib.sha256(
        (model + "\0" + system_instr + "\0" + json.dumps(raw_tools or [], sort_keys=True)).encode("utf-8")
    ).hexdigest()

    now = _time.time()
    cached = _GEMINI_CACHE_REGISTRY.get(cache_key)
    if cached and cached[1] > now:
        return cached[0]

    gemini_tool = _ollama_tools_to_gemini(raw_tools)
    gemini_tools_list = [_gtypes.Tool(google_search=_gtypes.GoogleSearch())]
    if gemini_tool is not None:
        gemini_tools_list.append(gemini_tool)

    cache_config_kwargs = {
        "system_instruction": system_instr,
        "tools": gemini_tools_list,
        "ttl": f"{_GEMINI_CACHE_TTL_SECONDS}s",
    }
    # tool_config fait partie du contenu mis en cache au même titre que
    # system_instruction/tools : l'API refuse qu'on le refixe sur la requête
    # une fois cached_content utilisé (erreur 400 INVALID_ARGUMENT sinon).
    if gemini_tool is not None:
        cache_config_kwargs["tool_config"] = _gtypes.ToolConfig(
            include_server_side_tool_invocations=True
        )

    try:
        cache = client.caches.create(
            model=model,
            config=_gtypes.CreateCachedContentConfig(**cache_config_kwargs),
        )
    except Exception:
        # Modele non supporte, contenu encore trop court selon l'API, quota, etc.
        # On se rabat silencieusement sur l'ancien comportement (pas de cache).
        return None

    _GEMINI_CACHE_REGISTRY[cache_key] = (cache.name, now + _GEMINI_CACHE_TTL_SECONDS - 60)
    return cache.name


# ── Compactage de l'historique de conversation ───────────────────────────────
# La fenêtre glissante (AI_HISTORY_LIMIT_CLOUD) tronque déjà les tours trop
# anciens sans rien renvoyer à leur sujet — mais le modèle perd toute mémoire
# de ce qui précède. On remplace la troncature sèche par un résumé court et
# cumulatif des tours qui sortent de la fenêtre, injecté dans le message
# système à la place du texte brut (beaucoup moins de tokens que les tours
# originaux, tout en gardant le fil de la conversation).

def _summarize_turns(model, turns, previous_summary=""):
    """
    Condense une liste de messages {"role", "content"} qui sortent de la
    fenêtre d'historique en un court résumé, fusionné avec le résumé
    précédent s'il y en a un.

    Essaie Gemini d'abord (rapide, peu coûteux) ; si aucune clé n'est
    disponible ou que l'appel échoue (hors-ligne, quota…), se rabat sur un
    modèle Ollama local (Gemma — CONSTANTS.AI_GEMINI_FALLBACK, le même
    modèle utilisé pour le fallback normal du chat). Ne lève jamais
    d'exception : retourne le résumé précédent inchangé si les deux échouent.
    """
    turns_text = "\n\n".join(
        f"{turn.get('role', '?').upper()} : {turn.get('content', '')}"
        for turn in turns
        if turn.get("role") in ("user", "assistant") and isinstance(turn.get("content"), str)
    ).strip()
    if not turns_text:
        return previous_summary

    prompt = (
        "Résume en un paragraphe court (10 lignes maximum) les faits, décisions "
        "et éléments de contexte importants de cet échange, à conserver pour la "
        "suite de la conversation. Si un résumé précédent est fourni, mets-le à "
        "jour et fusionne-le — ne le duplique pas.\n\n"
    )
    if previous_summary:
        prompt += f"RÉSUMÉ PRÉCÉDENT :\n{previous_summary}\n\n"
    prompt += f"NOUVEL ÉCHANGE À RÉSUMER :\n{turns_text}"

    api_key = _get_gemini_api_key()
    if api_key:
        try:
            from google import genai as _genai
            from google.genai import types as _gtypes
            client = _genai.Client(api_key=api_key)
            response = client.models.generate_content(
                model=model,
                contents=[_gtypes.Content(role="user", parts=[_gtypes.Part(text=prompt)])],
                config=_gtypes.GenerateContentConfig(temperature=0.3),
            )
            summary = (response.text or "").strip()
            if summary:
                return summary
        except Exception:
            pass  # Gemini indisponible (hors-ligne, quota, clé invalide…) : on tente Gemma.

    try:
        message = _ollama_chat_once(
            CONSTANTS.AI_OLLAMA_URL, CONSTANTS.AI_GEMINI_FALLBACK,
            [{"role": "user", "content": prompt}],
        )
        summary = re.sub(r"<think>.*?</think>", "", message.get("content") or "", flags=re.DOTALL).strip()
        return summary or previous_summary
    except Exception:
        return previous_summary


def _compact_history_summary(ai_conversation, history_limit, state):
    """
    Met à jour et retourne le résumé cumulatif des tours qui sortent de la
    fenêtre d'historique récente (les `history_limit` derniers tours restent
    envoyés bruts par ailleurs — cette fonction ne fait que produire le texte
    de résumé à injecter dans le message système).

    `state` est un dict mutable {"summary": str, "summarized_count": int},
    persisté dans .ai_conversation.json par l'appelant (voir _ai_save_history/
    _ai_load_history) pour survivre à un redémarrage.

    Le résumé est produit par _summarize_turns : Gemini en priorité (rapide,
    peu coûteux), avec repli automatique sur Gemma en local (Ollama) si
    Gemini n'est pas joignable — utile aussi pour les modèles Ollama locaux
    eux-mêmes, qui souffrent d'un prompt long (démarrage/traitement plus
    lents) : un historique compacté leur fait autant, sinon plus, de bien
    que pour Gemini.
    """
    already = state.get("summarized_count", 0)
    overflow_end = len(ai_conversation) - history_limit
    if overflow_end > already:
        new_turns = ai_conversation[already:overflow_end]
        if new_turns:
            state["summary"] = _summarize_turns(
                CONSTANTS.AI_GEMINI_MODEL, new_turns, state.get("summary", "")
            )
        state["summarized_count"] = overflow_end

    return state.get("summary", "")


def _gemini_chat_stream_with_tools(model, messages, tools=None, temperature=0.7):
    """
    Version Gemini de _ollama_chat_stream_with_tools.
    Génère les mêmes tuples : ("token", str), ("thinking", str), ("tool_calls", list).

    tool_calls est une liste d'éléments compatibles format Ollama :
      [{"function": {"name": str, "arguments": dict}}, ...]

    Nécessite GEMINI_API_KEY dans les variables d'environnement.
    """
    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        yield ("token", (
            "[Erreur : google-genai n'est pas installé. "
            "Exécute : pip install google-genai>=1.55.0]"
        ))
        return

    api_key = _get_gemini_api_key()

    if not api_key:
        yield ("token", (
            "[Erreur : la variable d'environnement GEMINI_API_KEY n'est pas définie. "
            "Veuillez la configurer dans votre environnement (.zshrc, .bashrc) ou créer un fichier .env contenant GEMINI_API_KEY=votre_cle.]"
        ))
        return

    try:
        # Timeout HTTP (read) : si le flux reste figé au-delà de ce délai sans
        # renvoyer de chunk, l'appel lève une erreur au lieu de pendre à
        # l'infini (sinon la boucle agent bloque toute l'app). Les retries
        # ci-dessous reprennent alors le tour.
        client = _genai.Client(
            api_key=api_key,
            http_options=_gtypes.HttpOptions(
                timeout=CONSTANTS.AI_GEMINI_STREAM_TIMEOUT_MS),
        )
    except Exception as exc:
        yield ("token", f"[Erreur initialisation Gemini : {exc}]")
        return

    # web_search et fetch_url sont remplacés par google_search natif :
    # Gemini gère la recherche en interne sans émettre de function_call,
    # donc aucun round de tool n'est consommé pour les recherches web.
    _WEB_TOOL_NAMES = {"web_search", "fetch_url"}
    tools_sans_web = [
        tool for tool in (tools or [])
        if tool.get("function", {}).get("name") not in _WEB_TOOL_NAMES
    ]
    gemini_tool = _ollama_tools_to_gemini(tools_sans_web)
    # Google Search natif toujours présent (remplace DuckDuckGo)
    gemini_tools_list = [_gtypes.Tool(google_search=_gtypes.GoogleSearch())]
    if gemini_tool is not None:
        gemini_tools_list.append(gemini_tool)

    system_instr, gemini_contents = _ollama_messages_to_gemini(messages)

    # Cache de contexte : si system_instr + tools sont identiques à un appel
    # récent, on réutilise le cache serveur au lieu de renvoyer ~9000 tokens
    # en clair. cached_content est alors exclusif avec system_instruction/tools/
    # tool_config (l'API les tire du cache, elle-même configurée avec les mêmes
    # valeurs) — on ne les fixe donc sur la requête que si pas de cache.
    _cached_content_name = _get_gemini_cached_content(client, model, system_instr, tools_sans_web)

    config_kwargs: dict = {}
    if not model.startswith("gemini-3.5"):
        config_kwargs["temperature"] = temperature
    if _cached_content_name:
        config_kwargs["cached_content"] = _cached_content_name
    else:
        if system_instr:
            config_kwargs["system_instruction"] = system_instr
        config_kwargs["tools"] = gemini_tools_list
        # Requis quand on mélange un outil natif (google_search) et des function declarations :
        if gemini_tool is not None:
            config_kwargs["tool_config"] = _gtypes.ToolConfig(
                include_server_side_tool_invocations=True
            )
    # Activer la réflexion (thinking) pour les modèles compatibles (Gemini 2.5, 3.1, 3.5, etc.)
    config_kwargs["thinking_config"] = _gtypes.ThinkingConfig(include_thoughts=True)
    config = _gtypes.GenerateContentConfig(**config_kwargs)

    # Accumulateur pour les function calls fragmentés sur plusieurs chunks
    import time as _time
    import re as _re_retry
    import base64 as _b64_mod

    _MAX_RETRIES = 3

    for _attempt in range(_MAX_RETRIES + 1):
        pending_tool_calls = []
        emitted_feedback = set()
        _emitted_tokens = False
        _emitted_thinking = False
        try:
            for chunk in client.models.generate_content_stream(
                model=model,
                contents=gemini_contents,
                config=config,
            ):
                for _fb_msg in _extract_gemini_feedback_messages(chunk):
                    if _fb_msg not in emitted_feedback:
                        emitted_feedback.add(_fb_msg)
                        yield ("token", f"\n[Gemini] {_fb_msg}\n")
                        _emitted_tokens = True

                if not chunk.candidates:
                    continue
                for part in chunk.candidates[0].content.parts:
                    # Tokens de réflexion (thinking)
                    if getattr(part, "thought", False) and part.text:
                        _emitted_thinking = True
                        yield ("thinking", part.text)
                    # Appel d'outil
                    elif part.function_call is not None:
                        fc = part.function_call
                        args = dict(fc.args) if fc.args else {}
                        tc_entry = {
                            "function": {"name": fc.name, "arguments": args}
                        }
                        # Gemini 3.x exige que le thought_signature associé à
                        # chaque function_call soit renvoyé tel quel au tour
                        # suivant (functionCall part), sinon 400 INVALID_ARGUMENT
                        # "Function call is missing a thought_signature" dès
                        # qu'un outil (MCP notamment) est appelé — encodé en
                        # base64 ici pour rester JSON-safe dans l'historique.
                        sig = getattr(part, "thought_signature", None)
                        if sig:
                            tc_entry["thought_signature"] = _b64_mod.b64encode(
                                sig).decode("ascii")
                        pending_tool_calls.append(tc_entry)
                    # Token de texte
                    elif part.text:
                        _emitted_tokens = True
                        yield ("token", part.text)

            # MALFORMED_FUNCTION_CALL : appel d'outil corrompu côté modèle,
            # transitoire — une nouvelle tentative suffit généralement.
            # ponytail: pas de backoff dédié, réutilise le budget de retry existant
            if (not pending_tool_calls and _attempt < _MAX_RETRIES and
                    any("MALFORMED_FUNCTION_CALL" in m for m in emitted_feedback)):
                yield ("token", "\n[Gemini] Appel d'outil corrompu – nouvelle tentative…\n")
                continue

            # Émettre les tool_calls accumulés en une seule fois
            if pending_tool_calls:
                yield ("tool_calls", pending_tool_calls)
            elif not _emitted_tokens and not _emitted_thinking:
                # Silence complet : flux coupé, réponse bloquée ou erreur silencieuse
                yield ("token", "*(Gemini n'a produit aucune réponse — réponse bloquée ou flux interrompu. Essaie de reformuler ou renvoie le message.)*")
            break  # succès

        except Exception as exc:
            _exc_str = str(exc)
            if ("429" in _exc_str or "RESOURCE_EXHAUSTED" in _exc_str) and _attempt < _MAX_RETRIES:
                _match = _re_retry.search(r"retryDelay[^0-9]*(\d+)", _exc_str)
                _delay = int(_match.group(1)) + 2 if _match else 62
                _retry_msg = _format_gemini_error(exc, prefix="Erreur Gemini")
                yield ("token", f"\n[{_retry_msg[1:-1]} – nouvelle tentative dans {_delay}s…]\n")
                _time.sleep(_delay)
            elif "503" in _exc_str or "UNAVAILABLE" in _exc_str:
                raise  # ponytail: laisser remonter → chaîne fallback Dashboard/SidePanel prend le relais
            elif _is_network_error(exc) and _attempt < _MAX_RETRIES:
                _delay = 5 * (_attempt + 1)
                yield ("token", f"\n[Connexion perdue avec Gemini – nouvelle tentative dans {_delay}s…]\n")
                _time.sleep(_delay)
            else:
                yield ("token", f"\n{_format_gemini_error(exc)}")
                break


def _gemini_generate_image(prompt, input_image_bytes=None, aspect_ratio="1:1", resolution="1K"):
    """
    Génère ou modifie une image avec Nano Banana 2.
    Utilise gemini-3.1-flash-image-preview avec bascule automatique (fallback) vers gemini-3.1-flash-image si surchargé ou indisponible.

    - prompt            : description textuelle de l'image souhaitée
    - input_image_bytes : bytes de l'image source pour édition (None = génération pure)
    - aspect_ratio      : format de sortie, ex. "1:1", "16:9", "3:2" …
    - resolution        : résolution : "512", "1K", "2K", "4K"

    Retourne (text_response: str, image_bytes: bytes | None).
    """
    try:
        from google import genai as _genai_img
        from google.genai import types as _gtypes_img
        import os as _os_img
        import io as _io_img
    except ImportError:
        return ("Erreur : bibliothèque google-genai non installée.", None)

    _api_key = _get_gemini_api_key()

    if not _api_key:
        return (
            "Erreur : la variable d'environnement GEMINI_API_KEY n'est pas définie. "
            "Veuillez la configurer dans votre environnement (.zshrc, .bashrc) ou créer un fichier .env contenant GEMINI_API_KEY=votre_cle.",
            None
        )

    import time as _time_gi
    import re as _re_gi

    _candidate_models = [
        "gemini-3.1-flash-image-preview",
        "gemini-3.1-flash-image",
    ]
    _last_error = None

    for _model in _candidate_models:
        _MAX_RETRIES_GI = 1  # Retries courts par modèle pour éviter d'attendre indéfiniment
        for _attempt_gi in range(_MAX_RETRIES_GI + 1):
            try:
                _client = _genai_img.Client(api_key=_api_key)

                # Construire le contenu : [prompt texte, image source optionnelle]
                # On évite d'envoyer response_format, qui peut provoquer l'erreur
                # "Extra inputs are not permitted" selon la version du SDK/API.
                _prompt_with_constraints = prompt
                if aspect_ratio != "1:1" or resolution != "1K":
                    _prompt_with_constraints += (
                        "\n\nContraintes de sortie :"
                        f" ratio {aspect_ratio}, taille {resolution}."
                        " Respecte ces contraintes si possible."
                    )

                _contents = [_prompt_with_constraints]
                if input_image_bytes:
                    try:
                        from PIL import Image as _PILImg, ImageOps as _PILOps
                        _pil = _PILImg.open(_io_img.BytesIO(input_image_bytes))
                        # Transposer EXIF et convertir en RGB pour supprimer la transparence RGBA potentiellement rejetée
                        _pil = _PILOps.exif_transpose(_pil).convert("RGB")
                        # Redimensionnement préventif au format 4K max pour préserver la qualité maximale
                        _pil.thumbnail((4096, 4096), _PILImg.Resampling.LANCZOS)
                        _buf = _io_img.BytesIO()
                        _pil.save(_buf, format="JPEG", quality=100, optimize=True)
                        _compressed_bytes = _buf.getvalue()
                        _contents.append(
                            _gtypes_img.Part.from_bytes(data=_compressed_bytes, mime_type="image/jpeg")
                        )
                    except Exception:
                        _contents.append(
                            _gtypes_img.Part.from_bytes(data=input_image_bytes, mime_type="image/jpeg")
                        )

                # Config minimale et robuste pour compatibilité multi-versions.
                _cfg_kwargs: dict = {"response_modalities": ["TEXT", "IMAGE"]}

                _response = _client.models.generate_content(
                    model=_model,
                    contents=_contents,
                    config=_gtypes_img.GenerateContentConfig(**_cfg_kwargs),
                )

                _text_out = ""
                _image_out = None
                _feedback_messages = _extract_gemini_feedback_messages(_response)

                # Les SDK Gemini peuvent exposer les parts à différents niveaux
                # (_response.parts ou _response.candidates[*].content.parts).
                _parts = []
                try:
                    if getattr(_response, "parts", None):
                        _parts.extend(_response.parts)
                except Exception:
                    pass
                if not _parts:
                    try:
                        for _cand in (getattr(_response, "candidates", None) or []):
                            _content = getattr(_cand, "content", None)
                            _cand_parts = getattr(_content, "parts", None) if _content is not None else None
                            if _cand_parts:
                                _parts.extend(_cand_parts)
                    except Exception:
                        pass

                for _part in _parts:
                    if getattr(_part, "thought", False):
                        continue
                    _part_text = getattr(_part, "text", None)
                    if _part_text:
                        _text_out += _part_text
                        continue
                    _inline = getattr(_part, "inline_data", None)
                    if _inline is not None and getattr(_inline, "data", None):
                        _image_out = _inline.data  # bytes PNG/JPEG

                if _feedback_messages:
                    _feedback_text = "\n".join(f"[Gemini] {_m}" for _m in _feedback_messages)
                    if _text_out:
                        _text_out = (_text_out + "\n\n" + _feedback_text).strip()
                    elif _image_out is None:
                        _text_out = _feedback_text

                return (_text_out.strip(), _image_out)

            except Exception as _exc:
                _exc_str = str(_exc)
                _last_error = _exc
                # Retry automatique très rapide sur quota 429 ou erreur 503
                if ("429" in _exc_str or "503" in _exc_str or "UNAVAILABLE" in _exc_str) and _attempt_gi < _MAX_RETRIES_GI:
                    _delay_gi = 2.0
                    _m_gi = _re_gi.search(r'"retryDelay":\s*"(\d+(?:\.\d+)?)s"', _exc_str)
                    if _m_gi:
                        _delay_gi = float(_m_gi.group(1)) + 1.0
                    _time_gi.sleep(_delay_gi)
                    continue
                # Si erreur irrémédiable ou épuisement des essais, on tente le modèle suivant immédiatement
                break

    return (_format_gemini_error(_last_error, prefix="ERREUR Gemini Image"), None)


def _gemini_refine_image_prompt(
    intent_prompt,
    user_request="",
    mode="edit_image",
    source_filename="",
    model="gemini-3.5-flash",
):
    """
    Raffine un prompt d'intention en prompt d'édition/génération très précis
    pour Nano Banana 2.

    Retourne toujours une chaîne exploitable : en cas d'échec, renvoie intent_prompt.
    """
    intent_prompt = (intent_prompt or "").strip()
    if not intent_prompt:
        return ""

    _mode_label = "édition" if mode == "edit_image" else "génération"
    _source_line = f"- Fichier source: {source_filename}\n" if source_filename else ""
    _user_req_block = user_request.strip() if user_request else "(non fourni)"

    _instruction = (
        "Tu es directeur artistique et expert en prompt engineering pour un "
        "modèle image-to-image / text-to-image (Nano Banana 2).\n"
        "Ta mission : transformer l'intention en UN prompt d'image ultra précis, "
        "concret et directement exécutable.\n\n"
        "Couvre systématiquement, en une formulation fluide (pas une liste) :\n"
        "1. SUJET principal + détails clés\n"
        "2. STYLE / medium (photo réaliste, aquarelle, 3D, illustration…)\n"
        "3. LUMIÈRE & ambiance (direction, dureté, heure, humeur)\n"
        "4. COMPOSITION / cadrage / angle\n"
        "5. PALETTE de couleurs dominante\n"
        "6. NIVEAU DE DÉTAIL & rendu (netteté, texture, qualité)\n\n"
        "Règles strictes :\n"
        "- Réponds UNIQUEMENT avec le prompt final : pas de commentaire, pas de "
        "Markdown, pas de guillemets, pas de préambule.\n"
        "- Interdit les mots vagues ('améliore', 'plus beau', 'sympa') sans "
        "critère visuel concret.\n"
        "- Pour une ÉDITION : préserver composition, proportions, cadrage, "
        "perspective et identité du sujet, sauf demande explicite de les changer ; "
        "ne décrire QUE ce qui change et ce qui doit rester intact.\n"
        "- Intention ambiguë → interprétation la plus sûre et conservatrice.\n"
        "- Écrire en français.\n\n"
        "Exemple —\n"
        "Intention : un chat mignon\n"
        "Prompt : Portrait rapproché d'un chaton roux tigré aux grands yeux "
        "verts, fourrure duveteuse très détaillée, assis sur un plaid en laine "
        "crème ; lumière douce de fin d'après-midi venant de la gauche, ambiance "
        "chaleureuse ; cadrage serré légèrement en plongée, faible profondeur de "
        "champ, arrière-plan flou ; palette chaude ocre et crème ; rendu "
        "photoréaliste net, textures fines.\n\n"
        f"Mode : {_mode_label}\n"
        f"{_source_line}"
        f"Demande utilisateur originale :\n{_user_req_block}\n\n"
        f"Intention brute à raffiner :\n{intent_prompt}\n"
    )

    result, _ = _gemini_interactions_create(_instruction, model=model)
    return result.strip() if result.strip() else intent_prompt


def _gemini_critique_image(image_bytes, goal, model="gemini-3.5-flash"):
    """
    Critique visuelle d'une image par rapport à un objectif.
    Retourne "OK" si l'image atteint clairement l'objectif, sinon une liste
    courte et actionnable de ce qu'il faut corriger.

    Ne lève jamais : en l'absence de vision/clé/erreur, retourne "OK" pour ne
    pas déclencher de régénération inutile (arrêt sûr de la boucle).
    """
    try:
        from google import genai as _genai_c
        from google.genai import types as _gt_c
        import io as _io_c
    except ImportError:
        return "OK"
    api_key = _get_gemini_api_key()
    if not api_key:
        return "OK"
    _instr = (
        "Tu es directeur artistique. Voici une image et un objectif.\n"
        f"OBJECTIF : {goal}\n\n"
        "Si l'image atteint clairement l'objectif, réponds EXACTEMENT « OK » "
        "(rien d'autre). Sinon, liste en 1 à 4 puces courtes et ACTIONNABLES "
        "les défauts visuels concrets à corriger pour l'atteindre (pas de "
        "généralités, pas de compliments). Écris en français."
    )
    try:
        client = _genai_c.Client(api_key=api_key)
        try:
            from PIL import Image as _PILc, ImageOps as _POc
            _pil = _POc.exif_transpose(
                _PILc.open(_io_c.BytesIO(image_bytes))
            ).convert("RGB")
            _pil.thumbnail((1536, 1536), _PILc.Resampling.LANCZOS)
            _buf = _io_c.BytesIO()
            _pil.save(_buf, format="JPEG", quality=85)
            _img_part = _gt_c.Part.from_bytes(
                data=_buf.getvalue(), mime_type="image/jpeg"
            )
        except Exception:
            _img_part = _gt_c.Part.from_bytes(
                data=image_bytes, mime_type="image/jpeg"
            )
        resp = client.models.generate_content(
            model=model, contents=[_instr, _img_part]
        )
        return (getattr(resp, "text", "") or "OK").strip() or "OK"
    except Exception:
        return "OK"


def _iterate_image_loop(source_path, goal, max_passes,
                        refiner_model="gemini-3.5-flash",
                        aspect_ratio="1:1", resolution="1K"):
    """
    Améliore itérativement une image jusqu'à atteindre `goal`.

    À chaque passe : critique visuelle du rendu courant (vision) ; si conforme
    ("OK"), arrêt anticipé ; sinon on raffine un prompt d'édition (objectif +
    critique) et on régénère via Nano Banana. Chaque version est sauvée en
    nouveau fichier « <base>_iterN.<ext> » à côté de la source.

    Retourne un dict, sans jamais lever :
      {"final_path": str|None,
       "passes": [ {"pass": int, "path": str, "critique": str, "ok": bool,
                    "prompt": str, "error": str|None} ],
       "error": str|None}
    """
    if not source_path or not _os.path.isfile(source_path):
        return {"final_path": None, "passes": [],
                "error": f"Image introuvable : {source_path}"}
    try:
        with open(source_path, "rb") as _f:
            cur_bytes = _f.read()
    except Exception as exc:
        return {"final_path": None, "passes": [],
                "error": f"Lecture impossible : {exc}"}

    folder = _os.path.dirname(source_path)
    base, ext = _os.path.splitext(_os.path.basename(source_path))
    ext = ext or ".png"
    passes = []
    cur_path = source_path

    for i in range(1, max(1, int(max_passes)) + 1):
        critique = _gemini_critique_image(cur_bytes, goal, model=refiner_model)
        if critique.strip().upper().startswith("OK"):
            passes.append({"pass": i, "path": cur_path, "critique": "OK",
                           "ok": True, "prompt": "", "error": None})
            break
        intent = (
            f"Objectif visuel : {goal}\n"
            f"Défauts à corriger sur l'image actuelle : {critique}"
        )
        edit_prompt = _gemini_refine_image_prompt(
            intent_prompt=intent, user_request=goal, mode="edit_image",
            source_filename=_os.path.basename(cur_path), model=refiner_model,
        )
        text, new_bytes = _gemini_generate_image(
            edit_prompt, input_image_bytes=cur_bytes,
            aspect_ratio=aspect_ratio, resolution=resolution,
        )
        if not new_bytes:
            passes.append({"pass": i, "path": cur_path, "critique": critique,
                           "ok": False, "prompt": edit_prompt, "error": text})
            return {"final_path": cur_path, "passes": passes,
                    "error": f"Passe {i} : aucune image générée ({text})"}
        new_path = _os.path.join(folder, f"{base}_iter{i}{ext}")
        try:
            with open(new_path, "wb") as _fout:
                _fout.write(new_bytes)
        except Exception as exc:
            return {"final_path": cur_path, "passes": passes,
                    "error": f"Écriture impossible : {exc}"}
        passes.append({"pass": i, "path": new_path, "critique": critique,
                       "ok": False, "prompt": edit_prompt, "error": None})
        cur_bytes, cur_path = new_bytes, new_path

    return {"final_path": cur_path, "passes": passes, "error": None}


# Outil autonome (annoncé uniquement par Dashboard, qui possède le dispatch).
_IMAGE_ITERATE_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "iterate_image",
            "description": (
                "Améliore ITÉRATIVEMENT une image existante du dossier ouvert "
                "jusqu'à atteindre un objectif : à chaque passe l'image est "
                "évaluée visuellement puis régénérée pour corriger ses défauts, "
                "avec arrêt anticipé dès que l'objectif est atteint. Utilise cet "
                "outil UNIQUEMENT quand l'utilisateur demande explicitement "
                "d'itérer / améliorer / peaufiner une image sur plusieurs passes "
                "(« jusqu'à ce que ce soit bon »). Pour une seule modification, "
                "utilise edit_image."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "source_filename": {
                        "type": "string",
                        "description": "Image source à améliorer (dans le dossier ouvert).",
                    },
                    "goal": {
                        "type": "string",
                        "description": (
                            "Objectif visuel précis à atteindre — ce qui doit "
                            "être vrai de l'image finale (concret, pas vague)."
                        ),
                    },
                    "passes": {
                        "type": "integer",
                        "description": (
                            "Nombre max de passes. Arrêt anticipé si l'objectif "
                            "est atteint. Laisser vide pour la valeur par défaut."
                        ),
                    },
                },
                "required": ["source_filename", "goal"],
            },
        },
    }
]


# ─── Interactions API ─────────────────────────────────────────────────────────

def _gemini_interactions_create(
    input_text,
    model="gemini-3.5-flash",
    previous_interaction_id=None,
    system_instruction=None,
    background=False,
    agent=None,
):
    """
    Appel simple Interactions API Gemini (non-streaming, sans outils).
    Supporte les conversations à états via previous_interaction_id.

    Agents disponibles :
      "deep-research-preview-04-2026"     — Deep Research rapide
      "deep-research-max-preview-04-2026" — Deep Research exhaustif
      "antigravity-preview-05-2026"       — Agent généraliste (code, web, fichiers)

    Les agents nécessitent background=True. Utiliser _gemini_interactions_get()
    pour récupérer le résultat.

    Retourne (output_text: str, interaction_id: str | None).
    """
    try:
        from google import genai as _genai_ia
    except ImportError:
        return ("", None)

    api_key = _get_gemini_api_key()
    if not api_key:
        return ("", None)

    try:
        client = _genai_ia.Client(api_key=api_key)
    except Exception:
        return ("", None)

    kwargs = {"input": input_text}
    if agent:
        kwargs["agent"] = agent
    else:
        kwargs["model"] = model
    if previous_interaction_id:
        kwargs["previous_interaction_id"] = previous_interaction_id
    if system_instruction:
        kwargs["system_instruction"] = system_instruction
    if background:
        kwargs["background"] = True

    try:
        interaction = client.interactions.create(**kwargs)
        return (interaction.output_text or "", interaction.id)
    except Exception as exc:
        return (_format_gemini_error(exc), None)


def _gemini_interactions_get(interaction_id):
    """
    Récupère l'état d'une Interaction (polling pour background=True).
    Retourne (status: str, output_text: str, error: str | None).
    status ∈ {"completed", "failed", "in_progress", "unknown"}
    """
    try:
        from google import genai as _genai_ig
    except ImportError:
        return ("unknown", "", "google-genai not installed")

    api_key = _get_gemini_api_key()
    if not api_key:
        return ("unknown", "", "GEMINI_API_KEY not set")

    try:
        client = _genai_ig.Client(api_key=api_key)
        interaction = client.interactions.get(interaction_id)
        status = getattr(interaction, "status", "unknown")
        text = getattr(interaction, "output_text", "") or ""
        return (status, text, None)
    except Exception as exc:
        return ("unknown", "", str(exc))


# ─── Musique — Lyria ──────────────────────────────────────────────────────────

def _gemini_generate_music(prompt, model="lyria-3-clip-preview", images=None):
    """
    Génère de la musique via Lyria 3 (Interactions API, google-genai >= 2.9.0).
    Modèles : "lyria-3-clip-preview" (30 s, 48 kHz stéréo MP3)
              "lyria-3-pro-preview"  (~2 min, structuré)
    images   : liste de chemins d'images (PIL requis) pour input multimodal.
    Retourne (audio_bytes: bytes | None, lyrics: str | None, error: str | None).
    Le 2e bloc texte (métadonnées JSON : BPM, clé, structure) est ignoré.
    """
    try:
        from google import genai as _genai_m
    except ImportError:
        return (None, None, "google-genai non installé")

    api_key = _get_gemini_api_key()
    if not api_key:
        return (None, None, "GEMINI_API_KEY non défini")

    try:
        if images:
            from PIL import Image as _PIL_Image
            input_content = [prompt] + [_PIL_Image.open(p) for p in images]
        else:
            input_content = prompt

        client = _genai_m.Client(api_key=api_key)
        interaction = client.interactions.create(
            model=model,
            input=input_content,
            response_modalities=["audio", "text"],
        )

        import base64
        text_parts = []
        audio_bytes = None

        for step in (getattr(interaction, "steps", None) or []):
            stype = getattr(step, "type", None)
            if stype == "text":
                t = getattr(step, "text", None)
                if t:
                    text_parts.append(t)
            elif stype == "audio":
                raw = getattr(step, "data", None)
                if raw and audio_bytes is None:
                    audio_bytes = base64.b64decode(raw)
            elif stype == "model_output":
                for block in (getattr(step, "content", None) or []):
                    if getattr(block, "text", None):
                        text_parts.append(block.text)
                    elif getattr(block, "type", None) == "audio":
                        raw = getattr(block, "data", None)
                        if raw and audio_bytes is None:
                            audio_bytes = base64.b64decode(raw)

        # Fallback inline_data (bytes directs, pas base64)
        if audio_bytes is None:
            for step in (getattr(interaction, "steps", None) or []):
                inline = getattr(step, "inline_data", None)
                if inline:
                    audio_bytes = inline.data
                    break

        lyrics = text_parts[0] if text_parts else None

        if audio_bytes is None:
            return (None, lyrics, "Aucune donnée audio dans la réponse")

        return (audio_bytes, lyrics, None)
    except Exception as exc:
        return (None, None, _format_gemini_error(exc))


# ─── Voix — TTS ───────────────────────────────────────────────────────────────

def _gemini_tts(text, voice_name="Puck", tts_model="gemini-3.1-flash-tts-preview", language_code=None):
    """
    Génère de l'audio TTS via l'API Gemini.

    Retourne les bytes PCM bruts (int16, mono, 24 kHz) ou None en cas d'erreur.
    Lève ImportError si google-genai n'est pas installé.
    """
    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        return None

    api_key = _get_gemini_api_key()
    if not api_key:
        return None

    try:
        client = _genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model=tts_model,
            contents=text,
            config=_gtypes.GenerateContentConfig(
                response_modalities=["AUDIO"],
                speech_config=_gtypes.SpeechConfig(
                    voice_config=_gtypes.VoiceConfig(
                        prebuilt_voice_config=_gtypes.PrebuiltVoiceConfig(
                            voice_name=voice_name,
                        )
                    ),
                    language_code=language_code,
                ),
            ),
        )
        return response.candidates[0].content.parts[0].inline_data.data
    except Exception:
        return None


def _gemini_tts_stream(text, voice_name="Puck", tts_model="gemini-3.1-flash-tts-preview", sample_rate=24000, language_code=None, stop_event=None, preroll_ms=0):
    """
    Génère et joue le TTS via l'API Gemini en streaming audio réel :
    une seule requête ``generate_content_stream``, lecture dès le premier
    chunk audio reçu (comme le mode "live", mais texte lu tel quel — pas
    de session conversationnelle qui reformule).

    Remplace l'ancien découpage du texte en phrases + plusieurs requêtes
    ``generate_content`` complètes : ça n'apportait aucun gain de latence
    pour une réponse courte (un seul "chunk" = exactement le même coût
    qu'un appel unique, retour user), et le streaming réel est de toute
    façon plus rapide même pour les textes longs.

    La réception réseau (producteur) et l'écriture audio (consommateur)
    tournent dans deux threads séparés, avec un petit coussin de
    pré-lecture (``preroll_ms``) avant la première écriture : sans ça, une
    pause de génération entre deux phrases (gigue réseau) affamait
    directement le flux de sortie et produisait un accroc audible entre
    les phrases (retour user) — même mécanisme que le mode "live".

    language_code (ex : "fr", "en") : accent de la synthèse.
    stop_event (threading.Event) : si activé, interrompt immédiatement
    la lecture et la génération en cours (barge-in).
    Bloquant : attend la fin de la lecture avant de retourner.
    """
    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        raise ImportError("google-genai n'est pas installé")

    api_key = _get_gemini_api_key()
    if not api_key:
        raise RuntimeError("Clé API Gemini introuvable — vérifier GEMINI_API_KEY")

    import queue as _queue
    import threading as _threading
    import sounddevice as _sd

    client = _genai.Client(api_key=api_key)
    config = _gtypes.GenerateContentConfig(
        response_modalities=["AUDIO"],
        speech_config=_gtypes.SpeechConfig(
            voice_config=_gtypes.VoiceConfig(
                prebuilt_voice_config=_gtypes.PrebuiltVoiceConfig(
                    voice_name=voice_name,
                )
            ),
            language_code=language_code,
        ),
    )

    # Découpe en morceaux de ~600 caractères aux frontières de phrases
    # (retour user : les réponses longues coupaient net après ~1/3 de
    # l'audio en un seul appel generate_content_stream — plafond côté
    # serveur sur la durée générée par requête, propre au streaming de ce
    # modèle TTS preview, absent en non-streaming). Plus long que l'ancien
    # découpage par phrases (~300 car., qui pipelinait des appels bloquants
    # un par un) : moins de requêtes, tout en restant sous ce plafond.
    import re as _re
    raw_sentences = _re.split(r'(?<=[.!?])\s+', text.strip())
    text_chunks: list = []
    current_chunk = ""
    for sentence in raw_sentences:
        if not sentence.strip():
            continue
        candidate = (current_chunk + " " + sentence).strip() if current_chunk else sentence
        if current_chunk and len(candidate) > 600:
            text_chunks.append(current_chunk)
            current_chunk = sentence
        else:
            current_chunk = candidate
    if current_chunk:
        text_chunks.append(current_chunk)
    if not text_chunks:
        text_chunks = [text]

    _SENTINEL = object()
    audio_queue: _queue.Queue = _queue.Queue(maxsize=40)
    error_holder: list = [None]

    def _producer():
        try:
            for text_chunk in text_chunks:
                if stop_event is not None and stop_event.is_set():
                    break
                for chunk in client.models.generate_content_stream(
                    model=tts_model, contents=text_chunk, config=config,
                ):
                    if stop_event is not None and stop_event.is_set():
                        break
                    if not chunk.candidates:
                        continue
                    for part in chunk.candidates[0].content.parts:
                        inline = getattr(part, "inline_data", None)
                        if inline and getattr(inline, "data", None):
                            audio_queue.put(inline.data)
        except Exception as exc:
            error_holder[0] = exc
        finally:
            audio_queue.put(_SENTINEL)

    _threading.Thread(target=_producer, daemon=True).start()

    output_stream = _sd.RawOutputStream(
        samplerate=sample_rate, channels=1, dtype="int16",
        blocksize=1024, latency="low")
    output_stream.start()
    got_audio = False
    preroll_bytes = int(sample_rate * 2 * max(preroll_ms, 0) / 1000.0)
    prebuf = bytearray()
    started = preroll_bytes <= 0
    try:
        while True:
            if stop_event is not None and stop_event.is_set():
                break
            try:
                item = audio_queue.get(timeout=0.1)
            except Exception:
                continue
            if item is _SENTINEL:
                if error_holder[0] is not None:
                    raise error_holder[0]
                if not got_audio and (stop_event is None or not stop_event.is_set()):
                    raise RuntimeError(
                        f"Aucun audio généré — vérifier la clé API et le modèle TTS ({tts_model})"
                    )
                if prebuf and (stop_event is None or not stop_event.is_set()):
                    output_stream.write(bytes(prebuf))
                break
            got_audio = True
            if started:
                output_stream.write(item)
            else:
                prebuf += item
                if len(prebuf) >= preroll_bytes:
                    output_stream.write(bytes(prebuf))
                    prebuf = bytearray()
                    started = True
    finally:
        output_stream.stop()
        output_stream.close()


def _gemini_live_tts_stream(
    text,
    model="gemini-3.1-flash-live-preview",
    voice_name="Kore",
    sample_rate=24000,
    language_code=None,
    stop_event=None,
    preroll_ms=0,
):
    """
    Génère et joue le TTS via l'API Gemini Live (connexion WebSocket persistante).

    Contrairement à _gemini_tts_stream qui enchaîne plusieurs requêtes indépendantes,
    cette fonction utilise une seule session Live : la voix est parfaitement cohérente
    du début à la fin (même intonation, même timbre, pas de rupture entre les phrases).

    text           : texte complet (str) OU queue.Queue de fragments de texte pour
                     une lecture incrémentale — pousser les morceaux au fil de la
                     génération, puis None pour signaler la fin. Même session Live
                     dans les deux cas : voix et langue restent verrouillées.
    model          : modèle Gemini Live (ex. "gemini-3.5-flash-live")
    voice_name     : voix Gemini (ex. "Kore", "Puck" — mêmes noms que le TTS classique)
    sample_rate    : fréquence de sortie PCM (24000 Hz par défaut)
    language_code  : code langue ISO 639-1 (ex. "fr") ou None pour auto-détection
    stop_event     : threading.Event — si activé, interrompt la lecture (barge-in)
    preroll_ms     : coussin audio accumulé avant de démarrer la lecture (jitter
                     buffer) — lisse les à-coups réseau au prix d'un léger retard.

    Bloquant : attend la fin de la lecture avant de retourner.
    Lève ImportError si google-genai n'est pas installé ou si le modèle n'est pas disponible.
    """
    import asyncio as _asyncio
    import queue as _queue
    import threading as _threading

    text_is_stream = isinstance(text, _queue.Queue)

    api_key = _get_gemini_api_key()
    if not api_key:
        raise RuntimeError("Clé API Gemini introuvable — vérifier GEMINI_API_KEY")

    audio_queue: _queue.Queue = _queue.Queue(maxsize=40)
    _SENTINEL = object()
    error_holder: list = [None]

    async def _run_live():
        try:
            from google import genai as _genai
            from google.genai import types as _gtypes

            client = _genai.Client(api_key=api_key)
            live_config = _gtypes.LiveConnectConfig(
                response_modalities=["AUDIO"],
                speech_config=_gtypes.SpeechConfig(
                    voice_config=_gtypes.VoiceConfig(
                        prebuilt_voice_config=_gtypes.PrebuiltVoiceConfig(
                            voice_name=voice_name,
                        )
                    ),
                    language_code=language_code,
                ),
            )
            async with client.aio.live.connect(model=model, config=live_config) as session:
                # Gemini 3.1 Live: pendant la conversation, envoyer le texte via
                # send_realtime_input (send_client_content sert surtout à l'amorçage
                # d'historique initial avec history_config dédié).
                feeder = None
                if text_is_stream:
                    # Mode incrémental : draine la file de fragments et les envoie
                    # au fil de l'eau, sans bloquer la boucle d'événements.
                    _loop = _asyncio.get_event_loop()

                    async def _feed():
                        while True:
                            frag = await _loop.run_in_executor(None, text.get)
                            if frag is None:
                                break
                            await session.send_realtime_input(text=frag)

                    feeder = _asyncio.create_task(_feed())
                else:
                    await session.send_realtime_input(text=text)
                async for message in session.receive():
                    if stop_event is not None and stop_event.is_set():
                        break
                    # Source officielle (Get_started_LiveAPI.py) : les données audio
                    # sont dans server_content.model_turn.parts[n].inline_data.data
                    server_content = getattr(message, "server_content", None)
                    if server_content:
                        model_turn = getattr(server_content, "model_turn", None)
                        if model_turn:
                            for part in getattr(model_turn, "parts", []):
                                inline = getattr(part, "inline_data", None)
                                if inline and getattr(inline, "data", None):
                                    audio_queue.put(inline.data)
                        if getattr(server_content, "turn_complete", False):
                            # En mode incrémental, un turn_complete peut tomber
                            # entre deux fragments : ne s'arrêter qu'une fois tout
                            # le texte envoyé (feeder terminé).
                            if feeder is None or feeder.done():
                                break
                    # Fallback : certaines versions SDK exposent .data directement
                    elif getattr(message, "data", None):
                        audio_queue.put(message.data)
                if feeder is not None and not feeder.done():
                    feeder.cancel()
        except Exception as exc:
            error_holder[0] = exc
        finally:
            audio_queue.put(_SENTINEL)

    def _producer():
        loop = _asyncio.new_event_loop()
        _asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(_run_live())
        finally:
            loop.close()

    _threading.Thread(target=_producer, daemon=True).start()

    # Lecture en streaming : on ouvre UN seul OutputStream sounddevice et on y
    # écrit chaque chunk dès qu'il arrive — pas de latence d'accumulation, pas de
    # saccades (un seul stream ouvert en continu, comme le cookbook officiel Google).
    import sounddevice as _sd_tts

    got_audio = False
    output_stream = _sd_tts.RawOutputStream(
        samplerate=sample_rate,
        channels=1,
        dtype="int16",
        blocksize=1024,
        latency="low",
    )
    output_stream.start()
    # Coussin de pré-lecture : on accumule preroll_ms d'audio avant la 1re
    # écriture pour absorber la gigue réseau (int16 mono → 2 octets/échantillon).
    preroll_bytes = int(sample_rate * 2 * max(preroll_ms, 0) / 1000.0)
    prebuf = bytearray()
    started = preroll_bytes <= 0
    try:
        while True:
            if stop_event is not None and stop_event.is_set():
                break
            try:
                item = audio_queue.get(timeout=0.1)
            except Exception:
                continue
            if item is _SENTINEL:
                if error_holder[0] is not None:
                    raise error_holder[0]
                if not got_audio and (stop_event is None or not stop_event.is_set()):
                    raise RuntimeError(
                        f"Aucun audio reçu du modèle Live ({model}) — "
                        "vérifier que le modèle est disponible sur votre compte"
                    )
                # Vider le coussin si la lecture n'a pas encore démarré (réponse
                # plus courte que le pré-buffer).
                if prebuf and (stop_event is None or not stop_event.is_set()):
                    output_stream.write(bytes(prebuf))
                break
            got_audio = True
            if started:
                output_stream.write(item)
            else:
                prebuf += item
                if len(prebuf) >= preroll_bytes:
                    output_stream.write(bytes(prebuf))
                    prebuf = bytearray()
                    started = True
    finally:
        # Attendre que le buffer interne du stream soit vidé avant de fermer
        output_stream.stop()
        output_stream.close()


def _voice_play_audio(pcm_bytes, sample_rate=24000, stop_event=None):
    """
    Joue des bytes PCM bruts (int16, mono) via sounddevice.

    stop_event (threading.Event) : si activé pendant la lecture, arrête
    immédiatement la lecture (barge-in).
    Bloquant : attend la fin de la lecture avant de retourner.
    Lève ImportError si sounddevice n'est pas installé.
    """
    import sounddevice as _sd
    import numpy as _np
    import time as _time

    audio_array = _np.frombuffer(pcm_bytes, dtype=_np.int16).astype(_np.float32) / 32768.0
    _sd.play(audio_array, samplerate=sample_rate)
    if stop_event is None:
        _sd.wait()
        return
    # Polling toutes les 20 ms pour réagir au barge-in
    total_seconds = len(audio_array) / sample_rate
    deadline = _time.monotonic() + total_seconds + 0.5
    while _time.monotonic() < deadline:
        if stop_event.is_set():
            _sd.stop()
            return
        _time.sleep(0.02)


# ── Dictée vocale — STT (push-to-talk) ────────────────────────────────────────

class _MicRecorder:
    """Enregistreur micro push-to-talk basé sur sounddevice.

    ``start()`` ouvre un flux d'entrée et accumule les échantillons ;
    ``stop()`` ferme le flux et retourne les octets WAV (PCM 16 bits, mono)
    prêts à être envoyés à Gemini, ou ``None`` si rien n'a été capté.
    Chaque instance ne sert qu'à un seul enregistrement.
    """

    def __init__(self, sample_rate=None):
        # sample_rate falsy (0/None) → fréquence native du micro. Forcer une
        # fréquence non supportée par le périphérique déforme l'audio (l'audio
        # part trop vite/lent) et casse la transcription.
        self.sample_rate = sample_rate or None
        self._stream = None
        self._device = None
        self._frames = []
        self._ready = False
        self._on_ready = None
        self._overflows = 0     # nb de callbacks signalant une perte (xrun BT)
        self._t_start = None    # horloge de démarrage (mesure des pertes)
        self.elapsed = 0.0      # durée réelle d'enregistrement (mur)

    @staticmethod
    def _resolve_input_device(_sd):
        """Retourne l'index du micro par défaut via une API stable.

        Laissé au défaut implicite, PortAudio tombe sur le « Mappeur de sons
        Microsoft » (device virtuel de Windows) qui interrompt la capture au
        bout de ~3 s. On cible donc le micro **par défaut de WASAPI** (l'API
        moderne de Windows, celle qu'utilise Wispr Flow), puis MME, puis
        DirectSound. Retourne None si rien de fiable → repli sur le défaut
        implicite de PortAudio.

        Repli machine-spécifique (variable d'environnement AI_VOICE_MIC_DEVICE,
        jamais committée) : sous-chaîne du nom de périphérique à préférer,
        insensible à la casse. Utile quand le device "par défaut" du système
        n'est en réalité pas câblé au micro physique — vu sur un Raspberry
        Pi où "default"/"pulse" (ALSA/PulseAudio) capturaient du silence
        pur alors que le device matériel brut du micro (DJI Mic 2 en USB)
        fonctionnait. Lister les devices disponibles :
          python3 -c "import sounddevice as sd; print(sd.query_devices())"
        """
        try:
            hostapis = _sd.query_hostapis()
            devices = _sd.query_devices()
        except Exception:
            return None

        override = _os.environ.get("AI_VOICE_MIC_DEVICE", "").strip().lower()
        if override:
            for candidate, info in enumerate(devices):
                if (info.get("max_input_channels", 0) > 0
                        and override in (info.get("name") or "").lower()):
                    return candidate

        # Noms écartés : Mappeur de sons (virtuel) et boucles de sortie.
        exclude = ("mapp", "mixage", "stereo mix", "loopback",
                   "primary", "principal")

        def _is_valid(index):
            if not isinstance(index, int) or index < 0:
                return False
            try:
                info = devices[index]
            except Exception:
                return False
            if info.get("max_input_channels", 0) <= 0:
                return False
            name = (info.get("name") or "").lower()
            return not any(token in name for token in exclude)

        for api_name in ("Windows WASAPI", "MME", "Windows DirectSound"):
            for api_index, hostapi in enumerate(hostapis):
                if hostapi.get("name") != api_name:
                    continue
                # 1) micro par défaut de cette API s'il est exploitable
                index = hostapi.get("default_input_device", -1)
                if _is_valid(index):
                    return index
                # 2) sinon, premier micro réel rattaché à cette API
                for candidate, info in enumerate(devices):
                    if (info.get("hostapi") == api_index
                            and _is_valid(candidate)):
                        return candidate
        return None

    def start(self, on_ready=None):
        """Ouvre le flux micro et commence à accumuler l'audio.

        Le périphérique met ~1 s à démarrer sous Windows (initialisation
        pilote) : aucun échantillon n'arrive avant. ``on_ready`` est appelé
        (depuis le thread audio) au tout premier échantillon capté, pour que
        l'UI n'affiche « enregistrement » qu'une fois le micro réellement actif
        et éviter de perdre le début de la phrase.
        """
        import sounddevice as _sd

        self._device = self._resolve_input_device(_sd)

        if not self.sample_rate:
            try:
                info = _sd.query_devices(
                    self._device if self._device is not None else None,
                    kind="input")
                self.sample_rate = int(info["default_samplerate"])
            except Exception:
                self.sample_rate = 44100

        import time as _time

        self._frames = []
        self._ready = False
        self._on_ready = on_ready
        self._overflows = 0
        self._t_start = _time.monotonic()

        def _callback(indata, frames, time_info, status):
            if status:
                # input_overflow = échantillons perdus (tampon trop court /
                # source Bluetooth qui hoquette).
                self._overflows += 1
            if not self._ready:
                self._ready = True
                if self._on_ready is not None:
                    try:
                        self._on_ready()
                    except Exception:
                        pass
            self._frames.append(indata.copy())

        # latency="high" → gros tampon d'entrée : absorbe la gigue du lien
        # Bluetooth (HFP) et limite les pertes d'échantillons en cours de route.
        self._stream = _sd.InputStream(
            device=self._device,
            samplerate=self.sample_rate,
            channels=1,
            dtype="int16",
            latency="high",
            callback=_callback,
        )
        self._stream.start()

    def stop(self):
        import io
        import time
        import wave

        import numpy as _np

        if self._t_start is not None:
            self.elapsed = time.monotonic() - self._t_start

        if self._stream is not None:
            # Laisser le tampon d'entrée se vider avant d'arrêter : en Bluetooth,
            # l'audio capté arrive au callback avec un retard de transport ; sans
            # cette pause, la fin de la phrase (encore « en vol ») est perdue.
            try:
                latency = float(getattr(self._stream, "latency", 0.0) or 0.0)
            except Exception:
                latency = 0.0
            time.sleep(min(max(latency + 0.3, 0.3), 1.0))
            try:
                self._stream.stop()
                self._stream.close()
            finally:
                self._stream = None
        if not self._frames:
            return None
        audio = _np.concatenate(self._frames, axis=0)

        # Retire les blancs en tête/queue avant l'envoi (retour user : moins
        # d'audio à transmettre/transcrire = plus rapide) — seuil d'amplitude
        # simple, pas de VAD dédiée, avec une marge de 150 ms de chaque côté
        # pour ne jamais couper le début/la fin de la parole.
        flat = audio.reshape(-1)
        above = _np.where(_np.abs(flat) > 500)[0]
        if above.size:
            pad = int(self.sample_rate * 0.15)
            start = max(0, above[0] - pad)
            end = min(len(flat), above[-1] + pad)
            flat = flat[start:end]

        # Plafonne aussi les blancs internes trop longs (pause de réflexion
        # en cours de phrase, retour user) : découpe en fenêtres de 20 ms,
        # énergie RMS par fenêtre, et toute pause interne au-delà de 500 ms
        # est tronquée à 500 ms — les pauses courtes naturelles entre les
        # mots restent intactes.
        frame_len = max(1, int(self.sample_rate * 0.02))
        n_frames = len(flat) // frame_len
        if n_frames > 1:
            usable = flat[:n_frames * frame_len].reshape(n_frames, frame_len)
            frame_rms = _np.sqrt(_np.mean(usable.astype(_np.float64) ** 2, axis=1))
            voiced = frame_rms > 300
            max_silence_frames = int(0.5 / 0.02)
            keep = []
            silence_run = 0
            for i, is_voiced in enumerate(voiced):
                if is_voiced:
                    silence_run = 0
                    keep.append(i)
                else:
                    silence_run += 1
                    if silence_run <= max_silence_frames:
                        keep.append(i)
            if keep:
                remainder = flat[n_frames * frame_len:]
                flat = _np.concatenate([usable[keep].reshape(-1), remainder])

        audio = flat.reshape(-1, 1)

        buffer = io.BytesIO()
        with wave.open(buffer, "wb") as wav_file:
            wav_file.setnchannels(1)
            wav_file.setsampwidth(2)  # int16 → 2 octets
            wav_file.setframerate(self.sample_rate)
            wav_file.writeframes(audio.tobytes())
        return buffer.getvalue()


def _gemini_transcribe_audio(wav_bytes, language_code="fr",
                             model="gemini-3.1-flash-lite"):
    """Transcrit un audio WAV via Gemini et retourne le texte (ou None).

    L'audio est envoyé comme Part inline (même mécanisme que les images),
    avec une consigne stricte pour n'obtenir que la transcription brute.
    Nécessite GEMINI_API_KEY. Retourne None en cas d'erreur ou d'audio vide.
    """
    try:
        from google import genai as _genai
        from google.genai import types as _gtypes
    except ImportError:
        return None

    api_key = _get_gemini_api_key()
    if not api_key or not wav_bytes:
        return None

    lang = language_code or "fr"
    prompt = (
        "Transcris fidèlement cet enregistrement audio. Réponds UNIQUEMENT "
        "avec le texte prononcé, sans guillemets, sans commentaire et sans "
        f"préambule. Langue attendue : {lang}. Si l'audio est vide ou "
        "inaudible, réponds par une chaîne vide."
    )
    # Pas de try/except ici : une erreur d'appel API (réseau, quota, clé
    # invalide…) doit remonter à l'appelant, qui la logge déjà
    # (Dashboard.pyw/_worker : "[ERREUR] Transcription : ..."). L'avaler
    # silencieusement transformait toute panne réelle en un opaque "Aucun
    # texte reconnu" sans piste de diagnostic.
    client = _genai.Client(api_key=api_key)
    response = client.models.generate_content(
        model=model,
        contents=[
            _gtypes.Content(
                role="user",
                parts=[
                    _gtypes.Part.from_bytes(
                        data=wav_bytes, mime_type="audio/wav"),
                    _gtypes.Part(text=prompt),
                ],
            )
        ],
    )
    text = (response.text or "").strip()
    return text or None


# ── Claude (Anthropic) ────────────────────────────────────────────────────────

def _get_anthropic_api_key():
    """Récupère ANTHROPIC_API_KEY depuis l'env, un fichier .env ou le shell de login."""
    import os
    import re
    import subprocess

    key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if key:
        return key

    for env_dir in [
        os.getcwd(),
        os.path.dirname(os.path.abspath(__file__)),
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    ]:
        env_path = os.path.join(env_dir, ".env")
        if os.path.exists(env_path):
            try:
                with open(env_path, "r", encoding="utf-8") as f:
                    for line in f:
                        if line.strip() and not line.startswith("#") and "=" in line:
                            k, v = line.split("=", 1)
                            if k.strip() == "ANTHROPIC_API_KEY":
                                key = v.strip().strip('"').strip("'")
                                if key:
                                    os.environ["ANTHROPIC_API_KEY"] = key
                                    return key
            except Exception:
                pass

    if os.name != "nt":
        for shell in ["/bin/zsh", "/bin/bash"]:
            if os.path.exists(shell):
                try:
                    # "-li" (login + interactif) : la plupart des .bashrc
                    # (Debian/Raspberry Pi OS...) s'arrêtent avant l'export
                    # de la clé si le shell n'est pas interactif.
                    result = subprocess.run(
                        [shell, "-li", "-c", "echo $ANTHROPIC_API_KEY"],
                        capture_output=True, text=True, timeout=2,
                    )
                    key = result.stdout.strip()
                    if key:
                        os.environ["ANTHROPIC_API_KEY"] = key
                        return key
                except Exception:
                    pass

        home = os.path.expanduser("~")
        for rc_file in [".zshrc", ".bashrc", ".bash_profile", ".profile"]:
            rc_path = os.path.join(home, rc_file)
            if os.path.isfile(rc_path):
                try:
                    with open(rc_path, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read()
                    # re.MULTILINE : sinon "$" ne matche que la fin du
                    # fichier, pas la fin de chaque ligne.
                    match = re.search(
                        r'(?:export\s+)?ANTHROPIC_API_KEY\s*=\s*["\']?(.*?)["\']?\s*(?:#|$)',
                        content, re.MULTILINE,
                    )
                    if match:
                        key = match.group(1).strip()
                        if key:
                            os.environ["ANTHROPIC_API_KEY"] = key
                            return key
                except Exception:
                    pass

    return ""


def _ollama_tools_to_claude(tools):
    """Convertit les définitions d'outils Ollama au format Anthropic Claude."""
    if not tools:
        return []
    claude_tools = []
    for tool in tools:
        fn = tool.get("function", tool)
        claude_tools.append({
            "name": fn.get("name", ""),
            "description": fn.get("description", ""),
            "input_schema": fn.get("parameters", {"type": "object", "properties": {}}),
        })
    return claude_tools


def _ollama_messages_to_claude(messages):
    """
    Convertit une liste de messages Ollama/Gemini au format Anthropic Claude.
    Retourne (system_prompt: str | None, claude_messages: list).

    Gère : system, user, assistant (avec/sans tool_calls), tool (résultats).
    Les IDs des tool_calls sont trackés pour apparier tool_use et tool_result.
    """
    import json as _json

    system_parts = []
    result = []
    _pending_ids = {}  # tool_name -> [id, ...]

    for msg in messages:
        role    = msg.get("role", "")
        content = msg.get("content", "")

        if role == "system":
            system_parts.append(content)

        elif role == "user":
            if result and result[-1]["role"] == "user":
                prev = result[-1]["content"]
                if isinstance(prev, list):
                    prev.append({"type": "text", "text": content})
                else:
                    result[-1]["content"] = str(prev) + "\n" + content
            else:
                result.append({"role": "user", "content": content})

        elif role == "assistant":
            tool_calls = msg.get("tool_calls", [])
            if tool_calls:
                blocks = []
                if content:
                    blocks.append({"type": "text", "text": content})
                for tc in tool_calls:
                    fn      = tc.get("function", {})
                    tc_name = fn.get("name", "")
                    tc_id   = tc.get("id") or f"toolu_{tc_name}_{len(result)}"
                    blocks.append({
                        "type":  "tool_use",
                        "id":    tc_id,
                        "name":  tc_name,
                        "input": fn.get("arguments", {}),
                    })
                    _pending_ids.setdefault(tc_name, []).append(tc_id)
                result.append({"role": "assistant", "content": blocks})
            else:
                if content:
                    result.append({"role": "assistant", "content": content})

        elif role == "tool":
            tool_name = msg.get("name", msg.get("tool_name", ""))
            ids   = _pending_ids.get(tool_name, [])
            tc_id = ids.pop(0) if ids else f"toolu_{tool_name}"
            tr_block = {"type": "tool_result", "tool_use_id": tc_id, "content": content}

            if (result and result[-1]["role"] == "user"
                    and isinstance(result[-1]["content"], list)
                    and result[-1]["content"]
                    and result[-1]["content"][-1].get("type") == "tool_result"):
                result[-1]["content"].append(tr_block)
            else:
                result.append({"role": "user", "content": [tr_block]})

    # Assurer que le premier message est un message user
    while result and result[0]["role"] != "user":
        result.pop(0)

    # Fusionner les messages consécutifs de même rôle (Claude ne les accepte pas séparés)
    merged = []
    for m in result:
        if merged and merged[-1]["role"] == m["role"]:
            prev_c = merged[-1]["content"]
            curr_c = m["content"]
            if isinstance(prev_c, str) and isinstance(curr_c, str):
                merged[-1]["content"] = prev_c + "\n" + curr_c
            elif isinstance(prev_c, list) and isinstance(curr_c, list):
                merged[-1]["content"].extend(curr_c)
            elif isinstance(prev_c, str) and isinstance(curr_c, list):
                merged[-1]["content"] = [{"type": "text", "text": prev_c}] + curr_c
            else:
                merged.append(m)
        else:
            merged.append(m)

    return ("\n\n".join(system_parts) if system_parts else None), merged


def _claude_chat_stream_with_tools(model, messages, tools=None, temperature=0.7):
    """
    Version Anthropic Claude de _gemini_chat_stream_with_tools.
    Génère les mêmes tuples : ("token", str), ("thinking", str), ("tool_calls", list).

    tool_calls items sont compatibles format Ollama :
      [{"id": str, "function": {"name": str, "arguments": dict}}, ...]

    Nécessite ANTHROPIC_API_KEY dans les variables d'environnement.
    """
    try:
        import anthropic as _anthropic
    except ImportError:
        yield ("token", (
            "[Erreur : anthropic n'est pas installé. "
            "Exécute : pip install anthropic]"
        ))
        return

    api_key = _get_anthropic_api_key()
    if not api_key:
        yield ("token", (
            "[Erreur : ANTHROPIC_API_KEY n'est pas définie. "
            "Configurez-la dans votre .zshrc ou dans un fichier .env : "
            "ANTHROPIC_API_KEY=votre_cle]"
        ))
        return

    try:
        client = _anthropic.Anthropic(api_key=api_key)
    except Exception as exc:
        yield ("token", f"[Erreur initialisation Claude : {exc}]")
        return

    system_prompt, claude_messages = _ollama_messages_to_claude(messages)
    if not claude_messages:
        yield ("token", "[Erreur : aucun message utilisateur valide]")
        return

    claude_tools = _ollama_tools_to_claude(tools) if tools else None

    kwargs = {
        "model":       model,
        "max_tokens":  16384,
        "temperature": temperature,
        "messages":    claude_messages,
    }
    # Prompt caching (GA, aucun beta header) : on met un cache_control sur les
    # deux gros préfixes statiques pour que les re-envois coûtent ~0,1×. Ordre
    # de rendu : tools → system → messages. Le breakpoint sur le dernier outil
    # cache tout le bloc d'outils et reste valide même quand le système change
    # (changer le système n'invalide pas le cache des outils). Le breakpoint
    # système cache system.md quand il est stable. En dessous du seuil minimal
    # du modèle, le cache ne se crée pas silencieusement — sans effet néfaste.
    _CACHE = {"type": "ephemeral"}
    if system_prompt:
        kwargs["system"] = [{"type": "text", "text": system_prompt,
                             "cache_control": _CACHE}]
    if claude_tools:
        claude_tools[-1] = {**claude_tools[-1], "cache_control": _CACHE}
        kwargs["tools"] = claude_tools

    tool_calls_out    = []
    current_tool      = None
    current_tool_json = ""

    try:
        with client.messages.stream(**kwargs) as stream:
            for event in stream:
                etype = getattr(event, "type", None)

                if etype == "content_block_start":
                    cb = getattr(event, "content_block", None)
                    if cb and getattr(cb, "type", None) == "tool_use":
                        current_tool      = {"id": cb.id, "name": cb.name}
                        current_tool_json = ""

                elif etype == "content_block_delta":
                    delta = getattr(event, "delta", None)
                    if delta is None:
                        continue
                    delta_type = getattr(delta, "type", None)
                    if delta_type == "text_delta":
                        yield ("token", delta.text)
                    elif delta_type == "thinking_delta":
                        yield ("thinking", delta.thinking)
                    elif delta_type == "input_json_delta":
                        current_tool_json += getattr(delta, "partial_json", "")

                elif etype == "content_block_stop":
                    if current_tool is not None:
                        try:
                            import json as _j
                            tool_input = _j.loads(current_tool_json) if current_tool_json else {}
                        except Exception as exc:
                            _logger.warning(
                                "arguments de tool call (Claude) illisibles, "
                                "remplacés par {} : %r — brut : %r", exc, current_tool_json)
                            tool_input = {}
                        tool_calls_out.append({
                            "id": current_tool["id"],
                            "function": {
                                "name":      current_tool["name"],
                                "arguments": tool_input,
                            },
                        })
                        current_tool      = None
                        current_tool_json = ""

    except Exception as exc:
        yield ("token", f"[Erreur Claude : {exc}]")
        return

    if tool_calls_out:
        yield ("tool_calls", tool_calls_out)
    _sd.wait()


# ==============================================================================
# ─── Outil édition partielle (Edit) ──────────────────────────────────────────
# ==============================================================================

def _edit_file(folder_path, filepath, old_string, new_string):
    """Remplace la première occurrence exacte de old_string par new_string dans un fichier."""
    resolved = _resolve_path(folder_path, filepath) if folder_path else filepath
    if not resolved or not _os.path.isfile(resolved):
        return f"[Erreur] Fichier introuvable : {filepath}"
    try:
        with open(resolved, "r", encoding="utf-8", errors="replace") as fh:
            content = fh.read()
        if old_string not in content:
            return (
                f"[Erreur] Chaîne introuvable dans {_os.path.basename(resolved)}. "
                "Vérifiez les espaces, sauts de ligne et la casse."
            )
        new_content = content.replace(old_string, new_string, 1)
        _backup_file(resolved)  # sauvegarde avant écrasement
        with open(resolved, "w", encoding="utf-8") as fh:
            fh.write(new_content)
        diff = new_content.count("\n") - content.count("\n")
        sign = "+" if diff >= 0 else ""
        return f"[OK] {_os.path.basename(resolved)} modifié ({sign}{diff} ligne(s))."
    except Exception as exc:
        return f"[Erreur] {exc}"


def _read_file_lines(folder_path, filepath, start_line=1, end_line=None):
    """Lit une plage de lignes d'un fichier, avec numéros de ligne."""
    resolved = _resolve_path(folder_path, filepath) if folder_path else filepath
    if not resolved or not _os.path.isfile(resolved):
        return f"Fichier introuvable : {filepath}"
    try:
        with open(resolved, "r", encoding="utf-8", errors="replace") as fh:
            lines = fh.readlines()
        total = len(lines)
        start = max(1, int(start_line or 1))
        end = min(total, int(end_line or total))
        if start > total:
            return f"Ligne de départ ({start}) dépasse le total ({total} lignes)."
        selected = lines[start - 1:end]
        header = f"[{_os.path.basename(resolved)} — lignes {start}–{end} / {total}]\n"
        return header + "".join(
            f"{start + i:>6} │ {line}" for i, line in enumerate(selected)
        )
    except Exception as exc:
        return f"Erreur : {exc}"


_READ_LINES_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_file_lines",
            "description": (
                "Lit une plage de lignes précise d'un fichier texte. "
                "Indispensable pour les gros fichiers (Dashboard.pyw, SidePanel.pyw…) : "
                "utiliser search_in_files pour trouver les numéros de ligne, "
                "puis read_file_lines pour lire uniquement la section pertinente. "
                "Retourne les lignes numérotées pour faciliter l'utilisation avec edit_file."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Chemin absolu ou relatif au dossier ouvert",
                    },
                    "start_line": {
                        "type": "integer",
                        "description": "Première ligne à lire (1-indexé, défaut : 1)",
                    },
                    "end_line": {
                        "type": "integer",
                        "description": "Dernière ligne à lire incluse (défaut : fin du fichier)",
                    },
                },
                "required": ["filepath"],
            },
        },
    },
]


_EDIT_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "edit_file",
            "description": (
                "Remplace une portion précise d'un fichier texte (old_string → new_string). "
                "Contrairement à create_file qui réécrit tout le fichier, edit_file effectue "
                "une substitution chirurgicale de la première occurrence exacte de old_string. "
                "Idéal pour corriger une ligne, modifier une valeur ou insérer du code à un endroit précis. "
                "Utilise read_file_content avant pour obtenir le texte exact à remplacer."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Chemin absolu ou relatif au dossier ouvert",
                    },
                    "old_string": {
                        "type": "string",
                        "description": "Texte exact à remplacer (doit être unique dans le fichier)",
                    },
                    "new_string": {
                        "type": "string",
                        "description": "Texte de remplacement",
                    },
                },
                "required": ["filepath", "old_string", "new_string"],
            },
        },
    },
]


# ==============================================================================
# ─── Outils recherche (Grep / Glob) ──────────────────────────────────────────
# ==============================================================================

import glob as _glob_mod


def _search_in_files(
    folder_path, pattern, path=None, file_glob="*",
    max_results=50, case_sensitive=False,
):
    """Recherche un pattern regex dans les fichiers texte d'un dossier (récursif)."""
    search_root = path if (path and _os.path.isdir(path)) else folder_path
    if not search_root:
        return "[Erreur] Aucun dossier spécifié."
    flags = 0 if case_sensitive else re.IGNORECASE
    try:
        compiled = re.compile(pattern, flags)
    except re.error as exc:
        return f"[Erreur] Pattern invalide : {exc}"
    results = []
    count = 0
    try:
        all_files = sorted(
            _glob_mod.glob(_os.path.join(search_root, "**", file_glob), recursive=True)
        )
    except Exception as exc:
        return f"[Erreur] {exc}"
    for filepath in all_files:
        if not _os.path.isfile(filepath):
            continue
        try:
            with open(filepath, "r", encoding="utf-8", errors="ignore") as fh:
                for lineno, line in enumerate(fh, 1):
                    if compiled.search(line):
                        rel = _os.path.relpath(filepath, search_root)
                        results.append(f"{rel}:{lineno}: {line.rstrip()}")
                        count += 1
                        if count >= max_results:
                            results.append(f"\n… limité à {max_results} résultats.")
                            return "\n".join(results)
        except Exception:
            continue
    if not results:
        return f"Aucun résultat pour « {pattern} » dans {_os.path.basename(search_root)}/."
    return f"{count} correspondance(s) :\n" + "\n".join(results)


def _find_files(folder_path, pattern, base_path=None, max_results=200):
    """Trouve des fichiers correspondant à un motif glob (récursif)."""
    search_root = base_path if (base_path and _os.path.isdir(base_path)) else folder_path
    if not search_root:
        return "[Erreur] Aucun dossier spécifié."
    try:
        matches = sorted(
            _glob_mod.glob(_os.path.join(search_root, "**", pattern), recursive=True)
        )
    except Exception as exc:
        return f"[Erreur] {exc}"
    if not matches:
        return f"Aucun fichier trouvé pour « {pattern} »."
    lines = []
    for p in matches[:max_results]:
        rel = _os.path.relpath(p, search_root)
        if _os.path.isdir(p):
            lines.append(f"📁 {rel}/")
        else:
            size = _os.path.getsize(p)
            lines.append(f"📄 {rel}  ({size:,} o)")
    if len(matches) > max_results:
        lines.append(f"… et {len(matches) - max_results} autres.")
    total = len(matches)
    shown = min(total, max_results)
    return f"{shown}/{total} fichier(s) :\n" + "\n".join(lines)


_SEARCH_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "search_in_files",
            "description": (
                "Recherche un motif regex dans le contenu des fichiers texte (équivalent de grep). "
                "Retourne les lignes correspondantes avec leur chemin et numéro de ligne. "
                "Utile pour trouver une fonction, une variable, une valeur de config "
                "ou tout texte dans des fichiers source."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "pattern": {
                        "type": "string",
                        "description": "Expression régulière à rechercher",
                    },
                    "path": {
                        "type": "string",
                        "description": "Dossier de recherche (absolu). Défaut : dossier ouvert.",
                    },
                    "file_glob": {
                        "type": "string",
                        "description": "Filtre de fichiers (ex. '*.py', '*.json'). Défaut : tous.",
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Nombre maximum de résultats. Défaut : 50.",
                    },
                    "case_sensitive": {
                        "type": "boolean",
                        "description": "Sensible à la casse. Défaut : false.",
                    },
                },
                "required": ["pattern"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "find_files",
            "description": (
                "Trouve des fichiers par motif glob récursif (équivalent de find). "
                "Exemples : '*.py' pour tous les scripts Python, '*.json' pour tous les JSON, "
                "'photo_*.jpg' pour les fichiers photo. Retourne chemins et tailles."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "pattern": {
                        "type": "string",
                        "description": "Motif glob, ex. '*.py', 'rapport*.pdf', '*.json'",
                    },
                    "base_path": {
                        "type": "string",
                        "description": "Dossier de base (absolu). Défaut : dossier ouvert.",
                    },
                    "max_results": {
                        "type": "integer",
                        "description": "Nombre maximum de résultats. Défaut : 200.",
                    },
                },
                "required": ["pattern"],
            },
        },
    },
]


# ==============================================================================
# ─── Outil git ────────────────────────────────────────────────────────────────
# ==============================================================================

_SAFE_GIT_SUBCMDS = {
    "status", "log", "diff", "show", "branch", "add", "commit",
    "push", "pull", "fetch", "checkout", "stash", "merge", "tag",
    "remote", "config", "--version", "init", "clone",
}


def _git_command(args, cwd=None):
    """Exécute une sous-commande git (liste autorisée seulement)."""
    if not args:
        return "[Erreur] Aucun argument fourni."
    subcmd = str(args[0]).lstrip("-")
    if subcmd not in _SAFE_GIT_SUBCMDS:
        allowed = ", ".join(sorted(_SAFE_GIT_SUBCMDS))
        return f"[Refusé] Sous-commande git non autorisée : {args[0]}.\nAutorisées : {allowed}"
    try:
        result = _subprocess.run(
            ["git"] + [str(a) for a in args],
            cwd=cwd,
            capture_output=True,
            text=True,
            timeout=30,
            encoding="utf-8",
            errors="replace",
        )
        out = (result.stdout or "").strip()
        err = (result.stderr or "").strip()
        if result.returncode != 0:
            return f"[git exit {result.returncode}]\n{err or out}"
        return out or "(commande exécutée sans sortie)"
    except FileNotFoundError:
        return "[Erreur] git n'est pas installé ou introuvable dans le PATH."
    except _subprocess.TimeoutExpired:
        return "[Timeout] La commande git a dépassé 30 secondes."
    except Exception as exc:
        return f"[Erreur] {exc}"


_GIT_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "git_command",
            "description": (
                "Exécute une commande git sur un dépôt local. "
                "Sous-commandes autorisées : status, log, diff, show, branch, add, commit, "
                "push, pull, fetch, checkout, stash, merge, tag, remote, config, init, clone. "
                "Exemples : ['status'], ['log', '--oneline', '-10'], ['diff', 'HEAD~1'], "
                "['commit', '-m', 'message'], ['add', 'fichier.py']."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "args": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "Arguments git sans le mot 'git', ex. ['log', '--oneline', '-5']",
                    },
                    "cwd": {
                        "type": "string",
                        "description": "Répertoire du dépôt (absolu). Défaut : dossier ouvert.",
                    },
                },
                "required": ["args"],
            },
        },
    },
]


# ==============================================================================
# ─── Outil gestion de tâches (TodoWrite) ─────────────────────────────────────
# ==============================================================================

_TASKS_FILE = _os.path.join(_DATA_DIR, ".tasks.json")
_STATUS_ICON = {"todo": "⬜", "in_progress": "🔄", "done": "✅"}


def _manage_tasks(action, task_id=None, title=None, status=None, notes=None):
    """CRUD sur une liste de tâches persistante (.tasks.json)."""
    import datetime as _dt_t

    def _load():
        try:
            with open(_TASKS_FILE, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:
            return []

    def _save(tasks):
        with open(_TASKS_FILE, "w", encoding="utf-8") as fh:
            json.dump(tasks, fh, ensure_ascii=False, indent=2)

    tasks = _load()

    if action == "list":
        if not tasks:
            return "Aucune tâche."
        lines = []
        for t in tasks:
            icon = _STATUS_ICON.get(t.get("status", "todo"), "⬜")
            line = f"{icon} [{t['id']}] {t['title']}"
            if t.get("notes"):
                line += f"\n      ↳ {t['notes']}"
            lines.append(line)
        return "\n".join(lines)

    elif action == "add":
        if not title:
            return "[Erreur] Paramètre 'title' requis."
        next_id = str(
            max((int(t["id"]) for t in tasks if str(t.get("id", "0")).isdigit()), default=0) + 1
        )
        tasks.append({
            "id": next_id,
            "title": title,
            "status": status or "todo",
            "notes": notes or "",
            "created": _dt_t.datetime.now().isoformat(),
        })
        _save(tasks)
        return f"[OK] Tâche #{next_id} ajoutée : {title}"

    elif action == "update":
        if task_id is None:
            return "[Erreur] Paramètre 'task_id' requis."
        for t in tasks:
            if str(t.get("id")) == str(task_id):
                if status is not None:
                    t["status"] = status
                if title is not None:
                    t["title"] = title
                if notes is not None:
                    t["notes"] = notes
                _save(tasks)
                return f"[OK] Tâche #{task_id} mise à jour."
        return f"[Erreur] Tâche #{task_id} introuvable."

    elif action == "delete":
        if task_id is None:
            return "[Erreur] Paramètre 'task_id' requis."
        before = len(tasks)
        tasks = [t for t in tasks if str(t.get("id")) != str(task_id)]
        if len(tasks) == before:
            return f"[Erreur] Tâche #{task_id} introuvable."
        _save(tasks)
        return f"[OK] Tâche #{task_id} supprimée."

    elif action == "clear":
        _save([])
        return "[OK] Liste de tâches vidée."

    return (
        f"[Erreur] Action inconnue : {action}. "
        "Valeurs acceptées : list, add, update, delete, clear."
    )


_TASK_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "manage_tasks",
            "description": (
                "Gère une liste de tâches persistante (équivalent de TodoWrite). "
                "Permet de suivre les travaux en cours avec des statuts : todo, in_progress, done. "
                "Actions : list (lister), add (ajouter), update (modifier), delete (supprimer), clear (vider)."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["list", "add", "update", "delete", "clear"],
                        "description": "Action à effectuer",
                    },
                    "task_id": {
                        "type": "string",
                        "description": "ID de la tâche (requis pour update et delete)",
                    },
                    "title": {
                        "type": "string",
                        "description": "Titre de la tâche (requis pour add)",
                    },
                    "status": {
                        "type": "string",
                        "enum": ["todo", "in_progress", "done"],
                        "description": "Statut de la tâche",
                    },
                    "notes": {
                        "type": "string",
                        "description": "Notes supplémentaires",
                    },
                },
                "required": ["action"],
            },
        },
    },
]


# ==============================================================================
# ─── Outil lecture PDF ────────────────────────────────────────────────────────
# ==============================================================================

def _parse_pdf_pages(pages_str, total):
    """Convertit '1-3,5' en liste d'indices 0-based valides."""
    indices = set()
    for part in str(pages_str).split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                indices.update(range(int(a.strip()) - 1, int(b.strip())))
            except ValueError:
                pass
        elif part.isdigit():
            indices.add(int(part) - 1)
    return sorted(i for i in indices if 0 <= i < total)


def _read_pdf(folder_path, filepath, pages=None, max_chars=40000):
    """Extrait le texte d'un PDF (pypdf en priorité, PyMuPDF en fallback)."""
    resolved = _resolve_path(folder_path, filepath) if folder_path else filepath
    if not resolved or not _os.path.isfile(resolved):
        return f"[Erreur] Fichier introuvable : {filepath}"

    def _format(parts, total):
        full = "\n\n".join(parts) or "(Aucun texte extractible)"
        if len(full) > max_chars:
            return full[:max_chars] + f"\n\n… (tronqué à {max_chars} caractères sur {len(full)})"
        return full

    # PyMuPDF en priorité : meilleure extraction, gère les encodages complexes
    try:
        import fitz as _fitz
        doc = _fitz.open(resolved)
        total = doc.page_count
        idxs = _parse_pdf_pages(pages, total) if pages else range(total)
        parts = []
        for i in idxs:
            if 0 <= i < total:
                text = doc[i].get_text().strip()
                if text:
                    parts.append(f"[Page {i + 1}/{total}]\n{text}")
        return _format(parts, total)
    except ImportError:
        pass

    # pypdf en fallback (pure Python, extraction basique)
    try:
        import pypdf as _pypdf
        reader = _pypdf.PdfReader(resolved)
        total = len(reader.pages)
        idxs = _parse_pdf_pages(pages, total) if pages else range(total)
        parts = []
        for i in idxs:
            if 0 <= i < total:
                text = (reader.pages[i].extract_text() or "").strip()
                if text:
                    parts.append(f"[Page {i + 1}/{total}]\n{text}")
        return _format(parts, total)
    except ImportError:
        pass

    return "[Erreur] Aucun module PDF disponible. Installez PyMuPDF : pip install pymupdf"


_PDF_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_pdf",
            "description": (
                "Extrait le texte d'un fichier PDF. "
                "Utile pour lire des contrats, factures, manuels ou documents PDF "
                "avec des modèles qui ne supportent pas nativement les PDF (Ollama, Claude). "
                "Supporte les chemins absolus ou relatifs au dossier ouvert."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Chemin du fichier PDF (absolu ou relatif au dossier ouvert)",
                    },
                    "pages": {
                        "type": "string",
                        "description": "Pages à extraire, ex. '1-5', '3', '1,4,7'. Défaut : toutes.",
                    },
                },
                "required": ["filepath"],
            },
        },
    },
]


# ==============================================================================
# ─── Outil sous-agent (délégation IA) ────────────────────────────────────────
# ==============================================================================

_NETWORK_EXC_NAMES = frozenset({
    "ConnectError", "TimeoutException", "ConnectTimeout", "ReadTimeout",
    "NetworkError", "ProxyError", "RemoteDisconnected", "IncompleteRead",
})


def _is_network_error(exc):
    """Vrai si l'exception signale une coupure réseau plutôt qu'une erreur API."""
    if type(exc).__name__ in _NETWORK_EXC_NAMES:
        return True
    if isinstance(exc, (ConnectionError, TimeoutError, OSError)):
        return True
    msg = str(exc).lower()
    return any(kw in msg for kw in (
        "connection refused", "failed to connect", "network is unreachable",
        "name or service not known", "nodename nor servname",
        "connection timed out", "timed out", "no route to host",
    ))


def _call_ai_text_with_fallback(msgs, model=None):
    """
    Appel IA texte avec chaîne de fallback automatique :
      1. model (ou AI_MODEL_TEXT)     — Gemini 3.5 Flash par défaut
      2. AI_GEMINI_FALLBACK_CLOUD     — Gemini 3.1 Pro si modèle indisponible/quota
      3. AI_GEMINI_FALLBACK (Ollama)  — Gemma local si coupure réseau

    Retourne (texte, modèle_utilisé, erreur_ou_None).
    """
    primary   = model or CONSTANTS.AI_MODEL_TEXT or ""
    cloud_fb  = getattr(CONSTANTS, "AI_GEMINI_FALLBACK_CLOUD", "")
    local_fb  = getattr(CONSTANTS, "AI_GEMINI_FALLBACK", "")

    def _stream_cloud(mdl):
        parts = []
        if mdl.startswith("gemini"):
            for evt, dat in _gemini_chat_stream_with_tools(mdl, msgs, tools=None):
                if evt == "token":
                    parts.append(dat)
        else:
            for evt, dat in _claude_chat_stream_with_tools(mdl, msgs, tools=None):
                if evt == "token":
                    parts.append(dat)
        return "".join(parts).strip() or "(Aucune réponse)"

    def _run_ollama(mdl):
        resp = _ollama_chat_once(CONSTANTS.AI_OLLAMA_URL, mdl, msgs)
        return resp.get("content", "(Aucune réponse)")

    if not (primary.startswith("gemini") or primary.startswith("claude")):
        # Modèle Ollama dès le départ — pas de fallback cloud
        try:
            return _run_ollama(primary), primary, None
        except Exception as exc:
            return f"[Erreur] {exc}", None, str(exc)

    # ── Niveau 1 : modèle principal ──────────────────────────────────────────
    try:
        return _stream_cloud(primary), primary, None
    except Exception as exc1:
        network_down = _is_network_error(exc1)

        # ── Niveau 2 : cloud fallback (si réseau OK mais modèle indisponible) ─
        if not network_down and cloud_fb and cloud_fb != primary:
            try:
                return _stream_cloud(cloud_fb), cloud_fb, str(exc1)
            except Exception as exc2:
                network_down = _is_network_error(exc2)
                if not network_down:
                    # Les deux Gemini ont échoué pour raison modèle → Ollama
                    if local_fb:
                        try:
                            return _run_ollama(local_fb), local_fb, f"{exc1} | {exc2}"
                        except Exception as exc3:
                            return (
                                f"[Erreur] {exc1} | {exc2} | Ollama: {exc3}",
                                None, str(exc3),
                            )
                    return f"[Erreur] {exc1} | {exc2}", None, str(exc2)
                # exc2 est réseau → tomber sur Ollama

        # ── Niveau 3 : Ollama local ───────────────────────────────────────────
        if local_fb:
            try:
                return _run_ollama(local_fb), local_fb, str(exc1)
            except Exception as exc3:
                return f"[Erreur réseau] {exc1} | Ollama: {exc3}", None, str(exc3)

        return f"[Erreur] {exc1}", None, str(exc1)


def _ask_subagent(task, context=None, model=None):
    """Délègue une tâche focalisée à un appel IA distinct (sans outils, résultat direct)."""
    user_content = f"Contexte :\n{context}\n\nTâche :\n{task}" if context else task
    msgs = [
        {"role": "system", "content": (
            "Tu es un assistant spécialisé appelé en sous-tâche. "
            "Réponds de façon concise et précise, en te concentrant uniquement sur la demande."
        )},
        {"role": "user", "content": user_content},
    ]
    result, used_model, err = _call_ai_text_with_fallback(msgs, model=model)
    primary = model or CONSTANTS.AI_MODEL_TEXT or ""
    if err and used_model and used_model != primary:
        return f"[↩ {used_model}]\n{result}"
    return result


_SUBAGENT_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "ask_subagent",
            "description": (
                "Délègue une sous-tâche à une instance IA distincte (appel focalisé sans outils). "
                "Utile pour paralléliser le raisonnement : rédiger un texte, traduire, résumer, "
                "classer ou analyser un contenu de façon autonome pendant que tu construis ta réponse. "
                "À utiliser pour des tâches qui ne nécessitent pas d'accès aux fichiers ou au terminal."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "task": {
                        "type": "string",
                        "description": "Tâche précise à confier au sous-agent",
                    },
                    "context": {
                        "type": "string",
                        "description": "Contexte additionnel (optionnel)",
                    },
                    "model": {
                        "type": "string",
                        "description": "Modèle à utiliser. Défaut : modèle actif.",
                    },
                },
                "required": ["task"],
            },
        },
    },
]


# ==============================================================================
# ─── Outil planification (Schedule) ──────────────────────────────────────────
# ==============================================================================

def _schedule_task(action, name=None, command=None, when=None):
    """Crée, liste ou supprime des tâches planifiées (schtasks Windows / crontab Unix)."""
    import platform as _plat
    import datetime as _dt

    is_win = _plat.system() == "Windows"
    _MARKER = "# SCHED_TASK:"

    # ── Helpers crontab (Linux / macOS) ──────────────────────────────────────

    def _crontab_read():
        r = _subprocess.run(["crontab", "-l"], capture_output=True,
                            text=True, encoding="utf-8", errors="replace")
        return r.stdout if r.returncode == 0 else ""

    def _crontab_write(content):
        r = _subprocess.run(["crontab", "-"], input=content, text=True,
                            encoding="utf-8", capture_output=True, timeout=15)
        return r.returncode == 0, (r.stderr or r.stdout).strip()

    def _crontab_remove(task_name):
        """Retire les lignes d'une tâche nommée ; retourne (found, new_content)."""
        marker = f"{_MARKER} {task_name}"
        lines = _crontab_read().splitlines()
        out, skip, found = [], False, False
        for line in lines:
            if skip:
                skip = False
                found = True
                continue
            if line.strip() == marker:
                skip = True
                continue
            out.append(line)
        return found, "\n".join(out) + "\n"

    def _when_to_cron(when_str):
        """
        Convertit 'HH:MM' (aujourd'hui) ou 'YYYY-MM-DD HH:MM' en expression cron.
        Accepte aussi une expression cron déjà formée (5 champs).
        """
        parts = when_str.strip().split()
        # Expression cron déjà formée : "30 14 * * 1"
        if len(parts) == 5 and all(
            c.isdigit() or c in "*,-/?" for c in "".join(parts)
        ):
            return " ".join(parts), None
        # YYYY-MM-DD HH:MM
        if len(parts) == 2 and "-" in parts[0]:
            try:
                y, mo, d = parts[0].split("-")
                h, mi = parts[1].split(":")
                return f"{int(mi)} {int(h)} {int(d)} {int(mo)} *", None
            except Exception:
                pass
        # HH:MM → aujourd'hui
        if len(parts) == 1 and ":" in parts[0]:
            try:
                h, mi = parts[0].split(":")
                now = _dt.datetime.now()
                return f"{int(mi)} {int(h)} {now.day} {now.month} *", None
            except Exception:
                pass
        return None, (
            f"Format 'when' non reconnu : {when_str!r}. "
            "Utilisez 'YYYY-MM-DD HH:MM', 'HH:MM' ou une expression cron '30 14 * * *'."
        )

    # ── list ─────────────────────────────────────────────────────────────────

    if action == "list":
        if is_win:
            r = _subprocess.run(
                ["schtasks", "/query", "/fo", "TABLE"],
                capture_output=True, text=True, timeout=15,
                encoding="utf-8", errors="replace",
            )
            return (r.stdout or r.stderr or "(Aucune tâche planifiée)").strip()
        else:
            out = _crontab_read().strip()
            return out if out else "(Aucune entrée crontab)"

    # ── create ────────────────────────────────────────────────────────────────

    elif action == "create":
        if not name or not command or not when:
            return "[Erreur] 'name', 'command' et 'when' sont requis."

        if is_win:
            parts = when.strip().split()
            if len(parts) == 2:
                try:
                    y, m, d = parts[0].split("-")
                    date_arg = f"{d}/{m}/{y}"
                except Exception:
                    date_arg = parts[0]
                args = ["schtasks", "/create", "/tn", name, "/tr", command,
                        "/sc", "ONCE", "/sd", date_arg, "/st", parts[1], "/f"]
            else:
                args = ["schtasks", "/create", "/tn", name, "/tr", command,
                        "/sc", "ONCE", "/st", when.strip(), "/f"]
            r = _subprocess.run(args, capture_output=True, text=True,
                                timeout=15, encoding="utf-8", errors="replace")
            if r.returncode == 0:
                return f"[OK] Tâche « {name} » planifiée pour {when} : {command}"
            return f"[Erreur schtasks] {(r.stderr or r.stdout).strip()}"

        else:
            cron_expr, err = _when_to_cron(when)
            if err:
                return f"[Erreur] {err}"
            # Remplacer si la tâche existe déjà
            _, new_content = _crontab_remove(name)
            new_content = (new_content.rstrip("\n") + "\n"
                           + f"{_MARKER} {name}\n"
                           + f"{cron_expr} {command}\n")
            ok, msg = _crontab_write(new_content)
            if ok:
                return f"[OK] Tâche « {name} » planifiée ({cron_expr}) : {command}"
            return f"[Erreur crontab] {msg}"

    # ── delete ────────────────────────────────────────────────────────────────

    elif action == "delete":
        if not name:
            return "[Erreur] 'name' est requis."

        if is_win:
            r = _subprocess.run(
                ["schtasks", "/delete", "/tn", name, "/f"],
                capture_output=True, text=True, timeout=15,
                encoding="utf-8", errors="replace",
            )
            if r.returncode == 0:
                return f"[OK] Tâche « {name} » supprimée."
            return f"[Erreur schtasks] {(r.stderr or r.stdout).strip()}"

        else:
            found, new_content = _crontab_remove(name)
            if not found:
                return f"[Erreur] Tâche « {name} » introuvable dans le crontab."
            ok, msg = _crontab_write(new_content)
            if ok:
                return f"[OK] Tâche « {name} » supprimée du crontab."
            return f"[Erreur crontab] {msg}"

    return f"[Erreur] Action inconnue : {action!r}. Valeurs : list, create, delete."


_SCHEDULE_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "schedule_task",
            "description": (
                "Planifie, liste ou supprime des tâches programmées via le planificateur OS "
                "(Planificateur de tâches Windows / crontab Linux-macOS). "
                "Permet d'exécuter automatiquement une commande à une heure donnée. "
                "Exemples : lancer un script de sauvegarde la nuit, "
                "déclencher un traitement d'images à heure précise."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["list", "create", "delete"],
                        "description": "list : lister. create : planifier. delete : supprimer.",
                    },
                    "name": {
                        "type": "string",
                        "description": "Nom unique de la tâche planifiée",
                    },
                    "command": {
                        "type": "string",
                        "description": "Commande à exécuter (chemin absolu recommandé)",
                    },
                    "when": {
                        "type": "string",
                        "description": "Heure d'exécution : 'HH:MM' (aujourd'hui) ou 'YYYY-MM-DD HH:MM'",
                    },
                },
                "required": ["action"],
            },
        },
    },
]


# ── http_request ─────────────────────────────────────────────────────────────

import urllib.request as _urllib_req
import urllib.error as _urllib_err
import json as _json_mod


def _http_request(method, url, headers=None, body=None, timeout=30):
    """Requête HTTP avec méthode, headers et body custom."""
    method = (method or "GET").upper()
    headers = dict(headers or {})
    data = None
    if body:
        if isinstance(body, dict):
            body = _json_mod.dumps(body)
            headers.setdefault("Content-Type", "application/json")
        data = body.encode("utf-8") if isinstance(body, str) else body
    try:
        req = _urllib_req.Request(url, data=data, headers=headers, method=method)
        with _urllib_req.urlopen(req, timeout=int(timeout or 30)) as resp:
            status = resp.status
            raw = resp.read().decode("utf-8", errors="replace")
            ct = resp.headers.get("Content-Type", "")
            if "json" in ct:
                try:
                    raw = _json_mod.dumps(_json_mod.loads(raw), indent=2, ensure_ascii=False)
                except Exception:
                    pass
            truncated = raw[:10000]
            suffix = "\n[…tronqué à 10 000 chars]" if len(raw) > 10000 else ""
            return f"[HTTP {status}] {method} {url}\n{truncated}{suffix}"
    except _urllib_err.HTTPError as exc:
        try:
            err_body = exc.read().decode("utf-8", errors="replace")[:2000]
        except Exception:
            err_body = ""
        return f"[HTTP {exc.code}] {method} {url}\n{err_body}"
    except Exception as exc:
        return f"[Erreur] {method} {url} : {exc}"


_HTTP_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "http_request",
            "description": (
                "Effectue une requête HTTP (GET, POST, PUT, DELETE, PATCH) "
                "avec méthode, headers et body personnalisables. "
                "Utile pour interagir avec des APIs REST, webhooks, services locaux."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "method": {
                        "type": "string",
                        "description": "Méthode HTTP : GET, POST, PUT, DELETE, PATCH",
                    },
                    "url": {
                        "type": "string",
                        "description": "URL cible (https://… ou http://localhost/…)",
                    },
                    "headers": {
                        "type": "object",
                        "description": (
                            "Headers HTTP (dict clé/valeur). "
                            "Ex : {\"Authorization\": \"Bearer token\", "
                            "\"Content-Type\": \"application/json\"}"
                        ),
                    },
                    "body": {
                        "type": "string",
                        "description": (
                            "Corps de la requête (string JSON, form-data, etc.). "
                            "Pour du JSON, passer une chaîne JSON valide."
                        ),
                    },
                    "timeout": {
                        "type": "integer",
                        "description": "Timeout en secondes (défaut : 30)",
                    },
                },
                "required": ["method", "url"],
            },
        },
    },
]


# ── ssh_command ───────────────────────────────────────────────────────────────
# Le mot de passe est fourni par l'appelant (résolu via credentials.py côté
# Dashboard/SidePanel, avec boîte de dialogue si besoin) : cette fonction ne
# le stocke jamais et ne le fait jamais réapparaître dans le texte retourné.

def _ssh_command(host, username, password, command, port=22, timeout=30):
    """Exécute une commande shell sur un hôte distant via SSH."""
    try:
        import paramiko
    except ImportError:
        return "[Erreur] Le module 'paramiko' est requis (pip install paramiko)."

    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        client.connect(
            host, port=int(port or 22), username=username, password=password,
            timeout=int(timeout or 30),
        )
        _stdin, stdout, stderr = client.exec_command(command, timeout=int(timeout or 30))
        out = stdout.read().decode("utf-8", errors="replace")
        err = stderr.read().decode("utf-8", errors="replace")
        exit_code = stdout.channel.recv_exit_status()
    except Exception as exc:
        return f"[Erreur SSH] {username}@{host} : {exc}"
    finally:
        client.close()

    result = out + (f"\n[stderr]\n{err}" if err else "")
    truncated = result[:10000]
    suffix = "\n[…tronqué à 10 000 chars]" if len(result) > 10000 else ""
    return f"[SSH exit={exit_code}] {username}@{host}:{port} $ {command}\n{truncated}{suffix}"


_SSH_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "ssh_command",
            "description": (
                "Exécute une commande shell sur un serveur distant via SSH "
                "(mot de passe résolu via un coffre à identifiants local, "
                "jamais visible du modèle). "
                "Utilise 'name' pour un serveur connu (voir CONSTANTS.SSH_SERVERS) "
                "— sinon précise host/username."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {
                        "type": "string",
                        "description": "Alias d'un serveur connu (CONSTANTS.SSH_SERVERS), remplace host/username",
                    },
                    "host": {
                        "type": "string",
                        "description": "Adresse ou nom d'hôte du serveur (ex : monobjet.example.com)",
                    },
                    "username": {
                        "type": "string",
                        "description": "Nom d'utilisateur SSH",
                    },
                    "command": {
                        "type": "string",
                        "description": "Commande shell à exécuter sur l'hôte distant",
                    },
                    "port": {
                        "type": "integer",
                        "description": "Port SSH (défaut : 22)",
                    },
                    "timeout": {
                        "type": "integer",
                        "description": "Timeout en secondes (défaut : 30)",
                    },
                },
                "required": ["command"],
            },
        },
    },
]


# ── read_spreadsheet ──────────────────────────────────────────────────────────

def _read_spreadsheet(folder_path, filepath, sheet=None, max_rows=100):
    """Lit un fichier tableur (CSV / .xlsx / .xls / .ods) et retourne un tableau texte."""
    resolved = _resolve_path(folder_path, filepath) if folder_path else filepath
    if not resolved or not _os.path.isfile(resolved):
        return f"[Erreur] Fichier introuvable : {filepath}"

    ext = _os.path.splitext(resolved)[1].lower()
    name = _os.path.basename(resolved)
    max_rows = min(int(max_rows or 100), 500)

    def _fmt(sheet_name, headers, rows, total, all_sheets):
        lines = []
        if len(all_sheets) > 1:
            others = ", ".join(s for s in all_sheets if s != sheet_name)
            lines.append(f"Feuille : « {sheet_name} »  (autres : {others})")
        else:
            lines.append(f"Feuille : « {sheet_name} »")
        lines.append(f"Colonnes ({len(headers)}) : {' | '.join(str(h) for h in headers)}")
        lines.append(f"Lignes : {len(rows)} affichées / {total} total")
        lines.append("")
        for row in rows:
            lines.append(" | ".join("" if c is None else str(c) for c in row))
        return "\n".join(lines)

    # ── CSV ──────────────────────────────────────────────────────────────────
    if ext == ".csv":
        import csv as _csv_mod
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                with open(resolved, "r", encoding=enc, newline="") as fh:
                    rows = list(_csv_mod.reader(fh))
                if not rows:
                    return f"[OK] {name} est vide."
                headers = rows[0]
                return _fmt("(CSV)", headers, rows[1:max_rows + 1],
                            max(0, len(rows) - 1), ["(CSV)"])
            except UnicodeDecodeError:
                continue
        return f"[Erreur] Encodage non reconnu pour {name}."

    # ── xlsx / xlsm ──────────────────────────────────────────────────────────
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            import openpyxl as _openpyxl
            wb = _openpyxl.load_workbook(resolved, read_only=True, data_only=True)
            all_sheets = wb.sheetnames
            ws = wb[sheet] if sheet and sheet in all_sheets else wb.active
            sname = ws.title
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
                if len(rows) > max_rows + 1:
                    break
            wb.close()
            if not rows:
                return f"[OK] Feuille « {sname} » vide."
            headers = [f"Col{i+1}" if c is None else str(c)
                       for i, c in enumerate(rows[0])]
            total = (ws.max_row or len(rows)) - 1
            return _fmt(sname, headers, rows[1:], total, all_sheets)
        except ImportError:
            return f"[Erreur] Module manquant pour {ext}. Installez : pip install openpyxl"

    # ── xls (ancien format) ───────────────────────────────────────────────────
    if ext == ".xls":
        try:
            import xlrd as _xlrd
            wb = _xlrd.open_workbook(resolved)
            all_sheets = wb.sheet_names()
            ws = wb.sheet_by_name(sheet) if sheet and sheet in all_sheets \
                else wb.sheet_by_index(0)
            rows = [[ws.cell_value(r, c) for c in range(ws.ncols)]
                    for r in range(min(ws.nrows, max_rows + 1))]
            if not rows:
                return f"[OK] Feuille « {ws.name} » vide."
            return _fmt(ws.name, [str(h) for h in rows[0]],
                        rows[1:], ws.nrows - 1, all_sheets)
        except ImportError:
            return f"[Erreur] Module manquant pour {ext}. Installez : pip install xlrd"

    # ── ods / fods ────────────────────────────────────────────────────────────
    if ext in (".ods", ".fods"):
        try:
            import odf.opendocument as _odfdoc
            import odf.table as _odftable
            import odf.text as _odftext

            doc = _odfdoc.load(resolved)
            sheets = doc.spreadsheet.getElementsByType(_odftable.Table)
            all_sheets = [s.getAttribute("name") for s in sheets]
            ws = next(
                (s for s in sheets if s.getAttribute("name") == sheet), None
            ) if sheet else (sheets[0] if sheets else None)
            if ws is None:
                return f"[Erreur] Aucune feuille trouvée dans {name}."
            sname = ws.getAttribute("name")
            rows = []
            for tr in ws.getElementsByType(_odftable.TableRow):
                cells = []
                for cell in tr.getElementsByType(_odftable.TableCell):
                    txt = " ".join(
                        (p.firstChild.data if p.firstChild else "")
                        for p in cell.getElementsByType(_odftext.P)
                    )
                    rep = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                    cells.extend([txt] * rep)
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    rows.append(cells)
                if len(rows) > max_rows + 1:
                    break
            if not rows:
                return f"[OK] Feuille « {sname} » vide."
            return _fmt(sname, rows[0], rows[1:], len(rows) - 1, all_sheets)
        except ImportError:
            return f"[Erreur] Module manquant pour {ext}. Installez : pip install odfpy"

    return (f"[Erreur] Format non supporté : {ext}. "
            f"Formats acceptés : .csv, .xlsx, .xlsm, .xls, .ods, .fods")


_SPREADSHEET_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_spreadsheet",
            "description": (
                "Lit un fichier tableur (CSV, .xlsx, .xls, .ods) et retourne "
                "les noms de colonnes et les premières lignes sous forme de tableau texte. "
                "Utile pour analyser des données, répondre à des questions sur un fichier Excel "
                "client, ou extraire des informations structurées."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Chemin du fichier (relatif au dossier ouvert ou absolu)",
                    },
                    "sheet": {
                        "type": "string",
                        "description": (
                            "Nom de la feuille à lire (optionnel — défaut : première feuille). "
                            "Les autres feuilles disponibles sont listées dans la réponse."
                        ),
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "Nombre de lignes à retourner (défaut : 100, max : 500)",
                    },
                },
                "required": ["filepath"],
            },
        },
    },
]
